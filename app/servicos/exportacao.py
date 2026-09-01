"""Tela 4 (Etapa 6) — exportar em Excel a lista de pedidos por fornecedor,
pronta para o comprador enviar por e-mail.

⚠ Mesma regra de compra.py: quantidade e valor do pedido se leem de
APP_DECISAO_PEDIDO ao vivo, com o FATOR_EXIBICAO CONGELADO na decisao
(MELHORIA A5). Exportar de COMPRAS_PEDIDO traria o valor do ultimo `dbt run`
- zero, enquanto a decisao ainda nao foi incorporada ao proximo build
(CONTEXTO.md §2). So entram linhas com PEDIDO > 0: pedido zero e decisao de
"nao comprar", e mandar isso ao fornecedor e ruido.

⚠ ESTE ARQUIVO SAI DA EMPRESA — decisao do usuario em 2026-08-19. So entram
as 9 colunas combinadas com o Diretor de Compras (codigo, cod. fabricante,
descricao, embalagem, qtd. pedida, unidade, qtd. em unidades, valor unitario
de referencia, valor de referencia). O valor de referencia usa VL_ENT_UNIT
(BH, valor da ultima entrada - o que o fornecedor ja pratica), NUNCA
CUSTO_TOT_OFICIAL (BW, custo carregado com ICMS-ST/creditos/ajustes
internos): mandar isso para o fornecedor entrega a estrutura tributaria e de
custo da CEDEP para quem esta do outro lado da negociacao. Nada de margem,
preco de venda, curva ABC, alerta, estoque, media de venda, credito, MVA ou
aliquota entra aqui, por nenhum motivo.
"""
from __future__ import annotations

import datetime as dt
import io
import re
import unicodedata

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.core import database

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
_AVISO = "Valores de referência, baseados na última entrada — não constituem preço acordado."

# fim exclusivo: filtra ATE 23:59:59.999 do dia informado, sem depender de
# como o Oracle trunca hora em comparacao de DATE.
def _janela(data_de: dt.date, data_ate: dt.date) -> tuple[dt.datetime, dt.datetime]:
    inicio = dt.datetime.combine(data_de, dt.time.min)
    fim = dt.datetime.combine(data_ate + dt.timedelta(days=1), dt.time.min)
    return inicio, fim


def fornecedores_com_pedido(data_de: dt.date, data_ate: dt.date) -> list[dict]:
    """Fornecedores com ao menos um PEDIDO > 0 gravado em APP_DECISAO_PEDIDO
    na janela, com contagem de SKU, total de unidades e valor de referencia
    (VL_ENT_UNIT, nunca CUSTO_TOT_OFICIAL - ver cabecalho do arquivo)."""
    inicio, fim = _janela(data_de, data_ate)
    return database.consultar(
        """
        select p.fornecedor,
               count(distinct p.codigo)                        as qtd_sku,
               sum(d.pedido * d.fator_exibicao)                 as total_unidades,
               sum(d.pedido * d.fator_exibicao * p.vl_ent_unit) as valor_total
          from compras_pedido p
          join app_decisao_pedido d on d.id_produto = p.codigo
         where d.pedido > 0
           and d.atualizado_em >= :inicio and d.atualizado_em < :fim
         group by p.fornecedor
         order by p.fornecedor
        """,
        {"inicio": inicio, "fim": fim},
    )


def itens_do_fornecedor(fornecedor: str, data_de: dt.date, data_ate: dt.date) -> list[dict]:
    """Linhas do pedido de um fornecedor na janela - so PEDIDO > 0. Fator e
    valor unitario vem CONGELADOS na propria decisao, nunca recalculados."""
    inicio, fim = _janela(data_de, data_ate)
    return database.consultar(
        """
        select p.codigo, p.cod_fab, p.descricao, p.embalagem, p.embal_compra,
               p.comprador,
               d.pedido                as pedido,
               d.fator_exibicao        as fator_exibicao,
               p.vl_ent_unit           as vl_ent_unit,
               d.atualizado_em         as atualizado_em
          from compras_pedido p
          join app_decisao_pedido d on d.id_produto = p.codigo
         where p.fornecedor = :fornecedor
           and d.pedido > 0
           and d.atualizado_em >= :inicio and d.atualizado_em < :fim
         order by p.codigo
        """,
        {"fornecedor": fornecedor, "inicio": inicio, "fim": fim},
    )


def _slug(texto: str) -> str:
    """Nome de fornecedor sem acento/espaco, para nome de arquivo."""
    sem_acento = unicodedata.normalize("NFKD", texto or "").encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^A-Za-z0-9]+", "_", sem_acento).strip("_").upper() or "FORNECEDOR"


def nome_arquivo(fornecedor: str) -> str:
    return f"pedido_{_slug(fornecedor)}_{dt.date.today():%Y-%m-%d}.xlsx"


def _br(valor: float, casas: int = 0) -> str:
    """1234.5 -> '1.234' / '1.234,50' - separador PT-BR so para o texto do
    cabecalho (as celulas da tabela usam number_format, nao string)."""
    return f"{valor:,.{casas}f}".replace(",", "§").replace(".", ",").replace("§", ".")


_COLUNAS = [
    ("CÓDIGO", "0"),
    ("CÓD. FABRICANTE", "@"),
    ("DESCRIÇÃO", "@"),
    ("EMBALAGEM", "@"),
    ("QTD. PEDIDA", "#,##0.##"),
    ("UNIDADE", "@"),
    ("QTD. EM UNIDADES", "#,##0.##"),
    ("VL. UNIT. REFERÊNCIA", "#,##0.0000"),
    ("VALOR DE REFERÊNCIA", "#,##0.00"),
]


def gerar_xlsx(fornecedor: str, itens: list[dict], cabecalho: dict) -> bytes:
    """Monta o .xlsx que sai para o fornecedor. `cabecalho` traz comprador,
    periodo, quem/quando gerou e os totais - tudo isso vira linhas de texto
    ACIMA da tabela, nunca coluna (para nao confundir Excel do fornecedor que
    filtra tudo)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Pedido"

    linhas_cabecalho = [
        (f"PEDIDO — {fornecedor}", True),
        (f"Comprador: {cabecalho.get('comprador') or '-'}", False),
        (f"Período: {cabecalho['data_de']:%d/%m/%Y} a {cabecalho['data_ate']:%d/%m/%Y}", False),
        (f"Gerado em {cabecalho['gerado_em']:%d/%m/%Y %H:%M} por {cabecalho.get('gerado_por') or '-'}", False),
        (
            f"{cabecalho['qtd_sku']} SKU(s) — {_br(cabecalho['total_unidades'])} unidade(s)"
            f" — R$ {_br(cabecalho['valor_total'], 2)} de referência",
            False,
        ),
        (_AVISO, False),
    ]
    for texto, destaque in linhas_cabecalho:
        ws.append([texto])
        cel = ws.cell(row=ws.max_row, column=1)
        cel.font = Font(bold=destaque, size=13 if destaque else 10, italic=not destaque and texto == _AVISO)
    ws.append([])

    linha_titulo = ws.max_row + 1
    ws.append([rotulo for rotulo, _ in _COLUNAS])
    for idx in range(1, len(_COLUNAS) + 1):
        cel = ws.cell(row=linha_titulo, column=idx)
        cel.fill = _HEADER_FILL
        cel.font = _HEADER_FONT
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    larguras = [max(len(rotulo) + 2, 12) for rotulo, _ in _COLUNAS]

    for item in itens:
        fator = item["fator_exibicao"] or 1
        pedido = item["pedido"] or 0
        vl_unit = item["vl_ent_unit"] or 0
        qtd_unidades = pedido * fator
        valor_ref = qtd_unidades * vl_unit
        embal_compra = item["embal_compra"]
        if fator and fator > 1 and embal_compra is not None:
            n = float(embal_compra)
            unidade = f"CAIXA C/ {int(n) if n.is_integer() else n}"
        else:
            unidade = "UNIDADE"

        valores = [
            item["codigo"], item["cod_fab"], item["descricao"], item["embalagem"],
            pedido, unidade, qtd_unidades, vl_unit, valor_ref,
        ]
        ws.append(valores)
        for i, valor in enumerate(valores):
            if valor is not None:
                larguras[i] = min(max(larguras[i], len(str(valor)) + 2), 45)

    ultima = ws.max_row
    for idx, (_, fmt) in enumerate(_COLUNAS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = larguras[idx - 1]
        if fmt == "@":
            continue
        for row in range(linha_titulo + 1, ultima + 1):
            ws.cell(row=row, column=idx).number_format = fmt

    ws.freeze_panes = f"A{linha_titulo + 1}"
    ws.auto_filter.ref = f"A{linha_titulo}:{get_column_letter(len(_COLUNAS))}{max(ultima, linha_titulo)}"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
