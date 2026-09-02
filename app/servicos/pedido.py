"""Etapa 9 — o fluxo de pedido de ponta a ponta: carrinho -> APP_PEDIDO(_ITEM),
lista, edição, máquina de estados, exportações (`sql/03_tabelas_pedido.sql`,
`v2/prototipo/PROTOTIPO.md` §2.4-2.8).

Substitui, para as telas de "Pedidos Salvos", o modelo antigo de
`app/servicos/compra.py` (APP_DECISAO_PEDIDO, PK em ID_PRODUTO, uma linha por
SKU sem agrupamento). Este módulo não toca em APP_DECISAO_PEDIDO nem em
APP_DECISAO_PRECO — schemas diferentes, tabelas diferentes.

── FATOR_EXIBICAO congelado (MELHORIA A5) ──────────────────────────────────
Toda gravação em APP_PEDIDO_ITEM (criação do carrinho, edição de item)
congela o FATOR_EXIBICAO **vigente no instante da gravação**, lido de
COMPRAS_PEDIDO. Quantidade em unidades reais = quantidade x fator_exibicao,
sempre lidos da MESMA linha do item — nunca do fator corrente do cadastro.
Isso vale inclusive quando o item é editado bem depois de criado: o fator é
re-congelado a cada toque na linha (é o mesmo raciocínio de
`APP_DECISAO_PEDIDO.fator_exibicao`, só que por item de pedido em vez de por
SKU solto).

── Origem do preço unitário quando não informado ───────────────────────────
`sql/03_tabelas_pedido.sql` (comentário de APP_PEDIDO_ITEM.preco_unitario) e
`PROTOTIPO.md §3.7` são explícitos: o preço nasce de `custoGerencial`
(COMPRAS_PEDIDO.CUSTO_TOT_GERENCIAL) no momento em que o produto entra no
pedido — não de VL_ENT_UNIT. VL_ENT_UNIT é a escolha da exportação para
FORNECEDOR (`exportacao.py`, arquivo que sai da empresa, para não revelar
custo carregado de imposto); aqui o preço é a decisão interna do comprador
sobre o próprio custo gerencial, editável em tela depois. Duas perguntas
diferentes, duas origens diferentes.

── Concorrência ─────────────────────────────────────────────────────────────
Duas pessoas editando o mesmo pedido ao mesmo tempo: toda operação que muda
um pedido (item, status, exclusão, exportação Winthor) faz
`select status ... for update` na linha de APP_PEDIDO antes de escrever. Isso
serializa as transações concorrentes no MESMO pedido (a segunda espera a
primeira commitar e enxerga o estado já atualizado) sem precisar de coluna de
versão/otimistic lock — não há tela colaborativa em tempo real no protótipo
que justifique o custo de um mecanismo mais fino, e "a segunda gravação vence
e reflete o que a primeira gravou" é um comportamento seguro, só não é
"avisa que alguém mexeu enquanto você editava". Decisão registrada aqui para
quem revisar depois: se a tela precisar desse aviso, é UPDATE ... WHERE
ATUALIZADO_EM = :versao_lida, e não está implementado.

Nomes em snake_case (vocabulário do banco); a tradução para camelCase mora em
`app/api/contrato.py`.
"""
from __future__ import annotations

import datetime as dt
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from app.core import auditoria, database
from app.servicos import exportacao

# ─────────────────────────────────────────────────────────────────────────────
# Máquina de estados — sql/03_tabelas_pedido.sql: o CHECK só garante o domínio
# de 4 valores; a ORDEM da transição é responsabilidade da aplicação.
# ─────────────────────────────────────────────────────────────────────────────
ORDEM_STATUS = ["Rascunho", "Orçamento Enviado", "Fechado", "Exportado"]
STATUS_EDITAVEIS = {"Rascunho", "Orçamento Enviado"}


class PedidoNaoEncontrado(Exception):
    """id_pedido não existe em APP_PEDIDO. A rota converte para 404."""


class TransicaoInvalida(Exception):
    """Pulo de etapa, volta de dois passos, ou exportar fora de Fechado. 409."""


class EdicaoNaoPermitida(Exception):
    """Editar item de pedido Fechado/Exportado. 409."""


class ProdutoInvalido(Exception):
    """Código não existe em COMPRAS_PEDIDO, é de outro fornecedor, ou não tem
    dado suficiente (fator/custo) para decidir quantidade/preço sem que o
    comprador informe manualmente. 422."""


# ─────────────────────────────────────────────────────────── leitura ───────

_AGG = """
      left join (
            select id_pedido, count(*) as qtd_itens,
                   sum(quantidade * fator_exibicao * preco_unitario) as valor_total
              from app_pedido_item
             group by id_pedido
           ) agg on agg.id_pedido = p.id_pedido
"""

_COLUNAS_CABECALHO = """
    p.id_pedido, p.fornecedor, p.status, p.criado_em, p.criado_por,
    p.atualizado_em, p.atualizado_por,
    nvl(agg.qtd_itens, 0) as qtd_itens,
    nvl(agg.valor_total, 0) as valor_total
"""


def listar_pedidos(filtros: dict, pagina: int, itens_por_pagina: int) -> tuple[list[dict], int]:
    """qtd_itens e valor_total somados no BANCO (sum/count sobre
    APP_PEDIDO_ITEM), não em Python — é o que a tela de lista mostra por
    pedido, PROTOTIPO.md §2.5."""
    condicoes: list[str] = []
    binds: dict = {}

    status_lista = filtros.get("status") or []
    if status_lista:
        marcas = ", ".join(f":st{i}" for i in range(len(status_lista)))
        binds.update({f"st{i}": s for i, s in enumerate(status_lista)})
        condicoes.append(f"p.status in ({marcas})")

    if filtros.get("fornecedor"):
        condicoes.append("p.fornecedor = :fornecedor")
        binds["fornecedor"] = filtros["fornecedor"]

    if filtros.get("busca"):
        condicoes.append(
            "(upper(p.fornecedor) like upper(:busca) or to_char(p.id_pedido) like :busca)"
        )
        binds["busca"] = f"%{filtros['busca'].strip()}%"

    onde = f"where {' and '.join(condicoes)}" if condicoes else ""

    total_linha = database.consultar_um(f"select count(*) as n from app_pedido p {onde}", binds)
    total = int(total_linha["n"]) if total_linha else 0

    pagina = max(1, pagina)
    linhas = database.consultar(
        f"""
        select {_COLUNAS_CABECALHO}
          from app_pedido p
          {_AGG}
          {onde}
         order by p.criado_em desc, p.id_pedido desc
        offset :offset rows fetch next :limite rows only
        """,
        {**binds, "offset": (pagina - 1) * itens_por_pagina, "limite": itens_por_pagina},
    )
    return linhas, total


def obter_pedido(id_pedido: int) -> dict | None:
    """Só o cabeçalho, com os agregados — mesma consulta de `listar_pedidos`
    para uma linha só, para o fragmento pós-gravação nunca divergir da lista."""
    return database.consultar_um(
        f"select {_COLUNAS_CABECALHO} from app_pedido p {_AGG} where p.id_pedido = :id_pedido",
        {"id_pedido": id_pedido},
    )


def obter_itens(id_pedido: int) -> list[dict]:
    """Itens com descrição/cód. fabricante/embalagem vindos de COMPRAS_PEDIDO
    por LEFT JOIN (nunca INNER: sql/03_tabelas_pedido.sql é explícito — o
    produto pode ter saído do catálogo depois que o pedido foi fechado, e o
    item precisa sobreviver como registro histórico)."""
    return database.consultar(
        """
        select i.id_pedido, i.id_produto as codigo, i.quantidade, i.fator_exibicao,
               i.preco_unitario, i.criado_em,
               p.descricao, p.cod_fab, p.embalagem, p.embal_compra
          from app_pedido_item i
          left join compras_pedido p on p.codigo = i.id_produto
         where i.id_pedido = :id_pedido
         order by i.id_produto
        """,
        {"id_pedido": id_pedido},
    )


def obter_detalhe(id_pedido: int) -> dict | None:
    cabecalho = obter_pedido(id_pedido)
    if cabecalho is None:
        return None
    cabecalho = dict(cabecalho)
    cabecalho["itens"] = obter_itens(id_pedido)
    return cabecalho


# ─────────────────────────────────────────────────────── salvar carrinho ───

def salvar_carrinho(
    itens_entrada: list[dict], usuario_login: str, usuario_id: int, ip: str | None = None
) -> list[dict]:
    """`itens_entrada`: [{"codigo", "quantidade", "preco_unitario"?}]. Cria UM
    APP_PEDIDO por FORNECEDOR (COMPRAS_PEDIDO.FORNECEDOR) presente no
    carrinho — exigência do formato de importação Winthor rotina 220
    (PROTOTIPO.md §2.4/§5). Tudo numa transação só: ou todos os pedidos do
    carrinho são criados, ou nenhum é.
    """
    normalizados: dict[int, dict] = {}
    for it in itens_entrada:
        quantidade = float(it.get("quantidade") or 0)
        if quantidade <= 0:
            continue  # nada a salvar dessa linha, mesmo raciocínio de "zero remove"
        codigo = int(it["codigo"])
        preco = it.get("preco_unitario")
        # última ocorrência do código vence, se o carrinho tiver duplicata
        normalizados[codigo] = {
            "codigo": codigo,
            "quantidade": quantidade,
            "preco_unitario": float(preco) if preco is not None else None,
        }
    if not normalizados:
        raise ProdutoInvalido("Nenhum item com quantidade maior que zero para salvar.")

    codigos = list(normalizados.keys())
    marcas = ", ".join(f":c{i}" for i in range(len(codigos)))
    binds = {f"c{i}": c for i, c in enumerate(codigos)}
    catalogo = database.consultar(
        f"select codigo, fornecedor, fator_exibicao, custo_tot_gerencial"
        f" from compras_pedido where codigo in ({marcas})",
        binds,
    )
    por_produto = {int(r["codigo"]): r for r in catalogo}
    faltando = [c for c in codigos if c not in por_produto]
    if faltando:
        raise ProdutoInvalido(
            "Produto(s) não encontrado(s) em COMPRAS_PEDIDO: "
            + ", ".join(str(c) for c in faltando)
        )

    grupos: dict[str, list[dict]] = {}
    for it in normalizados.values():
        prod = por_produto[it["codigo"]]
        fator = prod["fator_exibicao"]
        if fator is None or float(fator) <= 0:
            raise ProdutoInvalido(f"Produto {it['codigo']} sem FATOR_EXIBICAO válido em COMPRAS_PEDIDO.")
        preco = it["preco_unitario"]
        if preco is None:
            preco = prod["custo_tot_gerencial"]
        if preco is None or float(preco) <= 0:
            raise ProdutoInvalido(
                f"Produto {it['codigo']} sem custo gerencial calculado em COMPRAS_PEDIDO;"
                " informe precoUnitario manualmente."
            )
        grupos.setdefault(prod["fornecedor"], []).append({
            "codigo": it["codigo"],
            "quantidade": it["quantidade"],
            "fator_exibicao": float(fator),
            "preco_unitario": float(preco),
        })

    agora = dt.datetime.now()
    ids_criados: list[int] = []
    with database.transacao() as conn:
        for fornecedor, itens_grupo in grupos.items():
            cur = conn.cursor()
            id_var = cur.var(int)
            cur.execute(
                """
                insert into app_pedido (fornecedor, status, criado_em, criado_por, atualizado_em, atualizado_por)
                values (:fornecedor, 'Rascunho', :agora, :usuario, :agora, :usuario)
                returning id_pedido into :id
                """,
                {"fornecedor": fornecedor, "agora": agora, "usuario": usuario_login, "id": id_var},
            )
            id_pedido = int(id_var.getvalue()[0])

            cur.execute(
                """
                insert into app_pedido_status_hist
                    (id_pedido, status_anterior, status_novo, alterado_em, alterado_por)
                values (:id_pedido, null, 'Rascunho', :agora, :usuario)
                """,
                {"id_pedido": id_pedido, "agora": agora, "usuario": usuario_login},
            )

            for it in itens_grupo:
                cur.execute(
                    """
                    insert into app_pedido_item
                        (id_pedido, id_produto, quantidade, fator_exibicao, preco_unitario, criado_em)
                    values (:id_pedido, :codigo, :quantidade, :fator_exibicao, :preco_unitario, :agora)
                    """,
                    {
                        "id_pedido": id_pedido, "codigo": it["codigo"], "quantidade": it["quantidade"],
                        "fator_exibicao": it["fator_exibicao"], "preco_unitario": it["preco_unitario"],
                        "agora": agora,
                    },
                )
            cur.close()

            auditoria.registrar(
                conn, usuario_id, "CRIAR_PEDIDO", "APP_PEDIDO", str(id_pedido),
                {
                    "fornecedor": fornecedor, "qtd_itens": len(itens_grupo),
                    "itens": [{"codigo": it["codigo"], "quantidade": it["quantidade"]} for it in itens_grupo],
                    "usuario": usuario_login,
                },
                ip,
            )
            ids_criados.append(id_pedido)

    return [obter_pedido(i) for i in ids_criados]


# ─────────────────────────────────────────────────────── edição de item ───

def _bloquear_pedido(cur, id_pedido: int) -> dict:
    """Lê e trava (FOR UPDATE) o cabeçalho — serializa gravações concorrentes
    no mesmo pedido (ver docstring do módulo, seção Concorrência)."""
    cur.execute(
        "select status, fornecedor from app_pedido where id_pedido = :id for update",
        {"id": id_pedido},
    )
    linha = cur.fetchone()
    if linha is None:
        raise PedidoNaoEncontrado(f"Pedido {id_pedido} não encontrado.")
    return {"status": linha[0], "fornecedor": linha[1]}


def _exigir_editavel(status_atual: str) -> None:
    if status_atual not in STATUS_EDITAVEIS:
        raise EdicaoNaoPermitida(
            f"Pedido está em '{status_atual}'; só é possível editar em Rascunho"
            " ou Orçamento Enviado."
        )


def upsert_item(
    id_pedido: int, codigo: int, quantidade: float, preco_unitario: float | None,
    usuario_login: str, usuario_id: int, ip: str | None = None,
) -> dict:
    """Acrescenta ou atualiza um item. `quantidade <= 0` remove a linha —
    mesmo comportamento do protótipo (onBlur com campo zerado, §2.6)."""
    if quantidade <= 0:
        return remover_item(id_pedido, codigo, usuario_login, usuario_id, ip)

    agora = dt.datetime.now()
    with database.transacao() as conn:
        cur = conn.cursor()
        info = _bloquear_pedido(cur, id_pedido)
        _exigir_editavel(info["status"])

        cur.execute(
            "select fornecedor, fator_exibicao, custo_tot_gerencial"
            " from compras_pedido where codigo = :codigo",
            {"codigo": codigo},
        )
        prod = cur.fetchone()
        if prod is None:
            raise ProdutoInvalido(f"Produto {codigo} não encontrado em COMPRAS_PEDIDO.")
        fornecedor_produto, fator_exibicao, custo_gerencial = prod
        if fornecedor_produto != info["fornecedor"]:
            raise ProdutoInvalido(
                f"Produto {codigo} é do departamento '{fornecedor_produto}', diferente do"
                f" fornecedor deste pedido ('{info['fornecedor']}')."
            )
        if fator_exibicao is None or float(fator_exibicao) <= 0:
            raise ProdutoInvalido(f"Produto {codigo} sem FATOR_EXIBICAO válido em COMPRAS_PEDIDO.")

        preco = preco_unitario
        if preco is None:
            cur.execute(
                "select preco_unitario from app_pedido_item where id_pedido = :id and id_produto = :codigo",
                {"id": id_pedido, "codigo": codigo},
            )
            existente = cur.fetchone()
            preco = float(existente[0]) if existente is not None else custo_gerencial
        if preco is None or float(preco) <= 0:
            raise ProdutoInvalido(
                f"Produto {codigo} sem custo gerencial calculado em COMPRAS_PEDIDO;"
                " informe precoUnitario manualmente."
            )

        cur.execute(
            """
            merge into app_pedido_item t
            using (select :id_pedido as id_pedido, :codigo as id_produto from dual) s
               on (t.id_pedido = s.id_pedido and t.id_produto = s.id_produto)
             when matched then update set
                  quantidade = :quantidade, fator_exibicao = :fator_exibicao,
                  preco_unitario = :preco_unitario, criado_em = :agora
             when not matched then insert
                  (id_pedido, id_produto, quantidade, fator_exibicao, preco_unitario, criado_em)
                  values (:id_pedido, :codigo, :quantidade, :fator_exibicao, :preco_unitario, :agora)
            """,
            {
                "id_pedido": id_pedido, "codigo": codigo, "quantidade": quantidade,
                "fator_exibicao": float(fator_exibicao), "preco_unitario": float(preco), "agora": agora,
            },
        )
        cur.execute(
            "update app_pedido set atualizado_em = :agora, atualizado_por = :usuario where id_pedido = :id",
            {"agora": agora, "usuario": usuario_login, "id": id_pedido},
        )
        cur.close()

        auditoria.registrar(
            conn, usuario_id, "GRAVAR_ITEM_PEDIDO", "APP_PEDIDO_ITEM", f"{id_pedido}/{codigo}",
            {
                "quantidade": quantidade, "fator_exibicao": float(fator_exibicao),
                "preco_unitario": float(preco), "usuario": usuario_login,
            },
            ip,
        )

    return obter_detalhe(id_pedido)


def remover_item(
    id_pedido: int, codigo: int, usuario_login: str, usuario_id: int, ip: str | None = None
) -> dict:
    agora = dt.datetime.now()
    with database.transacao() as conn:
        cur = conn.cursor()
        info = _bloquear_pedido(cur, id_pedido)
        _exigir_editavel(info["status"])
        cur.execute(
            "delete from app_pedido_item where id_pedido = :id and id_produto = :codigo",
            {"id": id_pedido, "codigo": codigo},
        )
        removeu = cur.rowcount > 0
        cur.execute(
            "update app_pedido set atualizado_em = :agora, atualizado_por = :usuario where id_pedido = :id",
            {"agora": agora, "usuario": usuario_login, "id": id_pedido},
        )
        cur.close()
        if removeu:
            auditoria.registrar(
                conn, usuario_id, "REMOVER_ITEM_PEDIDO", "APP_PEDIDO_ITEM", f"{id_pedido}/{codigo}",
                {"usuario": usuario_login}, ip,
            )

    detalhe = obter_detalhe(id_pedido)
    if detalhe is None:
        raise PedidoNaoEncontrado(f"Pedido {id_pedido} não encontrado.")
    return detalhe


# ────────────────────────────────────────────────────── máquina de estados ─

def _transicionar(
    id_pedido: int, passo: int, acao: str, usuario_login: str, usuario_id: int, ip: str | None
) -> dict:
    agora = dt.datetime.now()
    with database.transacao() as conn:
        cur = conn.cursor()
        info = _bloquear_pedido(cur, id_pedido)
        idx = ORDEM_STATUS.index(info["status"])
        novo_idx = idx + passo
        if novo_idx < 0 or novo_idx >= len(ORDEM_STATUS):
            if passo < 0:
                raise TransicaoInvalida(
                    f"Pedido já está em '{info['status']}' (etapa inicial); não há como desfazer."
                )
            raise TransicaoInvalida(
                f"Pedido já está em '{info['status']}' (última etapa); não há próxima etapa."
            )
        novo_status = ORDEM_STATUS[novo_idx]
        cur.execute(
            "update app_pedido set status = :novo, atualizado_em = :agora, atualizado_por = :usuario"
            " where id_pedido = :id",
            {"novo": novo_status, "agora": agora, "usuario": usuario_login, "id": id_pedido},
        )
        cur.execute(
            """
            insert into app_pedido_status_hist (id_pedido, status_anterior, status_novo, alterado_em, alterado_por)
            values (:id, :anterior, :novo, :agora, :usuario)
            """,
            {"id": id_pedido, "anterior": info["status"], "novo": novo_status, "agora": agora, "usuario": usuario_login},
        )
        cur.close()
        auditoria.registrar(
            conn, usuario_id, acao, "APP_PEDIDO", str(id_pedido),
            {"de": info["status"], "para": novo_status, "usuario": usuario_login}, ip,
        )
    return obter_detalhe(id_pedido)


def avancar_status(id_pedido: int, usuario_login: str, usuario_id: int, ip: str | None = None) -> dict:
    return _transicionar(id_pedido, +1, "AVANCAR_STATUS_PEDIDO", usuario_login, usuario_id, ip)


def voltar_status(id_pedido: int, usuario_login: str, usuario_id: int, ip: str | None = None) -> dict:
    return _transicionar(id_pedido, -1, "VOLTAR_STATUS_PEDIDO", usuario_login, usuario_id, ip)


def excluir_pedido(id_pedido: int, usuario_login: str, usuario_id: int, ip: str | None = None) -> None:
    """Sem confirmação — mesmo comportamento do protótipo (§2.5). ON DELETE
    CASCADE derruba itens e histórico junto (sql/03_tabelas_pedido.sql)."""
    with database.transacao() as conn:
        cur = conn.cursor()
        info = _bloquear_pedido(cur, id_pedido)
        cur.execute("delete from app_pedido where id_pedido = :id", {"id": id_pedido})
        cur.close()
        auditoria.registrar(
            conn, usuario_id, "EXCLUIR_PEDIDO", "APP_PEDIDO", str(id_pedido),
            {"fornecedor": info["fornecedor"], "status": info["status"], "usuario": usuario_login}, ip,
        )


# ──────────────────────────────────────────────────────────── exportações ──

_COLUNAS_PEDIDO = [
    ("CÓDIGO", "0"),
    ("CÓD. FABRICANTE", "@"),
    ("DESCRIÇÃO", "@"),
    ("EMBALAGEM", "@"),
    ("QUANTIDADE", "#,##0.####"),
    ("UNIDADE", "@"),
    ("QTD. EM UNIDADES", "#,##0.####"),
    ("PREÇO UNITÁRIO", "#,##0.0000"),
    ("VALOR TOTAL", "#,##0.00"),
]


def _montar_xlsx_legivel(cabecalho: dict, itens: list[dict]) -> bytes:
    """Excel do pedido — colunas legíveis, mesmo estilo visual de
    `exportacao.py` (reaproveitado daqui: _HEADER_FILL/_HEADER_FONT), mas
    documento INTERNO (mostra o preço decidido pelo comprador, não o
    VL_ENT_UNIT do arquivo que sai para o fornecedor)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Pedido"

    linhas_cabecalho = [
        (f"PEDIDO #{cabecalho['id_pedido']} — {cabecalho['fornecedor']}", True),
        (f"Status: {cabecalho['status']}", False),
        (
            f"Criado em {cabecalho['criado_em']:%d/%m/%Y %H:%M} por {cabecalho.get('criado_por') or '-'}",
            False,
        ),
        (
            f"Atualizado em {cabecalho['atualizado_em']:%d/%m/%Y %H:%M} por"
            f" {cabecalho.get('atualizado_por') or '-'}",
            False,
        ),
    ]
    for texto, destaque in linhas_cabecalho:
        ws.append([texto])
        cel = ws.cell(row=ws.max_row, column=1)
        cel.font = Font(bold=destaque, size=13 if destaque else 10)
    ws.append([])

    linha_titulo = ws.max_row + 1
    ws.append([rotulo for rotulo, _ in _COLUNAS_PEDIDO])
    for idx in range(1, len(_COLUNAS_PEDIDO) + 1):
        cel = ws.cell(row=linha_titulo, column=idx)
        cel.fill = exportacao._HEADER_FILL
        cel.font = exportacao._HEADER_FONT
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    larguras = [max(len(rotulo) + 2, 12) for rotulo, _ in _COLUNAS_PEDIDO]

    for item in itens:
        fator = float(item["fator_exibicao"] or 1)
        quantidade = float(item["quantidade"] or 0)
        preco = float(item["preco_unitario"] or 0)
        qtd_unidades = quantidade * fator
        valor_total = qtd_unidades * preco
        embal_compra = item.get("embal_compra")
        if fator > 1 and embal_compra is not None:
            n = float(embal_compra)
            unidade = f"CAIXA C/ {int(n) if n.is_integer() else n}"
        else:
            unidade = "UNIDADE"

        valores = [
            item["codigo"], item.get("cod_fab"), item.get("descricao"), item.get("embalagem"),
            quantidade, unidade, qtd_unidades, preco, valor_total,
        ]
        ws.append(valores)
        for i, valor in enumerate(valores):
            if valor is not None:
                larguras[i] = min(max(larguras[i], len(str(valor)) + 2), 45)

    ultima = ws.max_row
    for idx, (_, fmt) in enumerate(_COLUNAS_PEDIDO, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = larguras[idx - 1]
        if fmt == "@":
            continue
        for row in range(linha_titulo + 1, ultima + 1):
            ws.cell(row=row, column=idx).number_format = fmt

    ws.freeze_panes = f"A{linha_titulo + 1}"
    ws.auto_filter.ref = f"A{linha_titulo}:{get_column_letter(len(_COLUNAS_PEDIDO))}{max(ultima, linha_titulo)}"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def gerar_excel_pedido(id_pedido: int) -> tuple[bytes, str] | None:
    cabecalho = obter_pedido(id_pedido)
    if cabecalho is None:
        return None
    itens = obter_itens(id_pedido)
    conteudo = _montar_xlsx_legivel(cabecalho, itens)
    nome = f"pedido_{id_pedido}_{exportacao._slug(cabecalho['fornecedor'])}_{dt.date.today():%Y-%m-%d}.xlsx"
    return conteudo, nome


def _gerar_xlsx_winthor(dados: list[tuple[int, float, float]]) -> bytes:
    """PROTOTIPO.md §5 "Regra de exportação Winthor": EXATAMENTE 3 colunas
    (código, preço unitário com 2 casas, quantidade), SEM CABEÇALHO. Filial/
    fornecedor/comprador não entram — são digitados manualmente na rotina 220
    do Winthor na hora de importar."""
    wb = Workbook()
    ws = wb.active
    for codigo, preco, quantidade in dados:
        ws.append([codigo, preco, quantidade])
    for row in range(1, ws.max_row + 1):
        ws.cell(row=row, column=2).number_format = "0.00"
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def exportar_winthor(
    id_pedido: int, usuario_login: str, usuario_id: int, ip: str | None = None
) -> tuple[bytes, str]:
    """Só em Fechado (PROTOTIPO.md §2.5); avança automaticamente para
    Exportado — na MESMA transação da leitura dos itens, para o arquivo
    baixado e o status novo nunca divergirem (ou os dois acontecem, ou
    nenhum). Quantidade da planilha é a REAL (quantidade x fator_exibicao
    CONGELADO do item) — é o que vai de fato para o estoque no Winthor
    (sql/03_tabelas_pedido.sql, comentário de APP_PEDIDO_ITEM.quantidade)."""
    with database.transacao() as conn:
        cur = conn.cursor()
        info = _bloquear_pedido(cur, id_pedido)
        if info["status"] != "Fechado":
            raise TransicaoInvalida(
                "Exportar para o Winthor exige o pedido em 'Fechado'"
                f" (status atual: '{info['status']}')."
            )
        cur.execute(
            "select id_produto, quantidade, fator_exibicao, preco_unitario"
            " from app_pedido_item where id_pedido = :id order by id_produto",
            {"id": id_pedido},
        )
        linhas = cur.fetchall()
        dados = [
            (int(codigo), round(float(preco), 2), float(quantidade) * float(fator))
            for codigo, quantidade, fator, preco in linhas
        ]
        conteudo = _gerar_xlsx_winthor(dados)

        agora = dt.datetime.now()
        cur.execute(
            "update app_pedido set status = 'Exportado', atualizado_em = :agora,"
            " atualizado_por = :usuario where id_pedido = :id",
            {"agora": agora, "usuario": usuario_login, "id": id_pedido},
        )
        cur.execute(
            """
            insert into app_pedido_status_hist (id_pedido, status_anterior, status_novo, alterado_em, alterado_por)
            values (:id, 'Fechado', 'Exportado', :agora, :usuario)
            """,
            {"id": id_pedido, "agora": agora, "usuario": usuario_login},
        )
        cur.close()
        auditoria.registrar(
            conn, usuario_id, "EXPORTAR_WINTHOR_PEDIDO", "APP_PEDIDO", str(id_pedido),
            {"qtd_itens": len(dados), "usuario": usuario_login}, ip,
        )

    nome = f"winthor_pedido_{id_pedido}_{dt.date.today():%Y-%m-%d}.xlsx"
    return conteudo, nome
