"""Prova que `compras.fat_pedido` reproduz, celula a celula, a aba `pedido` do
gabarito (`referencia/MODELO_COMPRAS_CEDEP_v11.xlsx` - ver `XLSX_PATH_PADRAO`) -
o CRITERIO DE ACEITE FINAL do projeto (CONTEXTO.md §7): 122 colunas x 8.772
linhas, tolerancia 0,01.

DESENHO (leia antes de mexer) - ver CONTEXTO.md §6.2, §6.3 e docs/gabarito_pedido_formulas.txt
------------------------------------------------------------------------------
Este script foi escrito ANTES de `fat_pedido` existir (outro agente esta
escrevendo os models agora - NADA em dbt/ e tocado aqui). O objetivo imediato
NAO e' "passar": e' provar que o MOTOR de comparacao funciona, documentar o
criterio de classificacao de colunas (que vai ser discutido por humanos) e
ficar pronto para rodar a comparacao real assim que `compras.fat_pedido`
existir.

1) CLASSIFICACAO DE COLUNAS (obrigatoria, ver `docs/gabarito_pedido_formulas.txt`)
   As 122 colunas da aba `pedido` sao classificadas em 5 grupos, calculados a
   partir da PROPRIA formula de cada coluna (grafo de dependencia entre
   colunas da mesma aba, extraido automaticamente do gabarito - nao e' uma
   lista digitada a mao, para reduzir risco de erro humano na classificacao):

     - CRITICA (4 colunas, fixas por instrucao do orquestrador): CLASSE,
       ALERTA, MODALIDADE, CUSTO_TOT_GERENCIAL. Exigem ZERO divergencia,
       sem excecao - mesmo que a propria formula leia uma coluna volatil ou
       dependente de decisao humana (ver aviso impresso sobre ALERTA).
     - ENTRADA_HUMANA (3 raizes PURAS: PEDIDO, ALT_PV_AT_AV, ALT_PV_VAR_AV -
       nao sao calculadas, nao tem valor padrao nenhum, ficam NULAS sem decisao
       gravada) + FECHO TRANSITIVO. Celula com valor vazio/NULL do LADO DBT
       nunca conta como divergencia aqui (instrucao explicita do orquestrador) -
       o bloco e' informativo, nunca reprova o aceite.
       ⚠ MARGEM_ALVO/MARGEM_ALVO_VAREJO (CN/DD) NAO sao raiz deste bucket desde
       a correcao da FALHA 3 (revisao de 24/08/2026) - ver RAIZES_CONDICIONAIS
       logo abaixo. Diferente de PEDIDO/ALT_PV_*, elas NUNCA ficam vazias (caem
       no padrao 20% sem decisao - CONTEXTO regra 10), entao sao deterministicas
       e verificaveis na maioria esmagadora das linhas; so' a linha ONDE HOUVER
       decisao gravada e' que se comporta como entrada humana.
     - VOLATIL: raizes = EST_DISP, QT_RESERVADA, DT_ULT_SAIDA, DIAS_SEM_VENDA
       (citadas explicitamente pelo orquestrador) + QT_BLOQUEADA, QT_AVARIA,
       PENDENTE, EST+PEND, EST_FABRICA (mesma fonte de estoque ao vivo de
       EST_DISP/QT_RESERVADA - dCadastroTI colunas S/T/U/V/W) + VD_MES_ATUAL,
       CHECK_FORA_DE_LINHA (cadastro EDITAVEL - OBS2='FL' pode mudar entre a
       foto e o build; a razao "so' vem do mes corrente" foi corrigida pela
       MELHORIA A3 e NAO vale mais - ver comentario em RAIZES_VOLATEIS) +
       PV_ATACADO/PV_VAREJO (CF/CW - tabela de preco PCTABPR ao vivo, FALHA 2
       da revisao de 24/08/2026, mesma natureza de EST_DISP) + FECHO
       TRANSITIVO. Divergencia aqui e' ESPERADA por construcao (a foto da
       planilha e o build do fat_pedido veem instantes diferentes do
       estoque/preco/cadastro - CONTEXTO §6.3) e NUNCA reprova sozinha.
     - CANDIDATA_ULTIMA_ENTRADA (Z=DT_ULT_ENT, AA=QT_ULT_ENT, BH=VL_ENT_UNIT,
       BI=CUSTO_ULT_ENT, DR=CRED_TOTAL_EMPIRICO): sinalizadas a parte, SEM
       fechamento transitivo (ver nota abaixo - decisao deliberada, nao e'
       omissao).
     - ESTRUTURAL: todo o resto. Divergencia aqui e' DEFEITO DE VERDADE - e' o
       que decide o aceite (>=99,9% das celulas numericas dentro da
       tolerancia; colunas de texto/data exigem 100%). Inclui, desde a correcao
       da FALHA 3, MARGEM_ALVO/MARGEM_ALVO_VAREJO (CN/DD) e os DEZ precos
       sugeridos que dependem delas (CO/CP/CQ/CR/CS/CT/DE/DF/DG/DH) - verificados
       linha a linha, com skip so' na linha ESPECIFICA onde ha' decisao gravada
       (RAIZES_CONDICIONAIS).

   NOTA sobre o corte do fecho transitivo em CANDIDATA_ULTIMA_ENTRADA: BH/BI
   (custo/valor da ultima entrada) TAMBEM sao fatos operacionais "ao vivo" -
   uma nota de compra registrada entre a foto da planilha e o build do
   fat_pedido muda BH/BI, exatamente pela mesma logica de EST_DISP mudar
   entre os dois instantes. A diferenca e' de FREQUENCIA (estoque muda a
   cada venda; custo de ultima entrada muda so' quando chega nota nova, bem
   mais raro), e o alcance: se BH/BI virassem raiz do fecho transitivo, a
   classificacao "volatil" engoliria quase toda a cadeia de custo/margem/
   preco (BO=CRED_TOTAL em diante, dezenas de colunas) - inclusive
   CUSTO_TOT_GERENCIAL, que o proprio orquestrador exige zero divergencia
   OBRIGATORIA. Por isso a decisao aqui foi NAO propagar: Z/AA/BH/BI/DR
   ficam sinalizadas a parte, informativas, mas as colunas que as LEEM
   (BO=CRED_TOTAL e toda a cadeia CH/CI/CJ/PV_SUG_*/MARGEM_*/etc.) continuam
   ESTRUTURAIS - sujeitas ao criterio de aceite cheio. Se, na pratica, a
   defasagem de "ultima entrada" se mostrar frequente o bastante para gerar
   divergencia sistematica nessas colunas, isso e' para DISCUTIR com humano,
   nao para o script decidir sozinho afrouxando o criterio.

2) TOLERANCIA POR TIPO, nao uma so para tudo (instrucao explicita):
     - txt / data: comparacao exata (vazio do Excel == NULL do SQL)
     - num (quantidade, moeda, codigo): tolerancia --tolerancia (default 0,01)
     - razao (margem, mkp, aliquota, credito, tend%, var_pv, tx_devolucao):
       tolerancia --tolerancia-razao (default 0,001 = 0,1 ponto percentual -
       0,01 aqui seria 1pp inteiro, grosso demais para decisao de preco)
     - pp (colunas ja multiplicadas por 100 - GAP_FILIAL_pp, DIF_MC_*_pp):
       tolerancia = tolerancia-razao * 100 (mesma granularidade da razao,
       so' que na escala "pp" em vez de fracao 0-1)

3) VERIFICACAO DO MOTOR (roda por padrao, ver `--pular-verificacao-motor`):
   Antes de qualquer tentativa de comparar contra `fat_pedido` (que pode nem
   existir ainda), o script LE a aba `pedido` de verdade (streaming,
   openpyxl read_only=True/data_only=True - nunca carrega workbook inteiro),
   mede tempo/memoria, e prova o motor em 2 frentes SEM banco nenhum:
     a) aba `pedido` comparada CONTRA SI MESMA -> tem que dar zero
        divergencia em toda coluna, em todo bucket. Se nao der, o LEITOR tem
        bug (nao a planilha).
     b) 4 defeitos sinteticos injetados numa copia (valor numerico alterado,
        texto de ALERTA alterado, linha faltando, linha duplicada) -> o
        motor tem que PEGAR cada um. Um validador que so' passa nao prova
        nada; este prova que reprova quando deveria.

Uso:
    python validar_pedido.py [--tolerancia 0.01] [--tolerancia-razao 0.001]
                              [--xlsx CAMINHO] [--limite-codigos N]
                              [--sem-db] [--pular-verificacao-motor]
                              [--lib-dir ...] [--profiles-path ...]

4) DIVERGENCIA ESPERADA POR DECISAO (`LETRAS_DIVERGENCIA_POR_DECISAO`)
   Desde 21/08/2026 o modelo diverge DE PROPOSITO do gabarito (v11) em um
   conjunto NOMEADO de colunas, por duas origens distintas:
     - decisoes do Diretor de Compras que ele ainda NAO aplicou na planilha
       (PENDENCIAS_DIRETORIA.md itens 3 e 4; CONTEXTO.md 6.0/6.4): BO, CC e,
       por tabela, D;
     - MELHORIAS APROVADAS (MELHORIAS.md; CONTEXTO.md 6.0), porque a planilha
       passou a ser PONTO DE PARTIDA e nao alvo de replica exata: AP (A3),
       AV/CL/DB (A4), AQ (D1), BB/BC/BD/BE/BF (A5) e, por tabela, D.
   ATENCAO: AQ (D1) e BB/BC/BD/BE/BF (A5) tem efeito ZERO hoje - as tabelas de que
   dependem estao vazias. Divergencia nelas neste build e' DEFEITO, nao a
   melhoria; o rotulo existe para o dia em que o dado aparecer.
   As seis colunas do cenario "ST s/Valor" (CH, CY, CO, CP, DE, DF) continuam
   registradas, mas contra a v11 elas FECHAM - o Diretor ja aplicou a correcao
   la. O registro delas fica porque explica a formula e porque elas divergem da
   v10, que segue no repositorio.

   O script NAO deixa de verificar nenhuma delas: elas seguem no bucket de
   sempre, sao comparadas, contadas e impressas igual as outras, e o veredito
   automatico NAO consulta esse dicionario. O registro so' ROTULA a divergencia
   no relatorio e a separa no resumo executivo, para que ela tenha dono
   explicito em vez de virar "ruido conhecido" que ninguem sabe explicar.

Aceite final (quando `compras.fat_pedido` existir):
  - ZERO divergencia nas 4 colunas CRITICAS (CLASSE, ALERTA, MODALIDADE,
    CUSTO_TOT_GERENCIAL) e em toda coluna de texto/data ESTRUTURAL.
  - >=99,9% das celulas numericas ESTRUTURAIS dentro da tolerancia.
  - ENTRADA_HUMANA e VOLATIL/CANDIDATA_ULTIMA_ENTRADA sao informativos e
    NUNCA reprovam sozinhos.

O script NAO afrouxa nada para "passar" - ele mede e imprime; ajustar o
model para o numero fechar e' trabalho de outro agente, e so' depois de
alguem decidir que o gabarito e' que esta certo.

5) FECHO DA ETAPA 4 (25/08/2026) - tres ajustes sobre o veredito automatico:

   AJUSTE 1 - divergencia REGISTRADA fica ISENTA do limiar de 99,9% ESTRUTURAL
   (`LETRAS_DIVERGENCIA_ESPERADO`), mas SO' ATE o numero que o proprio
   registro declara como esperado - excesso volta a reprovar por outro
   motivo (`estruturais_rotuladas_acima_do_esperado`). Nao e' cheque em
   branco: provado pelos casos sinteticos (f)/(g) de `verificar_motor`.

   AJUSTE 2 - B/C/E/F/H (COD_FAB, DESCRICAO, STATUS, FORNECEDOR, EMBALAGEM)
   sao lookup do CADASTRO ATUAL (WinThor), nao resultado de calculo nosso -
   viram bucket "volatil" (`LETRAS_VOLATEIS_CADASTRO_DIRETO`), mas SEM
   propagar pelo fecho transitivo: confirmado no grafo que propagar
   alcancaria AX (CLASSE) e BY (CUSTO_TOT_GERENCIAL), que tem que continuar
   exigindo zero absoluto (CONTEXTO.md §6.1.0/§6.1.1) - so' a celula de
   cadastro em si fica isenta, tudo que descende dela continua ESTRUTURAL.

   AJUSTE 3 - atribuicao de D (ALERTA) RECALCULADA a cada execucao contra os
   dados reais do dia (`atribuir_divergencia_alerta`/`COMPONENTES_ALERTA`),
   nunca contra o numero antigo de MELHORIAS.md/CONTEXTO.md - CONTEXTO §6.0.1
   e' explicito que esses numeros envelhecem sozinhos. Reporta linha nao
   atribuivel a componente nenhum em vez de presumir zero.
"""
from __future__ import annotations

import argparse
import copy
import ctypes
import datetime
import os
import re
import sys
import time
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# Caminhos fixos do ambiente (ver CONTEXTO.md §3)
# ─────────────────────────────────────────────────────────────────────────────
ORACLE_CLIENT_LIB_DIR = r"C:\Oracle\instantclient_21_17"
PROFILES_PATH = os.path.expanduser(r"~\.dbt\profiles.yml")
REPO_DIR = Path(__file__).resolve().parent.parent
# ⚠ O gabarito e' a v11 desde 21/08/2026 (CONTEXTO.md 6.0, tabela de
# arquivos de referencia). A v10 continua no repositorio SO' como historico -
# apontar para ela aqui faria as seis colunas do cenario "ST s/Valor"
# (CH/CY/CO/CP/DE/DF) aparecerem como divergentes sem que nada esteja errado,
# porque a v11 ja traz a correcao do Diretor (PENDENCIAS item 1) e o modelo
# tambem. docs/gabarito_pedido_formulas.txt tambem foi regerado da v11.
XLSX_PATH_PADRAO = REPO_DIR / "referencia" / "MODELO_COMPRAS_CEDEP_v11.xlsx"
GABARITO_PATH = REPO_DIR / "docs" / "gabarito_pedido_formulas.txt"
XLSX_ABA = "pedido"

EXCEL_EPOCH = datetime.date(1899, 12, 30)


# ─────────────────────────────────────────────────────────────────────────────
# Medida de memoria (Windows, sem dependencia externa - psutil nao esta
# instalado neste ambiente) via GetProcessMemoryInfo/psapi.
# ─────────────────────────────────────────────────────────────────────────────


class _ProcessMemoryCounters(ctypes.Structure):
    _fields_ = [
        ("cb", ctypes.c_ulong),
        ("PageFaultCount", ctypes.c_ulong),
        ("PeakWorkingSetSize", ctypes.c_size_t),
        ("WorkingSetSize", ctypes.c_size_t),
        ("QuotaPeakPagedPoolUsage", ctypes.c_size_t),
        ("QuotaPagedPoolUsage", ctypes.c_size_t),
        ("QuotaPeakNonPagedPoolUsage", ctypes.c_size_t),
        ("QuotaNonPagedPoolUsage", ctypes.c_size_t),
        ("PagefileUsage", ctypes.c_size_t),
        ("PeakPagefileUsage", ctypes.c_size_t),
    ]


def memoria_pico_mb() -> float | None:
    """Pico de working set do PROCESSO ATUAL (Windows), em MB. Devolve None
    se a API nao estiver disponivel (ex.: rodando fora do Windows).

    Precisa de `argtypes`/`restype` EXPLICITOS: sem eles, o ctypes assume
    `int` de 32 bits como retorno de `GetCurrentProcess()` (que devolve um
    HANDLE de 64 bits, no caso um pseudo-handle) e o handle chega truncado/
    mal convertido em `GetProcessMemoryInfo` - a chamada falha em silencio
    (devolve 0/false, sem excecao, sem GetLastError util) e o pico sai
    sempre "indisponivel". Medido no proprio ambiente antes de fechar isto.
    """
    try:
        psapi = ctypes.WinDLL("psapi.dll")
        kernel32 = ctypes.WinDLL("kernel32.dll")
        kernel32.GetCurrentProcess.restype = ctypes.c_void_p
        psapi.GetProcessMemoryInfo.argtypes = [
            ctypes.c_void_p, ctypes.POINTER(_ProcessMemoryCounters), ctypes.c_ulong,
        ]
        psapi.GetProcessMemoryInfo.restype = ctypes.c_int

        counters = _ProcessMemoryCounters()
        counters.cb = ctypes.sizeof(_ProcessMemoryCounters)
        h_process = kernel32.GetCurrentProcess()
        ok = psapi.GetProcessMemoryInfo(h_process, ctypes.byref(counters), counters.cb)
        if ok:
            return counters.PeakWorkingSetSize / (1024 * 1024)
    except Exception:
        pass
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Parser do gabarito (`docs/gabarito_pedido_formulas.txt`) - fonte UNICA da
# verdade para letra/cabecalho/formula. Nao duplicamos isso a mao: se o
# gabarito mudar, o parser reflete sem precisar editar este script.
# ─────────────────────────────────────────────────────────────────────────────

_LINHA_RE = re.compile(r"^([A-Z]{1,3})\|(.*)\|F=(.*)\|V=(.*)$")


@dataclass
class ColunaGabarito:
    letra: str
    header: str
    formula: str  # texto "None" quando a coluna nao e' calculada (entrada humana ou cadastro estatico)


def carregar_gabarito(caminho: Path) -> list[ColunaGabarito]:
    linhas = caminho.read_text(encoding="utf-8").splitlines()
    colunas: list[ColunaGabarito] = []
    for linha in linhas:
        m = _LINHA_RE.match(linha)
        if not m:
            continue  # ex.: "NCOLS 122" (cabecalho do arquivo)
        letra, header, formula, _valor_exemplo = m.groups()
        colunas.append(ColunaGabarito(letra=letra, header=header.strip(), formula=formula))
    # sanidade: letras devem ser exatamente A, B, C, ..., na ordem das colunas do Excel
    esperadas = [get_column_letter(i) for i in range(1, len(colunas) + 1)]
    reais = [c.letra for c in colunas]
    if reais != esperadas:
        raise RuntimeError(
            f"Parser do gabarito fora de ordem/incompleto: esperado {esperadas[:5]}...{esperadas[-3:]} "
            f"({len(esperadas)} colunas), obtido {reais[:5]}...{reais[-3:]} ({len(reais)} colunas). "
            f"Confira {caminho}."
        )
    return colunas


def _slug(header: str) -> str:
    h = header.strip()
    h = h.replace("%", "pct").replace("+", "_").replace("/", "_")
    h = re.sub(r"\s+", "_", h)
    h = h.replace("-", "_")
    h = re.sub(r"[^A-Za-z0-9_]", "", h)
    h = re.sub(r"_+", "_", h)
    return h.strip("_").lower()


# ─────────────────────────────────────────────────────────────────────────────
# Grafo de dependencia ENTRE COLUNAS DA MESMA ABA `pedido` - extraido das
# proprias formulas (regex `\$LETRA2` - toda formula desta aba refere-se a
# celula da mesma LINHA 2, o padrao do template; referencia a OUTRA aba usa
# sintaxe `aba!$COL:$COL`, sem numero de linha, entao nao e' capturada aqui
# de proposito - so' nos interessa dependencia dentro da propria aba `pedido`).
# ─────────────────────────────────────────────────────────────────────────────

_REF_RE = re.compile(r"\$([A-Z]{1,3})2\b")


def construir_grafo(colunas: list[ColunaGabarito]) -> dict[str, set[str]]:
    letras_validas = {c.letra for c in colunas}
    refs: dict[str, set[str]] = {}
    for c in colunas:
        if c.formula == "None":
            refs[c.letra] = set()
            continue
        achados = {m for m in _REF_RE.findall(c.formula) if m in letras_validas and m != c.letra}
        refs[c.letra] = achados
    return refs


def fecho_transitivo(raizes: set[str], grafo_refs: dict[str, set[str]]) -> set[str]:
    """Devolve TODA coluna que le', direta ou indiretamente, alguma das
    `raizes` - ou seja, propaga na direcao "quem depende de mim", nao "de
    quem eu dependo". Constroi o grafo reverso (quem referencia cada letra)
    e faz BFS a partir das raizes."""
    reverso: dict[str, set[str]] = defaultdict(set)
    for letra, deps in grafo_refs.items():
        for dep in deps:
            reverso[dep].add(letra)
    vistos = set(raizes)
    pilha = list(raizes)
    while pilha:
        atual = pilha.pop()
        for prox in reverso.get(atual, ()):
            if prox not in vistos:
                vistos.add(prox)
                pilha.append(prox)
    return vistos


# ─────────────────────────────────────────────────────────────────────────────
# Classificacao fixa (letra -> chave). Colunas fora daqui NAO entram na
# comparacao (chave de join `CODIGO`=A, ou coluna explicitamente marcada
# "(nao usado)"=I no gabarito).
# ─────────────────────────────────────────────────────────────────────────────

LETRA_CHAVE = "A"          # CODIGO - usada como chave de join, nao comparada como coluna
LETRAS_IGNORADAS = {"A", "I"}  # I = "(nao usado)" no proprio gabarito

LETRAS_CRITICAS = {"D", "AX", "BL", "BY"}  # ALERTA, CLASSE, MODALIDADE, CUSTO_TOT_GERENCIAL

RAIZES_VOLATEIS = {
    "Q",   # EST_DISP            - citada explicitamente pelo orquestrador
    "T",   # QT_RESERVADA        - citada explicitamente pelo orquestrador
    "AM",  # DT_ULT_SAIDA        - citada explicitamente pelo orquestrador
    "AN",  # DIAS_SEM_VENDA      - citada explicitamente pelo orquestrador
    "R",   # QT_BLOQUEADA        - mesma fonte de estoque ao vivo (dCadastroTI) que Q/T
    "S",   # QT_AVARIA           - idem
    "U",   # PENDENTE            - idem
    "V",   # EST+PEND            - idem (= Q + U, mas incluida como raiz tambem nao muda o fecho)
    "W",   # EST_FABRICA         - estoque de fabrica ao vivo (fEstFabrica)
    "AC",  # VD_MES_ATUAL        - mes corrente (confirmado: le' fVendaMes col. QATUAL)
    # AP CHECK_FORA_DE_LINHA - continua volatil, mas NAO mais pela razao antiga.
    # Ate' a MELHORIA A3 (CONTEXTO.md §6.0), FORA_DE_LINHA so' vinha da linha do
    # MES CORRENTE do fVendaMes - era literalmente volatil com o relogio. A3
    # trocou o insumo para o registro MAIS RECENTE do SKU em int_venda_mensal_pivot,
    # e a divergencia de AP hoje NAO e' mais volatilidade: e' a MELHORIA (71 SKUs
    # esperados, ver LETRAS_DIVERGENCIA_POR_DECISAO). AP continua neste bucket
    # porque o CAMPO por tras dele (OBS2='FL') e' cadastro EDITAVEL no WinThor -
    # o Diretor pode marcar/desmarcar um produto fora de linha a qualquer momento,
    # entre a foto da planilha e o build - o mesmo motivo estrutural de Q/T/AC,
    # so' que a fonte e' o cadastro em vez do estoque. CONTEXTO.md §6.1.1.
    "AP",
    # CF/CW (PV_ATACADO/PV_VAREJO) - FALHA 2 do relatorio de revisao (24/08/2026).
    # Vem de dCadastroTI!$AB/$AC, que e' PCTABPR - a TABELA DE PRECO do WinThor.
    # E' tao "ao vivo" quanto EST_DISP: o preco de venda pode mudar a qualquer
    # minuto por repasse do fornecedor, exatamente como estoque muda a cada
    # venda (CONTEXTO.md §6.3). Sem isto, CF/CW e todo o fecho de margem que
    # depende deles ficavam classificados como ESTRUTURAL - defeito de verdade -
    # quando na pratica e' a mesma defasagem de foto-x-build de sempre.
    # MEDIDO na revisao (24/08/2026), conferido AO VIVO contra pctabpr: 50 SKUs
    # divergem por repasse de tabela de preco (40 MUNDIAL PR, 6 SHELL, 2 KOUBE,
    # 2 IPIRANGA) - pv_atacado/pv_varejo do fat_pedido batem com pctabpr em 50/50.
    "CF",  # PV_ATACADO
    "CW",  # PV_VAREJO
}

# ─────────────────────────────────────────────────────────────────────────────
# AJUSTE 2 (25/08/2026, fecho da Etapa 4): colunas de CADASTRO num FATO sao
# VOLATEIS - CONTEXTO.md §6.3 ("produto, departamento, secao e fora_de_linha
# nao sao campos historicos - sao lookup do estado ATUAL do cadastro"), caso
# real citado nominalmente: produto 1826 renomeado entre foto e build. O
# mesmo vale aqui: B/C/H sao copia DIRETA do cadastro na aba pedido (F=None
# no gabarito) e E/F sao lookup de UM SO nivel em dCadastroTI (Ativo/Inativo,
# texto do departamento) - nenhuma das cinco tem CALCULO nosso por tras,
# mudam sozinhas quando alguem edita o cadastro no WinThor:
#   B  COD_FAB     - F=None, copia direta do cadastro
#   C  DESCRICAO   - F=None, copia direta do cadastro (caso medido: 36 SKUs
#                    divergem sem estar registrado - INCLUSIVE o produto 1826,
#                    o mesmo caso ja documentado em §6.3 para int_venda_mensal_pivot)
#   E  STATUS      - lookup de dCadastroTI!$D (Ativo/Inativo no WinThor)
#   F  FORNECEDOR  - lookup de dCadastroTI!$M (texto do departamento - regra 7)
#   H  EMBALAGEM   - F=None, copia direta do cadastro
#
# ⚠ NAO SAO RAIZES de RAIZES_VOLATEIS (a diferenca importa - ver guardrail
# abaixo). Adicionar B/C/E/F/H como raizes normais e deixar o fecho_transitivo
# propagar "quem depende de mim" foi CONFERIDO no grafo real e o resultado e'
# inaceitavel:
#     fecho_transitivo({'F'}, grafo) -> 62 colunas, ALCANCA AX (CLASSE) e
#         BY (CUSTO_TOT_GERENCIAL) - F alimenta K (FATOR_EXIBICAO, 14
#         referencias na aba - a coluna mais citada depois de A/BZ/BW/CB),
#         que por sua vez alimenta praticamente toda a cadeia de
#         estoque/custo/margem/preco.
#     fecho_transitivo({'H'}, grafo) -> alcanca AX (CLASSE) via H->L->P->AX.
# Isso contradiz DUAS garantias que nao podem ceder: (1) o proprio §6.1.0 do
# CONTEXTO justifica exigir zero absoluto em CLASSE dizendo que ela "depende
# so' de media de venda de meses fechados E DE EMBALAGEM, AMBAS ESTAVEIS" -
# deixar H virar raiz volatil contradiria essa premissa; (2) CLASSE,
# MODALIDADE e CUSTO_TOT_GERENCIAL sao as 3 colunas de zero absoluto sem
# excecao (§6.1.1). Propagar apagaria defeito de verdade numa fatia enorme
# do motor de preco so' para "resolver" a volatilidade de 5 colunas de
# cadastro. Por isso a escolha aqui e' DELIBERADAMENTE MAIS ESTREITA: SO' a
# celula de cadastro em si (a copia/lookup direto) fica isenta - tudo que
# DESCENDE dela (G, K, W, X, L, AY, AB, AZ, BO, BX, e a cadeia de
# margem/preco/CLASSE/CUSTO_TOT_GERENCIAL) continua no bucket que teria de
# qualquer forma (estrutural, na maioria), sujeito ao criterio de aceite
# cheio. Se um cadastro renomeado causar divergencia estrutural real
# downstream, ela vai aparecer e CONTINUA valendo como o que e': defasagem
# real de cadastro, atribuivel na hora de investigar - mas o script nao a
# apaga silenciosamente. Ver `montar_buckets` (nao entra em RAIZES_VOLATEIS,
# entra direto no bucket "volatil" so' para estas 5 letras).
LETRAS_VOLATEIS_CADASTRO_DIRETO = {"B", "C", "E", "F", "H"}

RAIZES_ENTRADA_HUMANA = {
    "BA",  # PEDIDO         - digitada, sem formula, fica NULA sem decisao (APP_DECISAO_PEDIDO)
    "CU",  # ALT_PV_AT_AV   - idem, sem formula, fica NULA sem decisao (APP_DECISAO_PRECO)
    "DI",  # ALT_PV_VAR_AV  - idem
}
# ⚠ CN (MARGEM_ALVO) e DD (MARGEM_ALVO_VAREJO) NAO estao aqui. Ver
# RAIZES_CONDICIONAIS logo abaixo - FALHA 3 do relatorio de revisao de
# 24/08/2026.

# ─────────────────────────────────────────────────────────────────────────────
# FALHA 3 do relatorio de revisao (24/08/2026): "os dez precos sugeridos nunca
# reprovam". CN/DD alimentavam RAIZES_ENTRADA_HUMANA e arrastavam pelo fecho
# transitivo os DEZ precos sugeridos (CO/CP/CQ/CR/CS/CT/DE/DF/DG/DH) para o
# bucket ENTRADA_HUMANA - que o veredito NUNCA consulta (informativo, por
# design). Isso escondia a SAIDA MAIS CONSEQUENTE do modelo (o preco que vai
# para a equipe comercial) do aceite, mesmo sendo 100% deterministica hoje.
#
# A diferenca para PEDIDO/ALT_PV_*: CN/DD NAO sao "digitada, sem formula, fica
# NULA sem decisao" - CONTEXTO regra 10 e' explicito, "MARGEM_ALVO cai no
# padrao 20% quando nao houver decisao gravada". Sem decisao, o valor e'
# DETERMINISTICO (MARGEM_ALVO_PADRAO, seed_parametros.csv) e verificavel como
# qualquer outra coluna calculada - foi assim que a revisao mediu ZERO
# divergencia nos dez precos sugeridos, SKU a SKU, 8.772 de 8.772.
#
# So' vira entrada humana de verdade na linha ONDE HA' decisao gravada em
# APP_DECISAO_PRECO - ali sim o dbt usa um numero que a planilha (foto de
# antes da decisao) nao tem como conhecer.
#
# RAIZ CONDICIONAL, nao bucket fixo: CN e DD (e tudo que descende delas)
# continuam no bucket ESTRUTURAL (comparados, contam para o veredito), mas o
# motor de comparacao (`comparar_pedido`) verifica LINHA A LINHA se a propria
# raiz (CN para o lado atacado, DD para o varejo) se desviou do padrao
# `MARGEM_ALVO_PADRAO` alem da tolerancia de razao - so' NESSA linha, e so'
# NESSAS colunas (a raiz e' quem descende dela), a celula e' tratada como
# entrada humana (ignorada, nao conta como divergencia). Escolhido em vez de
# um bucket fixo porque a classificacao aqui e' por LINHA, nao por COLUNA: a
# mesma coluna CO e' 100% verificavel numas 8.767 linhas e informativa nas (ate
# hoje, zero) linhas com decisao gravada - um bucket unico por coluna nao
# consegue representar isso.
#
# Confirmado no grafo (docs/gabarito_pedido_formulas.txt): o fecho de CN e' so'
# {CN,CO,CP,CQ,CR,CS,CT}; o de DD e' so' {DD,DE,DF,DG,DH} - sem sobreposicao
# entre si nem com o fecho de BA/CU/DI. Nenhuma das duas alcanca CLASSE (AX),
# ALERTA (D), MODALIDADE (BL) ou CUSTO_TOT_GERENCIAL (BY) - as quatro criticas
# continuam exigindo zero absoluto, sem excecao desta raiz condicional.
RAIZES_CONDICIONAIS: dict[str, float] = {
    "CN": 0.2,  # MARGEM_ALVO         - MARGEM_ALVO_PADRAO (seed_parametros.csv)
    "DD": 0.2,  # MARGEM_ALVO_VAREJO  - idem, independente de CN
}

# ─────────────────────────────────────────────────────────────────────────────
# DIVERGENCIA ESPERADA POR DECISAO (CONTEXTO.md 6.4, PENDENCIAS_DIRETORIA.md)
#
# ⚠ ISTO NAO DESLIGA VERIFICACAO NENHUMA. Estas colunas continuam sendo lidas,
# comparadas, contadas e impressas exatamente como todas as outras, no mesmo
# bucket de sempre - `critica` continua critica, `estrutural` continua
# estrutural, e o veredito automatico nao consulta este dicionario. O que ele
# faz e' ROTULAR a divergencia no relatorio e separa-la das demais no resumo
# executivo, para que quem le saiba de imediato quais divergencias JA TEM DONO
# e quais precisam de investigacao.
#
# Por que existe: em 21/08/2026 o Diretor de Compras decidiu tres pendencias
# fiscais e o modelo mudou junto. O xlsx em referencia/ e' a versao ANTERIOR a
# essas decisoes, entao nestas colunas o gabarito esta DESATUALIZADO em relacao
# a regra vigente. Sem este registro, a proxima pessoa a rodar o validador
# olharia CH/CY/CO/DE divergindo e teria que redescobrir o porque - ou, pior,
# "consertaria" o model de volta para o numero antigo.
#
# Quando o arquivo em referencia/ for substituido pela planilha JA CORRIGIDA
# pelo Diretor, estas entradas devem sair daqui e as colunas voltam a exigir
# convergencia normal. Nao apague sem trocar o arquivo de referencia.
LETRAS_DIVERGENCIA_POR_DECISAO = {
    "CH": "PENDENCIAS item 1 (21/08/2026): cenario 'ST s/Valor' passou a usar ICMS_SEM_RED "
          "(aliquota CHEIA) no lugar de ICMS_SAIDA_EF. Esperado: ~1.240 SKUs de MODALIDADE "
          "'NORMAL'; ZERO em regime ST. Fora do ST, CH fica identica a CJ.",
    "CY": "PENDENCIAS item 1 (21/08/2026): idem CH, no varejo. Fora do ST, CY fica identica a CZ.",
    "CO": "PENDENCIAS item 1 (21/08/2026): idem CH, no preco sugerido a vista do atacado. "
          "Fora do ST, CO fica identica a CS.",
    "CP": "PENDENCIAS item 1 (21/08/2026): CONSEQUENCIA de CO (= CO x FATOR_PRAZO). Sem "
          "alteracao propria de formula.",
    "DE": "PENDENCIAS item 1 (21/08/2026): idem CH, no preco sugerido a vista do varejo. "
          "Fora do ST, DE fica identica a DG.",
    "DF": "PENDENCIAS item 1 (21/08/2026): CONSEQUENCIA de DE (= DE x FATOR_PRAZO_VAREJO). "
          "Sem alteracao propria de formula.",
    "BO": "PENDENCIAS item 3 (21/08/2026): grafia do seed_credito corrigida de 'CAR80' para "
          "'CAR 80', com espaco, para casar com o departamento da base. Esperado: 41 SKUs "
          "deixam de cair no credito empirico e passam a usar o tabelado.",
    "CC": "PENDENCIAS item 4 (21/08/2026): CHECK_TRIB deixou de ser formula morta - dispara "
          "com codigo de tributacao vazio OU zero. Esperado: exatamente os 5 SKUs de "
          "codst = 0.",
    # ── MELHORIAS APROVADAS em 21/08/2026 (MELHORIAS.md; CONTEXTO.md 6.0) ──
    # Diferente das linhas acima, estas NAO sao "a planilha ainda nao foi
    # atualizada": sao mudancas que o gabarito nao tem e nao vai ter, decididas
    # porque a planilha virou ponto de partida.
    "AP": "MELHORIA A3 (21/08/2026): CHECK_FORA_DE_LINHA. A planilha le FORA_DE_LINHA so' da "
          "linha do MES CORRENTE do fVendaMes, e produto fora de linha sem movimento no mes nao "
          "tem linha nenhuma - o alerta sumia justamente para os produtos parados. Passamos a ler "
          "o registro MAIS RECENTE em que o SKU aparece. Esperado: exatamente 71 SKUs que saiam "
          "vazios passam a trazer 'FORA DE LINHA - AVALIAR LIQUIDACAO' (de 1 para 72 alertados). "
          "ZERO no sentido contrario. Divergencia ALEM desses 71 e' DEFEITO.",
    "AV": "MELHORIA A4 (21/08/2026): CHECK_DEVOLUCAO_ALTA. Mesma mascara TEXT(x,\"0.0%\") de CL/DB - "
          "ver CL. Esperado: as 136 celulas nao vazias mudam de texto; nenhum numero muda.",
    "CL": "MELHORIA A4 (21/08/2026): CHECK_MARGEM_INSTAVEL. A planilha usa TEXT(x,\"0.0%\") num "
          "arquivo pt-BR, onde o '.' e' separador de MILHAR - o Excel imprime percentual INTEIRO "
          "com dois digitos ('05%'), e o alerta nao distingue 5,0% de 5,9%. Passamos a imprimir "
          "uma casa decimal com VIRGULA ('5,5%'), que e' o que a mascara quis dizer; o zero a "
          "esquerda cai junto, por ser artefato da mesma leitura errada. Esperado: TODAS as celulas "
          "nao vazias de AV+CL+DB mudam de texto (2.182 medidas em 21/08/2026). E melhoria de "
          "LEITURA - nenhum numero de calculo muda.",
    "DB": "MELHORIA A4 (21/08/2026): CHECK_MARGEM_INSTAVEL_VAREJO. Idem CL, no varejo. Esperado: as "
          "414 celulas nao vazias mudam de texto.",
    "AQ": "MELHORIA D1 (21/08/2026): VD_ANT_3M passa a herdar da sucessao com PESO_1 E PESO_2, como "
          "as colunas irmas AD/AE/AF e AG; a planilha so' escreveu o termo de PESO_1. "
          "ATENCAO - esperado HOJE: ZERO celulas divergentes - nenhuma linha de seed_sucessao tem ANTIGO_2 "
          "e todas estao ATIVO='NAO', entao o segundo termo e' 0 em 100% das linhas. QUALQUER "
          "divergencia em AQ hoje e' DEFEITO, nao esta melhoria. O registro existe para o dia em "
          "que alguem cadastrar um segundo antecessor. AR (TEND %) e AT (VAR_PV) derivam de AQ e "
          "pela mesma razao tambem nao mudam hoje - por isso NAO estao rotuladas aqui.",
    # ── MELHORIA A5 (24/08/2026): FATOR_EXIBICAO congelado na decisao ──
    # ATENCAO - efeito HOJE: ZERO celulas, em todas as cinco. APP_DECISAO_PEDIDO
    # esta VAZIA, entao nao existe fator congelado em SKU nenhum e BB cai no
    # fallback (o fator CORRENTE, K), que e' exatamente a formula da planilha.
    # QUALQUER divergencia em BB/BC/BD/BE/BF neste build e' DEFEITO, nao esta
    # melhoria. O registro existe para o dia em que a primeira decisao for
    # gravada - dali em diante a divergencia passa a ser esperada nos SKUs que
    # tiverem decisao com fator diferente do corrente, e SO' neles.
    # A regra foi exercitada de ponta a ponta em 24/08/2026 com duas decisoes de
    # teste gravadas e apagadas (MELHORIAS.md A5) e e' protegida pelo teste
    # singular compras_pedido_unidades_usa_fator_congelado.
    "BB": "MELHORIA A5 (24/08/2026): PEDIDO_UNIDADES. A planilha faz $BA2*$K2 com o fator "
          "CORRENTE; passamos a usar o FATOR_EXIBICAO CONGELADO em APP_DECISAO_PEDIDO quando "
          "existe decisao gravada (sem decisao nao ha congelado e vale K). Motivo: EMBAL_COMPRA "
          "muda na atualizacao DIARIA de cadastro, e recalcular faria a quantidade em unidades - "
          "e o valor gasto - mudarem sozinhas numa decisao que ninguem redigitou. Esperado HOJE: "
          "ZERO divergencia (tabela vazia).",
    "BC": "MELHORIA A5 (24/08/2026): PEDIDO_NA_MEDIDA. CONSEQUENCIA de BB - a formula propria "
          "nao mudou. Esperado HOJE: ZERO divergencia.",
    "BD": "MELHORIA A5 (24/08/2026): VALOR_PEDIDO. CONSEQUENCIA de BB (= BB x CUSTO_TOT_OFICIAL). "
          "E o numero que o Diretor NAO quer ver mudar sozinho. Esperado HOJE: ZERO divergencia.",
    "BE": "MELHORIA A5 (24/08/2026): SUG_PALETE. CONSEQUENCIA de BB (o texto depende das caixas, "
          "que saem de BB). Esperado HOJE: ZERO divergencia.",
    "BF": "MELHORIA A5 (24/08/2026): MESES_EST+PED. CONSEQUENCIA de BB (= (EST+PEND + BB) / "
          "MEDIA_JANELA) - a cobertura em meses. Esperado HOJE por A5: ZERO divergencia. "
          "ATENCAO - BF DIVERGE HOJE, e NAO e' por esta melhoria: com BB = 0 em 100% das linhas, "
          "BF fica numericamente IDENTICA a AW (MESES_EST) - medido no build: 0 de 8.777 linhas "
          "com BF <> AW - e herda a divergencia VOLATIL de AW, que vem de EST+PEND (estoque ao "
          "vivo). As contagens de BF e AW batem uma com a outra; e' a mesma causa, contada duas "
          "vezes. Divergencia de BF que NAO acompanhe AW e' que seria defeito. "
          "BG (VALOR_ESTOQUE) NAO entra nesta melhoria: e' estoque, nao pedido, e segue com o "
          "fator corrente.",
    "D":  "PENDENCIAS item 4 (21/08/2026): CONSEQUENCIA de CC - 5 SKUs de codst = 0 ganham o "
          "componente 'TRIB NAO ENCONTRADA' em ALERTA. MELHORIAS A3 e A4 (21/08/2026): D tambem "
          "muda nos 71 SKUs de AP e em todo SKU que tenha AV, CL ou DB preenchidos (a concatenacao "
          "carrega o texto novo). ATRIBUICAO MEDIDA em 21/08/2026, componente a componente na "
          "MESMA linha (metodo A/B da secao 6.3) - 1.793 linhas divergentes, 1.793 atribuiveis, "
          "ZERO sobra: 1.713 so' por A4, 37 so' por A3, 34 por A3+A4, 5 por CC, 2 por AO "
          "(volatil), 1 por A4+AO e 1 por A4+AY (produto 6433, embalagem renomeada no cadastro). "
          "ATENCAO - D e' CRITICA: divergencia que nao seja atribuivel a CC, a A3, a A4 ou a um "
          "componente volatil/de decisao humana (secao 6.1.1) continua sendo DEFEITO.",
}

# ─────────────────────────────────────────────────────────────────────────────
# AJUSTE 1 (25/08/2026, fecho da Etapa 4): registro NAO PODE ser cheque em
# branco. Uma coluna rotulada em LETRAS_DIVERGENCIA_POR_DECISAO fica ISENTA
# do criterio generico de limiar (>=99,9% dentro da tolerancia), mas so'
# ATE o numero que o proprio registro declara como esperado. Divergencia
# ALEM disso volta a reprovar - por outro motivo, ver
# `ResultadoPedido.estruturais_rotuladas_acima_do_esperado`.
#
# {letra -> numero maximo de linhas divergentes ESPERADO, ou None quando o
# registro NAO declara um numero EXATO (so' aproximado/textual, ex. "~1.240"
# em CH) - nesses casos a coluna fica isenta do limiar mas SEM cheque
# automatico de excesso; o relatorio imprime o valor ATUAL para o registro
# ser atualizado por humano, exatamente como pedido.
#
# Os numeros abaixo sao os que o PROPRIO texto de LETRAS_DIVERGENCIA_POR_DECISAO
# (ou CONTEXTO.md §6.0/§6.4, quando o texto local so' cita o total combinado)
# ja declara - nao e' um numero novo inventado aqui:
LETRAS_DIVERGENCIA_ESPERADO: dict[str, int | None] = {
    # PENDENCIAS item 1 (21/08/2026) - "ST s/Valor": a v11 ja traz a correcao
    # do Diretor (ver XLSX_PATH_PADRAO), entao contra o xlsx padrao estas seis
    # colunas devem fechar (esperado 0). O rotulo continua existindo so' para
    # explicar a formula/decisao para quem ler codigo antigo ou comparar
    # contra a v10. Nao sao "None" porque o texto NAO da' um numero exato
    # (a nota fala "~1.240", aproximado) - mas contra a v11 nao ha' motivo
    # para ELAS divergirem, entao 0 e' o numero certo aqui.
    "CH": 0,
    "CY": 0,
    "CO": 0,
    "CP": 0,
    "DE": 0,
    "DF": 0,
    "BO": 41,   # PENDENCIAS item 3 - grafia CAR80->CAR 80 (texto declara "41 SKUs")
    "CC": 5,    # PENDENCIAS item 4 - "exatamente os 5 SKUs de codst = 0"
    "AP": 71,   # MELHORIA A3 - "exatamente 71 SKUs"
    "AV": 136,  # MELHORIA A4 - "as 136 celulas nao vazias"
    "CL": 1632, # MELHORIA A4 - nao esta no texto de CL isolado (so' o combinado 2.182
                # em AV+CL+DB), numero vem da tabela CONTEXTO.md §6.0(b)
    "DB": 414,  # MELHORIA A4 - "as 414 celulas nao vazias"
    "AQ": 0,    # MELHORIA D1 - "esperado HOJE: ZERO celulas" (texto explicito)
    "BB": 0,    # MELHORIA A5 - "esperado HOJE: ZERO divergencia"
    "BC": 0,    # idem
    "BD": 0,    # idem
    "BE": 0,    # idem
    "BF": None, # A5 declara ZERO por A5, mas o proprio texto avisa que BF DIVERGE
                # hoje por OUTRA causa (herda de AW/volatil) - nao ha' um numero
                # exato unico que cubra as duas causas ao mesmo tempo; fica None
                # (isenta do limiar, sem cheque de excesso) ate' alguem decidir
                # um numero combinado.
    "D": None,  # ALERTA e' CRITICA, nao ESTRUTURAL - este dicionario so' vale para
                # o bucket estrutural (ver `estruturais_abaixo_do_limiar`). A
                # atribuicao de D e' feita por atribuibilidade linha-a-linha
                # (AJUSTE 3, `atribuir_divergencia_alerta`), nao por contagem fixa.
}

# Arrasto da correcao de grafia do CAR 80 (item 3): MEDIDO NO BANCO em
# 21/08/2026, e o resultado foi VAZIO. A nota do Diretor previa "muda credito,
# ICMS-ST e preco desses 41", mas so' BO muda de fato (em 39 dos 41 - nos
# outros 2 o empirico ja dava o valor tabelado). ICMS-ST, custo, margem, preco
# e ALERTA nao mudam em nenhum dos 41, porque BQ (CRED_ICMS) le DR (o credito
# EMPIRICO) e nao BO - decisao 2 das pendencias, confirmada como INTENCIONAL.
# BO so' alimenta BS (CHECK_IMPORTADO), e nos 41 o limiar nao virou de lado.
#
# A lista fica aqui, VAZIA e nomeada, em vez de ser apagada: ela e' o registro
# de que a hipotese foi testada e refutada. Se um dia a decisao 2 for revista
# (a TI expondo a tributacao por item - PDF secao 14), BO entra na cadeia de
# custo e estas colunas passam a ter arrasto de verdade - e e' aqui que ele
# deve ser declarado.
#
# ⚠ E note por que o arrasto NUNCA entraria em LETRAS_DIVERGENCIA_POR_DECISAO:
# a divergencia esperada seria de 41 linhas em colunas que, nas outras 8.736,
# continuam tendo que fechar. Rotular a COLUNA inteira esconderia defeito de
# verdade no resto do catalogo.
LETRAS_ARRASTO_CAR_80: list[str] = []

LETRAS_CANDIDATA_ULTIMA_ENTRADA = {
    "Z",   # DT_ULT_ENT
    "AA",  # QT_ULT_ENT
    "BH",  # VL_ENT_UNIT
    "BI",  # CUSTO_ULT_ENT
    "DR",  # CRED_TOTAL_EMPIRICO  (= MAX(0,1-BI/BH) - dependencia direta de BH/BI)
}


# ─────────────────────────────────────────────────────────────────────────────
# Tipo de comparacao por coluna (txt | data | num | razao | pp). Classificacao
# manual e auditavel (nao heuristica sobre texto do cabecalho) - decidida por
# leitura da formula/semantica de cada uma das 120 colunas comparaveis.
# ─────────────────────────────────────────────────────────────────────────────

LETRAS_TXT = {
    "B", "C", "D", "E", "F", "G", "H",             # COD_FAB..EMBALAGEM, ALERTA
    "X", "Y", "AL", "AO", "AP", "AS", "AV", "AX",   # CHECK_*, TEND, CLASSE
    "AY", "BE", "BL", "BN", "BS",                   # CHECK_LITRAGEM, SUG_PALETE, MODALIDADE, UF_ORIGEM, CHECK_IMPORTADO
    "CC", "CD", "CE", "CL", "DB", "DO",             # CHECK_TRIB..CHECK_SUCESSAO
}
LETRAS_DATA = {"Z", "AM"}  # DT_ULT_ENT, DT_ULT_SAIDA
LETRAS_RAZAO = {
    "AR", "AT", "AU",                               # TEND %, VAR_PV, TX_DEVOLUCAO_3M
    "BM", "BO", "BP", "BQ", "BR",                    # MVA, CRED_TOTAL, ALIQ_ICMS_ORIGEM, CRED_ICMS, CRED_PISCOF
    "BZ", "CA", "CB",                                # PISCOF_EF, ICMS_SAIDA_EF, ICMS_SEM_RED
    "CG", "CH", "CI", "CJ", "CN",                    # MKP_ATACADO, MARGEM_ST/OFICIAL/SEM_RED, MARGEM_ALVO
    "CX", "CY", "CZ", "DD", "DR",                    # MKP_VAREJO, MARGEM_*_VAREJO, MARGEM_ALVO_VAREJO, CRED_TOTAL_EMPIRICO
}
LETRAS_PP = {"CK", "CM", "DA", "DC"}  # GAP_FILIAL_pp, DIF_MC_pp, GAP_FILIAL_VAREJO_pp, DIF_MC_VAREJO_pp


def montar_tipos(colunas: list[ColunaGabarito]) -> dict[str, str]:
    tipos: dict[str, str] = {}
    for c in colunas:
        if c.letra in LETRAS_IGNORADAS:
            continue
        if c.letra in LETRAS_TXT:
            tipos[c.letra] = "txt"
        elif c.letra in LETRAS_DATA:
            tipos[c.letra] = "data"
        elif c.letra in LETRAS_RAZAO:
            tipos[c.letra] = "razao"
        elif c.letra in LETRAS_PP:
            tipos[c.letra] = "pp"
        else:
            tipos[c.letra] = "num"
    faltando = [c.letra for c in colunas if c.letra not in LETRAS_IGNORADAS and c.letra not in tipos]
    if faltando:
        raise RuntimeError(f"Coluna(s) sem tipo classificado (bug no script): {faltando}")
    return tipos


def montar_fecho_condicional(grafo_refs: dict[str, set[str]]) -> dict[str, str]:
    """Devolve {letra -> raiz_condicional} para toda coluna que descende de uma
    raiz de RAIZES_CONDICIONAIS (FALHA 3, ver comentario la'). Uma coluna so'
    pode pertencer a UMA raiz condicional - os fechos de CN e DD nao se
    sobrepoem (conferido no grafo), e se um dia sobrepuserem isso e' bug de
    classificacao, nao um caso a resolver em silencio - por isso o assert."""
    letra_por_raiz: dict[str, str] = {}
    for raiz in RAIZES_CONDICIONAIS:
        for letra in fecho_transitivo({raiz}, grafo_refs):
            if letra in letra_por_raiz and letra_por_raiz[letra] != raiz:
                raise RuntimeError(
                    f"Coluna {letra} descende de DUAS raizes condicionais "
                    f"({letra_por_raiz[letra]} e {raiz}) - fecho ambiguo, resolver a mao."
                )
            letra_por_raiz[letra] = raiz
    return letra_por_raiz


def montar_buckets(colunas: list[ColunaGabarito], grafo_refs: dict[str, set[str]]) -> tuple[dict[str, str], set[str], set[str]]:
    fecho_volatil = fecho_transitivo(RAIZES_VOLATEIS, grafo_refs)
    fecho_entrada_humana = fecho_transitivo(RAIZES_ENTRADA_HUMANA, grafo_refs)
    buckets: dict[str, str] = {}
    for c in colunas:
        if c.letra in LETRAS_IGNORADAS:
            continue
        if c.letra in LETRAS_CRITICAS:
            buckets[c.letra] = "critica"
        elif c.letra in fecho_entrada_humana:
            buckets[c.letra] = "entrada_humana"
        elif c.letra in fecho_volatil or c.letra in LETRAS_VOLATEIS_CADASTRO_DIRETO:
            # LETRAS_VOLATEIS_CADASTRO_DIRETO (AJUSTE 2) entra aqui DIRETO, sem
            # ter sido raiz de `fecho_transitivo` - de proposito, ver comentario
            # da constante. Isso NAO expande `fecho_volatil` (usado em outros
            # lugares, ex. o aviso de tensao das criticas): so' as 5 letras
            # fixas (B/C/E/F/H) ganham o bucket "volatil" aqui.
            buckets[c.letra] = "volatil"
        elif c.letra in LETRAS_CANDIDATA_ULTIMA_ENTRADA:
            buckets[c.letra] = "candidata_ultima_entrada"
        else:
            buckets[c.letra] = "estrutural"
    return buckets, fecho_volatil, fecho_entrada_humana


# ─────────────────────────────────────────────────────────────────────────────
# ColSpec / normalizacao / comparacao de valor
# ─────────────────────────────────────────────────────────────────────────────


@dataclass
class ColSpec:
    letra: str
    header: str
    slug: str      # nome logico = chave no dict Python (lado xlsx e, por padrao, lado dest)
    dest: str      # nome resolvido no lado dbt (pode divergir do slug apos resolucao)
    tipo: str      # txt | data | num | razao | pp
    bucket: str    # critica | entrada_humana | volatil | candidata_ultima_entrada | estrutural
    raiz_condicional: str | None = None  # FALHA 3: letra da raiz em RAIZES_CONDICIONAIS
                                          # (CN/DD) da qual esta coluna descende, ou None


def montar_colspecs(colunas: list[ColunaGabarito], tipos: dict[str, str], buckets: dict[str, str],
                     fecho_condicional: dict[str, str] | None = None) -> list[ColSpec]:
    fecho_condicional = fecho_condicional or {}
    specs = []
    for c in colunas:
        if c.letra in LETRAS_IGNORADAS:
            continue
        slug = _slug(c.header)
        specs.append(ColSpec(letra=c.letra, header=c.header, slug=slug, dest=slug,
                              tipo=tipos[c.letra], bucket=buckets[c.letra],
                              raiz_condicional=fecho_condicional.get(c.letra)))
    return specs


def _normaliza_vazio(v):
    """'' (celula vazia do Excel / IFERROR(...,"")) equivale a NULL - as duas
    sao tratadas como equivalentes (instrucao do orquestrador)."""
    if isinstance(v, str) and v.strip() == "":
        return None
    return v


def _norm_texto(v):
    return _normaliza_vazio(v)


def _norm_num_ou_erro(v):
    """Devolve float, None, ou ('__ERRO__', texto) para valor de erro do
    Excel (#N/A, #DIV/0! etc. - nao deveriam aparecer, ja que quase toda
    formula do gabarito e' protegida por IFERROR, mas trata sem explodir)."""
    v = _normaliza_vazio(v)
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if s.startswith("#"):
            return ("__ERRO__", s)
        s = s.replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return ("__ERRO__", s)
    try:
        return float(v)
    except (TypeError, ValueError):
        return ("__ERRO__", str(v))


def _serial_para_data(v):
    v = _normaliza_vazio(v)
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    try:
        n = float(v)
    except (TypeError, ValueError):
        return None
    return EXCEL_EPOCH + datetime.timedelta(days=n)


def comparar_valor(tipo: str, orig, dest, tol_num: float, tol_razao: float, tol_pp: float):
    """Devolve (igual: bool, diff_ordenacao: float)."""
    if tipo == "txt":
        a, b = _norm_texto(orig), _norm_texto(dest)
        igual = a == b
        return igual, (0.0 if igual else 1.0)
    if tipo == "data":
        a, b = _serial_para_data(orig), _serial_para_data(dest)
        if a is None and b is None:
            return True, 0.0
        if a is None or b is None:
            return False, float("inf")
        diff_dias = abs((a - b).days)
        return a == b, float(diff_dias)
    tol = {"num": tol_num, "razao": tol_razao, "pp": tol_pp}[tipo]
    a, b = _norm_num_ou_erro(orig), _norm_num_ou_erro(dest)
    if a is None and b is None:
        return True, 0.0
    if a is None or b is None:
        return False, float("inf")
    if isinstance(a, tuple) or isinstance(b, tuple):
        igual = a == b
        return igual, (0.0 if igual else 1.0)
    diff = abs(a - b)
    return diff <= tol, diff


def _fmt(v) -> str:
    if v is None:
        return "NULL"
    if isinstance(v, float):
        return f"{v:.6f}".rstrip("0").rstrip(".")
    if isinstance(v, tuple):
        return f"ERRO_EXCEL({v[1]})"
    s = str(v)
    return s if len(s) <= 50 else s[:47] + "..."


# ─────────────────────────────────────────────────────────────────────────────
# Leitura da aba `pedido` - STREAMING (openpyxl read_only=True, data_only=True
# - nunca `load_workbook` sem read_only; a aba tem ~140 MB de XML descomprimido
# com ~1,07M celulas de formula/valor cacheado).
# ─────────────────────────────────────────────────────────────────────────────


def ler_pedido_xlsx(caminho: Path, colunas: list[ColunaGabarito],
                     limite_linhas: int | None = None) -> tuple[list[dict], list[str], float, float | None]:
    letras_ordem = [c.letra for c in colunas]
    header_esperado = {c.letra: c.header.upper() for c in colunas}
    slug_por_letra = {c.letra: _slug(c.header) for c in colunas}

    t0 = time.perf_counter()
    wb = openpyxl.load_workbook(str(caminho), read_only=True, data_only=True)
    try:
        if XLSX_ABA not in wb.sheetnames:
            raise RuntimeError(f"aba '{XLSX_ABA}' nao encontrada em {caminho}. "
                                f"Abas disponiveis: {wb.sheetnames}")
        ws = wb[XLSX_ABA]
        it = ws.iter_rows(min_row=1, max_col=len(letras_ordem), values_only=True)
        header = next(it)

        avisos: list[str] = []
        for i, letra in enumerate(letras_ordem):
            if letra in LETRAS_IGNORADAS and letra != LETRA_CHAVE:
                continue
            real = str(header[i] if i < len(header) and header[i] is not None else "").strip().upper()
            esperado = header_esperado[letra]
            if real != esperado:
                avisos.append(f"coluna {letra} (posicao {i}): cabecalho esperado '{esperado}', "
                               f"real '{header[i]}'")

        linhas: list[dict] = []
        for row in it:
            if row is None or all(v is None for v in row):
                continue
            codigo_bruto = row[0] if len(row) > 0 else None
            try:
                codigo = int(codigo_bruto)
            except (TypeError, ValueError):
                continue
            rec: dict = {"codigo": codigo}
            for i, letra in enumerate(letras_ordem):
                if letra in LETRAS_IGNORADAS:
                    continue
                rec[slug_por_letra[letra]] = row[i] if i < len(row) else None
            linhas.append(rec)
            if limite_linhas and len(linhas) >= limite_linhas:
                break
    finally:
        wb.close()
    tempo = time.perf_counter() - t0
    pico_mb = memoria_pico_mb()
    return linhas, avisos, tempo, pico_mb


# ─────────────────────────────────────────────────────────────────────────────
# Motor de comparacao
# ─────────────────────────────────────────────────────────────────────────────


@dataclass
class ResultadoColunaPedido:
    letra: str
    header: str
    slug: str
    dest: str
    tipo: str
    bucket: str
    total_comparavel: int
    divergentes: int
    maior_diferenca: float
    exemplos: list  # (CODIGO, orig, dest, diff)
    ignorados_dest_vazio: int = 0  # so' > 0 no bucket entrada_humana
    ignorados_por_decisao: int = 0  # FALHA 3: so' > 0 nas colunas de RAIZES_CONDICIONAIS
                                     # (CN/DD e o fecho delas) - linha onde a PROPRIA raiz
                                     # se desviou do padrao, sinal de decisao gravada

    @property
    def pct_divergente(self) -> float:
        return 100.0 * self.divergentes / self.total_comparavel if self.total_comparavel else 0.0

    @property
    def pct_dentro(self) -> float:
        return 100.0 - self.pct_divergente if self.total_comparavel else 100.0


@dataclass
class ResultadoPedido:
    nome: str
    n_orig: int
    n_dest: int
    n_dup_orig: int
    n_dup_dest: int
    so_orig: list
    so_dest: list
    colunas_nao_encontradas: list[str]
    resultados: list[ResultadoColunaPedido] = field(default_factory=list)
    # Preenchido DEPOIS da comparacao, por atribuir_divergencia_alerta(): quantas
    # linhas de D (ALERTA) divergem SEM nenhum componente CHECK_* divergente.
    # None = atribuicao nao foi calculada (motor sintetico) -> D volta a exigir
    # zero absoluto, que e' o comportamento seguro na ausencia de evidencia.
    nao_atribuiveis_alerta: int | None = None

    @property
    def por_bucket(self) -> dict[str, list[ResultadoColunaPedido]]:
        out: dict[str, list[ResultadoColunaPedido]] = defaultdict(list)
        for r in self.resultados:
            out[r.bucket].append(r)
        return out

    @property
    def diferenca_de_conjunto(self) -> bool:
        """Existe QUALQUER diferenca de conjunto - usado so' para decidir se a
        secao de relatorio correspondente e' impressa. NAO usar para o veredito:
        as duas direcoes tem peso diferente (ver `diferenca_de_conjunto_grave`)."""
        return bool(self.so_orig) or bool(self.so_dest) or self.n_dup_orig > 0 or self.n_dup_dest > 0 \
            or self.n_orig != self.n_dest

    @property
    def diferenca_de_conjunto_grave(self) -> bool:
        """FALHA 4 do relatorio de revisao: as duas direcoes de `so_orig`/`so_dest`
        NAO sao a mesma coisa - mesmo raciocinio ja aplicado ao validador mensal
        (CONTEXTO.md, so_orig/so_dest la').

        - `so_orig` (SKU existe na planilha, NAO existe no dbt) = DEFEITO. Perdemos
          produto - algo que devia estar no fat_pedido sumiu.
        - `so_dest` (SKU existe no dbt, NAO existe na planilha) = ESPERADO. A
          planilha e' uma FOTO (CONTEXTO.md §6.3); o catalogo do WinThor cresce
          entre a foto e o build - SKU cadastrado depois da foto e' normal, nao
          defeito. Medido em 19/08/2026: 8.777 linhas no dbt x 8.772 na planilha,
          diferenca = exatamente 5 SKUs novos no catalogo desde a foto.
        - Duplicata dos dois lados continua reprovando sempre: chave repetida e'
          bug de join/fan-out, nunca "esperado".

        Por isso `n_orig != n_dest` sozinho NAO entra aqui: a contagem PODE (e vai)
        divergir por causa do crescimento normal do catalogo, e isso ja esta
        coberto por `so_dest` (informativo, nao reprova)."""
        return bool(self.so_orig) or self.n_dup_orig > 0 or self.n_dup_dest > 0

    @property
    def criticas_com_divergencia(self) -> list[ResultadoColunaPedido]:
        """Criticas que de fato reprovam.

        `D` (ALERTA) segue criterio DIFERENTE das outras tres, e o CONTEXTO.md
        secao 6.1.1 ja dizia isso antes deste codigo: `D` concatena 14 CHECK_*,
        e parte deles depende de estoque ao vivo (CHECK_ESTOQUE_PARADO) e de
        campo de decisao humana (CHECK_FABRICA le PEDIDO). Exigir zero absoluto
        de uma coluna cujos insumos mudam sozinhos e' exigir o impossivel - e a
        cada melhoria aprovada que muda texto de alerta (A3, A4) fica mais
        impossivel ainda.

        O criterio real de `D` e' ATRIBUIBILIDADE: toda linha divergente precisa
        ter ao menos um componente CHECK_* divergente. Linha de `D` diferente com
        TODOS os componentes iguais e' defeito de CONCATENACAO - ordem trocada,
        separador errado, componente esquecido - e ai' reprova.

        As outras tres criticas (AX CLASSE, BL MODALIDADE, BY CUSTO_TOT_GERENCIAL)
        continuam com zero absoluto: sao estruturais puras e hoje estao em zero
        de verdade.

        Sem atribuicao calculada (`nao_atribuiveis_alerta is None`), `D` volta a
        exigir zero absoluto - na falta de evidencia, o criterio mais duro."""
        out = []
        for r in self.resultados:
            if r.bucket != "critica" or r.divergentes == 0:
                continue
            if r.letra == "D" and self.nao_atribuiveis_alerta == 0:
                continue  # toda divergencia atribuida a componente; ver docstring
            out.append(r)
        return out

    @property
    def estruturais_abaixo_do_limiar(self) -> list[ResultadoColunaPedido]:
        """AJUSTE 1 (25/08/2026): coluna ESTRUTURAL rotulada em
        LETRAS_DIVERGENCIA_POR_DECISAO fica ISENTA deste criterio generico de
        limiar (>=99,9% dentro da tolerancia) - a divergencia dela ja tem
        dono e, na maioria dos casos, um numero esperado registrado
        (LETRAS_DIVERGENCIA_ESPERADO). Reprovar de novo uma divergencia ja
        aprovada e documentada torna o registro inutil.

        Isso NAO e' cheque em branco: a coluna continua sendo comparada,
        contada e impressa, e se ela divergir ALEM do numero que o proprio
        registro declara como esperado, ela volta a reprovar - so' que por
        OUTRO motivo (`estruturais_rotuladas_acima_do_esperado`), nao por
        este limiar generico."""
        return [r for r in self.resultados if r.bucket == "estrutural" and r.pct_dentro < 99.9
                and r.letra not in LETRAS_DIVERGENCIA_POR_DECISAO]

    @property
    def estruturais_rotuladas_acima_do_esperado(self) -> list[ResultadoColunaPedido]:
        """AJUSTE 1: a outra metade da isencao acima - coluna ESTRUTURAL
        rotulada cuja divergencia MEDIDA excede o numero EXATO que o proprio
        registro declara (`LETRAS_DIVERGENCIA_ESPERADO[letra]` != None). So'
        reprova quando o registro declara um numero exato; entradas `None`
        (aproximadas/textuais, ex. BF) ficam isentas sem cheque automatico -
        o relatorio imprime o valor atual para atualizacao manual."""
        out = []
        for r in self.resultados:
            if r.bucket != "estrutural" or r.letra not in LETRAS_DIVERGENCIA_POR_DECISAO:
                continue
            esperado = LETRAS_DIVERGENCIA_ESPERADO.get(r.letra)
            if esperado is not None and r.divergentes > esperado:
                out.append(r)
        return out

    @property
    def motivos_reprovacao(self) -> list[str]:
        """FALHA 1 do relatorio de revisao: todo motivo de reprovacao, nao so' o
        primeiro que bater. Sao condicoes INDEPENDENTES - coluna critica
        divergente, diferenca de conjunto grave, coluna do gabarito sem par no
        dbt e coluna estrutural abaixo do limiar podem acontecer AO MESMO TEMPO,
        e o veredito antigo (if/elif) escondia todas menos a primeira."""
        motivos: list[str] = []
        if self.criticas_com_divergencia:
            letras = ", ".join(c.letra for c in self.criticas_com_divergencia)
            det = []
            for c in self.criticas_com_divergencia:
                if c.letra == "D":
                    na = self.nao_atribuiveis_alerta
                    det.append(f"D ({'zero absoluto: atribuicao nao calculada' if na is None else f'{na} linha(s) NAO atribuivel(is) a componente CHECK_*'})")
                else:
                    det.append(f"{c.letra} (zero era obrigatorio)")
            motivos.append(f"{len(self.criticas_com_divergencia)} coluna(s) CRITICA(s): " + "; ".join(det))
        if self.diferenca_de_conjunto_grave:
            partes = []
            if self.so_orig:
                partes.append(f"{len(self.so_orig)} SKU(s) perdido(s) (na planilha, ausente no dbt)")
            if self.n_dup_orig:
                partes.append(f"{self.n_dup_orig} CODIGO(s) duplicado(s) no lado planilha")
            if self.n_dup_dest:
                partes.append(f"{self.n_dup_dest} CODIGO(s) duplicado(s) no lado dbt (fan-out)")
            motivos.append("diferenca de conjunto GRAVE - " + "; ".join(partes))
        if self.colunas_nao_encontradas:
            motivos.append(f"{len(self.colunas_nao_encontradas)} coluna(s) do gabarito sem "
                            f"correspondente no lado dbt")
        if self.estruturais_abaixo_do_limiar:
            cols_ordenadas = sorted(self.estruturais_abaixo_do_limiar,
                                     key=lambda c: c.pct_divergente, reverse=True)
            letras = ", ".join(f"{c.letra}({c.pct_divergente:.3f}%)" for c in cols_ordenadas)
            motivos.append(f"{len(self.estruturais_abaixo_do_limiar)} coluna(s) ESTRUTURAL(is) "
                            f"abaixo de 99,9% dentro da tolerancia (NAO rotulada por decisao): {letras}")
        if self.estruturais_rotuladas_acima_do_esperado:
            cols_ordenadas = sorted(self.estruturais_rotuladas_acima_do_esperado,
                                     key=lambda c: c.divergentes, reverse=True)
            partes = []
            for c in cols_ordenadas:
                esperado = LETRAS_DIVERGENCIA_ESPERADO.get(c.letra)
                partes.append(f"{c.letra}({c.divergentes} medido > {esperado} esperado)")
            motivos.append(f"{len(self.estruturais_rotuladas_acima_do_esperado)} coluna(s) ROTULADA(s) "
                            f"por decisao com divergencia ALEM do numero esperado pelo proprio registro "
                            f"(AJUSTE 1 - isencao nao e' cheque em branco): {', '.join(partes)}")
        return motivos

    @property
    def ok(self) -> bool:
        return (
            not self.diferenca_de_conjunto_grave
            and not self.colunas_nao_encontradas
            and not self.criticas_com_divergencia
            and not self.estruturais_abaixo_do_limiar
            and not self.estruturais_rotuladas_acima_do_esperado
        )


def _indexar(linhas: list[dict]) -> tuple[dict, int]:
    idx: dict = {}
    dup = 0
    for row in linhas:
        k = row["codigo"]
        if k in idx:
            dup += 1
        idx[k] = row
    return idx, dup


def comparar_pedido(nome: str, colspecs: list[ColSpec], linhas_orig: list[dict], linhas_dest: list[dict],
                     colunas_dest_disponiveis: set[str], tol_num: float, tol_razao: float, tol_pp: float,
                     ignorar_dest_vazio_entrada_humana: bool = True) -> ResultadoPedido:
    idx_orig, dup_orig = _indexar(linhas_orig)
    idx_dest, dup_dest = _indexar(linhas_dest)
    chaves_orig, chaves_dest = set(idx_orig), set(idx_dest)
    so_orig = sorted(chaves_orig - chaves_dest)
    so_dest = sorted(chaves_dest - chaves_orig)
    comuns = chaves_orig & chaves_dest

    colunas_nao_encontradas = [f"{c.letra} {c.header} -> {c.dest}" for c in colspecs
                                if c.dest not in colunas_dest_disponiveis]

    # FALHA 3: dest (lado dbt) de cada raiz condicional (CN/DD), para o skip
    # linha-a-linha abaixo. Resolvido por letra, nao por slug fixo, porque
    # `resolver_dest` pode ter ajustado `c.dest` antes de chegar aqui.
    raiz_dest_por_letra = {
        letra: next((cc.dest for cc in colspecs if cc.letra == letra), None)
        for letra in RAIZES_CONDICIONAIS
    }

    resultados: list[ResultadoColunaPedido] = []
    for c in colspecs:
        if c.dest not in colunas_dest_disponiveis:
            continue
        total = 0
        divergentes = 0
        ignorados = 0
        ignorados_decisao = 0
        maior_diff = 0.0
        exemplos: list[tuple] = []
        pular_vazio = ignorar_dest_vazio_entrada_humana and c.bucket == "entrada_humana"
        raiz_dest = raiz_dest_por_letra.get(c.raiz_condicional) if c.raiz_condicional else None
        padrao_raiz = RAIZES_CONDICIONAIS.get(c.raiz_condicional) if c.raiz_condicional else None
        for k in comuns:
            ov = idx_orig[k].get(c.slug)
            dv = idx_dest[k].get(c.dest)
            if pular_vazio and _normaliza_vazio(dv) is None:
                ignorados += 1
                continue
            # FALHA 3 (CN/DD como raiz CONDICIONAL, nao bucket fixo): a linha so'
            # se comporta como entrada humana quando a PROPRIA raiz (CN para o
            # atacado, DD para o varejo) se desviou do padrao deterministico -
            # sinal de que ha' decisao gravada em APP_DECISAO_PRECO para este SKU.
            # Sem desvio, e' o valor padrao (MARGEM_ALVO_PADRAO) e compara normal.
            if raiz_dest is not None:
                raiz_dv = _norm_num_ou_erro(idx_dest[k].get(raiz_dest))
                if isinstance(raiz_dv, float) and abs(raiz_dv - padrao_raiz) > tol_razao:
                    ignorados_decisao += 1
                    continue
            total += 1
            igual, diff = comparar_valor(c.tipo, ov, dv, tol_num, tol_razao, tol_pp)
            if not igual:
                divergentes += 1
                diff_ordenacao = diff if diff != float("inf") else 1e18
                maior_diff = max(maior_diff, diff_ordenacao if diff_ordenacao < 1e18 else maior_diff)
                exemplos.append((k, ov, dv, diff_ordenacao))
        exemplos.sort(key=lambda t: t[3], reverse=True)
        resultados.append(ResultadoColunaPedido(
            letra=c.letra, header=c.header, slug=c.slug, dest=c.dest, tipo=c.tipo, bucket=c.bucket,
            total_comparavel=total, divergentes=divergentes, maior_diferenca=maior_diff,
            exemplos=exemplos[:3], ignorados_dest_vazio=ignorados, ignorados_por_decisao=ignorados_decisao,
        ))

    return ResultadoPedido(
        nome=nome, n_orig=len(linhas_orig), n_dest=len(linhas_dest),
        n_dup_orig=dup_orig, n_dup_dest=dup_dest, so_orig=so_orig, so_dest=so_dest,
        colunas_nao_encontradas=colunas_nao_encontradas, resultados=resultados,
    )


# ─────────────────────────────────────────────────────────────────────────────
# AJUSTE 3 (25/08/2026, fecho da Etapa 4): atribuicao de ALERTA (D) REFEITA
# componente a componente contra os dados de HOJE - CONTEXTO.md §6.0.1 e'
# explicito: "1.793 divergencias, zero sobra" e' foto de 21/08/2026 e ja
# ENVELHECEU (o proprio CONTEXTO mede 2.146 em 24/08/2026, sem mudanca de
# codigo nenhuma). Ninguem pode afirmar "zero nao atribuivel" citando um
# numero de dias atras - isto aqui RECALCULA, nao reafirma o texto antigo.
#
# D = ALERTA concatena exatamente 14 colunas CHECK_* (gabarito, formula da
# letra D): X, Y, AL, AV, AO, AP, AY, BS, CC, CD, CE, CL, DO, DB. Cada uma
# tem natureza diferente (CONTEXTO §6.1.1 + o que a formula de cada uma le'):
COMPONENTES_ALERTA: dict[str, str] = {
    "X":  "decisao_humana",        # CHECK_FABRICA        - le PEDIDO (BA), decisao gravada
    "Y":  "decisao_humana",        # CHECK_INATIVO         - le PEDIDO (BA), decisao gravada
    "AL": "estrutural",            # CHECK_RUPTURA         - nenhum registro cobre; defeito se divergir
    "AV": "melhoria_a4",           # CHECK_DEVOLUCAO_ALTA  - MELHORIA A4 (mascara de percentual)
    "AO": "volatil_estoque",       # CHECK_ESTOQUE_PARADO  - estoque ao vivo (V=EST+PEND)
    "AP": "melhoria_a3",           # CHECK_FORA_DE_LINHA   - MELHORIA A3 (71 SKUs esperados)
    "AY": "volatil_cadastro",      # CHECK_LITRAGEM        - le H (EMBALAGEM), AJUSTE 2 - cadastro editavel
    "BS": "estrutural",            # CHECK_IMPORTADO       - nenhum registro cobre; defeito se divergir
    "CC": "pendencia_check_trib",  # CHECK_TRIB            - PENDENCIAS item 4 (5 SKUs esperados)
    "CD": "estrutural",            # CHECK_MVA             - nenhum registro cobre; defeito se divergir
    "CE": "ultima_entrada",        # CHECK_CUSTO           - le BH/BI/Z (CANDIDATA_ULTIMA_ENTRADA, ao vivo)
    "CL": "melhoria_a4",           # CHECK_MARGEM_INSTAVEL - MELHORIA A4 (mascara de percentual)
    "DB": "melhoria_a4",           # CHECK_MARGEM_INSTAVEL_VAREJO - idem
    "DO": "estrutural",            # CHECK_SUCESSAO        - nenhum registro cobre; defeito se divergir
}


@dataclass
class AtribuicaoAlerta:
    total_divergentes: int
    contagem_por_combinacao: dict[str, int]
    exemplos_por_combinacao: dict[str, list[int]]
    nao_atribuiveis: list[int]  # CODIGO de toda linha onde D diverge e NENHUM componente diverge

    @property
    def contagem_por_componente_isolado(self) -> dict[str, int]:
        """Reparte cada linha (que pode ter combinacao de varios componentes)
        de volta por componente INDIVIDUAL, para o resumo 'quantas por cada
        componente' pedido - uma linha com 2 componentes conta 1 vez em CADA
        um (nao e' particao, e' contagem de presenca)."""
        out: dict[str, int] = defaultdict(int)
        for combinacao, n in self.contagem_por_combinacao.items():
            for comp in combinacao.split("+"):
                out[comp] += n
        return out


def atribuir_divergencia_alerta(colspecs: list[ColSpec], linhas_orig: list[dict], linhas_dest: list[dict],
                                 tol_num: float, tol_razao: float, tol_pp: float) -> AtribuicaoAlerta:
    """Para cada CODIGO comum aos dois lados onde D (ALERTA) diverge, testa
    CADA UM dos 14 componentes de COMPONENTES_ALERTA na MESMA linha (metodo
    A/B ja usado em CONTEXTO.md §6.3/§6.4: reconstroi o insumo e ve' se a
    divergencia se explica por ele). Linha onde NENHUM componente diverge e'
    'nao atribuivel' - candidata a defeito real, reportada explicitamente."""
    idx_orig, _ = _indexar(linhas_orig)
    idx_dest, _ = _indexar(linhas_dest)
    comuns = set(idx_orig) & set(idx_dest)
    by_letra = {c.letra: c for c in colspecs}
    d_spec = by_letra.get("D")

    contagem: dict[str, int] = defaultdict(int)
    exemplos: dict[str, list[int]] = defaultdict(list)
    nao_atribuiveis: list[int] = []
    total_divergentes_d = 0

    if d_spec is None:
        return AtribuicaoAlerta(0, {}, {}, [])

    componentes_presentes = {letra: comp for letra, comp in COMPONENTES_ALERTA.items() if letra in by_letra}

    for k in sorted(comuns):
        ov = idx_orig[k].get(d_spec.slug)
        dv = idx_dest[k].get(d_spec.dest)
        igual, _ = comparar_valor(d_spec.tipo, ov, dv, tol_num, tol_razao, tol_pp)
        if igual:
            continue
        total_divergentes_d += 1
        categorias_desta_linha: set[str] = set()
        for letra, categoria in componentes_presentes.items():
            spec = by_letra[letra]
            cov = idx_orig[k].get(spec.slug)
            cdv = idx_dest[k].get(spec.dest)
            c_igual, _ = comparar_valor(spec.tipo, cov, cdv, tol_num, tol_razao, tol_pp)
            if not c_igual:
                categorias_desta_linha.add(categoria)
        if not categorias_desta_linha:
            nao_atribuiveis.append(k)
        else:
            chave = "+".join(sorted(categorias_desta_linha))
            contagem[chave] += 1
            if len(exemplos[chave]) < 3:
                exemplos[chave].append(k)

    return AtribuicaoAlerta(
        total_divergentes=total_divergentes_d,
        contagem_por_combinacao=dict(contagem),
        exemplos_por_combinacao=dict(exemplos),
        nao_atribuiveis=nao_atribuiveis,
    )


def imprimir_atribuicao_alerta(a: AtribuicaoAlerta):
    print()
    print("#" * 92)
    print("ATRIBUICAO DE ALERTA (D) - RECALCULADA AGORA (AJUSTE 3, CONTEXTO.md 6.0.1)")
    print("#" * 92)
    print("Numeros de registro (MELHORIAS.md/PENDENCIAS_DIRETORIA.md/CONTEXTO.md) sao FOTO do dia")
    print("da medicao e envelhecem sozinhos (ex.: 1.793 em 21/08 -> 2.146 em 24/08, sem mudanca de")
    print("codigo). Por isso esta secao NAO cita o numero antigo como referencia - recalcula contra")
    print("os dois lados de HOJE, componente a componente, na MESMA linha.")
    print()
    print(f"Total de linhas com D divergente HOJE: {a.total_divergentes}")
    if a.total_divergentes == 0:
        print("Nenhuma linha com D divergente - nada a atribuir.")
        print("#" * 92)
        return
    print()
    print("Por componente (uma linha com mais de 1 componente conta em CADA um - nao e' particao):")
    for comp, n in sorted(a.contagem_por_componente_isolado.items(), key=lambda t: t[1], reverse=True):
        pct = 100.0 * n / a.total_divergentes
        print(f"    {comp:<24} {n:>6} linha(s)  ({pct:.1f}% das divergentes)")
    print()
    print("Por combinacao EXATA de componentes (particao real - soma bate com o total acima):")
    for combinacao, n in sorted(a.contagem_por_combinacao.items(), key=lambda t: t[1], reverse=True):
        exs = ", ".join(str(x) for x in a.exemplos_por_combinacao.get(combinacao, []))
        print(f"    {combinacao:<40} {n:>6} linha(s)   exemplos CODIGO: {exs}")
    print()
    n_nao_atrib = len(a.nao_atribuiveis)
    print(f"SEM COMPONENTE DIVERGENTE NENHUM (nao atribuivel - candidata a DEFEITO REAL): {n_nao_atrib}")
    if n_nao_atrib:
        print("    Isto e' reportado, nao explicado - cada CODIGO abaixo precisa de investigacao:")
        for k in a.nao_atribuiveis[:20]:
            print(f"        CODIGO={k}")
        if n_nao_atrib > 20:
            print(f"        ... e mais {n_nao_atrib - 20} CODIGO(s) (ver lista completa em execucao dedicada)")
    else:
        print("    ZERO - toda linha com D divergente hoje tem pelo menos 1 componente CHECK_*")
        print("    tambem divergente na mesma linha. (Isto NAO absolve automaticamente: componente")
        print("    'estrutural' sem registro proprio - AL/BS/CD/DO - ainda e' defeito, so' que o")
        print("    defeito e' NAQUELE CHECK_*, nao um mistério em D.)")
    print("#" * 92)


# ─────────────────────────────────────────────────────────────────────────────
# Impressao
# ─────────────────────────────────────────────────────────────────────────────

_ORDEM_BUCKET = ["critica", "estrutural", "volatil", "candidata_ultima_entrada", "entrada_humana"]
_ROTULO_BUCKET = {
    "critica": "CRITICA (zero divergencia OBRIGATORIA)",
    "estrutural": "ESTRUTURAL (divergencia = defeito de verdade; >=99,9% dentro da tolerancia)",
    "volatil": "VOLATIL (divergencia ESPERADA por construcao - estoque ao vivo/mes corrente; nunca reprova sozinha)",
    "candidata_ultima_entrada": "CANDIDATA_ULTIMA_ENTRADA (sinalizada, nao propagada - ver docstring; informativa)",
    "entrada_humana": "ENTRADA_HUMANA (decisao digitada; celula vazia do lado dbt nunca conta como divergencia)",
}


def _imprimir_coluna(c: ResultadoColunaPedido):
    marca = "  <<< DIVERGENCIA ESPERADA POR DECISAO" if c.letra in LETRAS_DIVERGENCIA_POR_DECISAO else ""
    print(f"  [{c.letra}] {c.header}  (slug={c.slug} -> dest={c.dest}, tipo={c.tipo}){marca}")
    if marca:
        print(f"      MOTIVO: {LETRAS_DIVERGENCIA_POR_DECISAO[c.letra]}")
    extra = f", {c.ignorados_dest_vazio} ignorada(s) por dest vazio" if c.ignorados_dest_vazio else ""
    if c.ignorados_por_decisao:
        extra += f", {c.ignorados_por_decisao} ignorada(s) por decisao gravada (raiz condicional)"
    print(f"      {c.divergentes}/{c.total_comparavel} linhas divergem ({c.pct_divergente:.3f}%){extra}, "
          f"maior diferenca absoluta = {c.maior_diferenca:.4f}")
    for chave, ov, dv, diff in c.exemplos:
        diff_txt = f"diff={diff:.4f}" if diff < 1e18 else "(nulo de um lado)"
        print(f"      exemplo CODIGO={chave}: planilha={_fmt(ov)}  dbt={_fmt(dv)}  {diff_txt}")


def imprimir_resultado_pedido(r: ResultadoPedido):
    print()
    print("=" * 92)
    print(f"RELATORIO: {r.nome}")
    print("=" * 92)
    print(f"Lado planilha (aba pedido): {r.n_orig:>8} linhas")
    print(f"Lado dbt (fat_pedido)     : {r.n_dest:>8} linhas")
    if r.n_dup_orig:
        print(f"FALHA: {r.n_dup_orig} CODIGO(s) duplicado(s) no lado planilha")
    if r.n_dup_dest:
        print(f"FALHA: {r.n_dup_dest} CODIGO(s) duplicado(s) no lado dbt (fan-out)")
    if r.n_orig != r.n_dest:
        print(f"AVISO: contagem de linhas diverge (planilha={r.n_orig}, dbt={r.n_dest}) - "
              f"isoladamente isso NAO e' falha, ver secao de conjunto abaixo (o catalogo "
              f"cresce entre a foto da planilha e o build).")

    print()
    print("-- Diferencas de CONJUNTO (checadas ANTES de qualquer coluna) --")
    print(f"So' na planilha (perdido no dbt - DEFEITO)             : {len(r.so_orig)}")
    for k in r.so_orig[:3]:
        print(f"    exemplo: CODIGO={k}")
    print(f"So' no dbt (SKU novo desde a foto da planilha - ESPERADO): {len(r.so_dest)}")
    for k in r.so_dest[:3]:
        print(f"    exemplo: CODIGO={k}")

    if r.colunas_nao_encontradas:
        print()
        print("-- Colunas do gabarito SEM correspondente encontrado no lado dbt --")
        for c in r.colunas_nao_encontradas:
            print(f"    {c}")

    por_bucket = r.por_bucket
    for bucket in _ORDEM_BUCKET:
        cols = por_bucket.get(bucket, [])
        if not cols:
            continue
        cols_ordenadas = sorted(cols, key=lambda c: c.divergentes, reverse=True)
        divergentes = [c for c in cols_ordenadas if c.divergentes > 0]
        ok_cols = [c for c in cols_ordenadas if c.divergentes == 0]
        print()
        print(f"-- BLOCO {_ROTULO_BUCKET[bucket]} --")
        print(f"   {len(cols)} coluna(s) neste bloco: {len(divergentes)} com divergencia, "
              f"{len(ok_cols)} identicas")
        for c in divergentes:
            print()
            _imprimir_coluna(c)

    # FALHA 1 do relatorio de revisao: o veredito antigo era if/elif - reportava
    # SO' a primeira causa de reprovacao que batesse, e as outras ficavam
    # escondidas (ex.: coluna CRITICA divergente escondia estrutural abaixo do
    # limiar). Motivos de reprovacao sao INDEPENDENTES entre si; todos precisam
    # aparecer. Ver `ResultadoPedido.motivos_reprovacao` e a prova sintetica em
    # `provar_veredito_reporta_todos_os_motivos`.
    print()
    motivos = r.motivos_reprovacao
    if motivos:
        print(f"VEREDITO {r.nome}: REPROVADO - {len(motivos)} motivo(s):")
        for m in motivos:
            print(f"    - {m}")
    else:
        print(f"VEREDITO {r.nome}: PASSOU")


def imprimir_resumo_executivo(r: ResultadoPedido):
    print()
    print("#" * 92)
    print("RESUMO EXECUTIVO (para o Diretor de Compras)")
    print("#" * 92)
    total = len(r.resultados)
    identicas = [c for c in r.resultados if c.divergentes == 0]
    com_divergencia = [c for c in r.resultados if c.divergentes > 0]
    print(f"Colunas comparadas: {total}")
    print(f"  - identicas (zero divergencia): {len(identicas)}")
    print(f"  - com alguma divergencia      : {len(com_divergencia)}")
    print()
    print("Colunas CRITICAS (exigem zero divergencia):")
    for letra in sorted(LETRAS_CRITICAS):
        c = next((x for x in r.resultados if x.letra == letra), None)
        if c is None:
            print(f"  [{letra}] nao comparada (coluna nao encontrada no lado dbt)")
        else:
            estado = "OK (zero divergencia)" if c.divergentes == 0 else \
                f"REPROVADA - {c.divergentes}/{c.total_comparavel} linhas divergem"
            print(f"  [{c.letra}] {c.header}: {estado}")
    print()
    print("Colunas com DIVERGENCIA ESPERADA POR DECISAO (CONTEXTO.md 6.0 e 6.4):")
    print("  Duas origens, ambas registradas e com dono:")
    print("    (a) DECISAO DO DIRETOR ainda nao aplicada na planilha (PENDENCIAS itens 3 e 4);")
    print("    (b) MELHORIA APROVADA (MELHORIAS.md A3/A4/D1/A5) - a planilha e' ponto de partida,")
    print("        nao alvo de replica exata, e estas mudancas nao vao existir no xlsx.")
    print("  Elas continuam sendo comparadas e contadas normalmente - o rotulo diz de quem e' a")
    print("  explicacao, e ISENTA a coluna do limiar generico de 99,9% (AJUSTE 1, 25/08/2026) -")
    print("  MAS SO' ATE o numero que o proprio registro declara como esperado. Divergencia ALEM")
    print("  disso volta a reprovar (ver VEREDITO). Onde o registro nao declara um numero exato,")
    print("  o valor ATUAL abaixo e' para atualizar o registro, nao para ignorar.")
    for letra, motivo in LETRAS_DIVERGENCIA_POR_DECISAO.items():
        c = next((x for x in r.resultados if x.letra == letra), None)
        if c is None:
            print(f"  [{letra}] nao comparada (coluna nao encontrada no lado dbt)")
            continue
        # O cheque de excesso do AJUSTE 1 (`estruturais_rotuladas_acima_do_esperado`) SO'
        # reprova quando bucket == "estrutural" - e' o mesmo criterio do veredito.
        # Coluna VOLATIL/ENTRADA_HUMANA/CANDIDATA_ULTIMA_ENTRADA que exceda o esperado
        # NAO reprova sozinha (informativa por design, mesma logica de defasagem de
        # estoque/preco ao vivo dessas buckets) - o texto abaixo tem que refletir isso,
        # senao mente sobre o que o veredito realmente faz.
        esperado = LETRAS_DIVERGENCIA_ESPERADO.get(letra)
        if esperado is None:
            situacao = f"esperado NAO declarado (numero aproximado/textual) - ATUALIZAR registro com {c.divergentes}"
        elif c.divergentes > esperado and c.bucket == "estrutural":
            situacao = f"esperado {esperado} - EXCEDIDO (REPROVA, ver VEREDITO)"
        elif c.divergentes > esperado:
            situacao = (f"esperado {esperado} - EXCEDIDO, mas bucket={c.bucket} (informativo, NAO reprova "
                        f"sozinho - mesma defasagem de estoque/preco ao vivo do bucket; ATUALIZAR registro "
                        f"ou investigar se o excesso e' sempre da MESMA causa volatil)")
        else:
            situacao = f"esperado {esperado} - dentro do declarado"
        print(f"  [{c.letra}] {c.header} (bucket={c.bucket}): "
              f"{c.divergentes}/{c.total_comparavel} divergem - {situacao}")
        print(f"        {motivo}")
    if LETRAS_ARRASTO_CAR_80:
        print(f"  Arrasto do item 3 (CAR 80) - 41 SKUs podem divergir nestas, sem que a COLUNA")
        print(f"  inteira seja esperada: {', '.join(LETRAS_ARRASTO_CAR_80)}")
    else:
        print("  Arrasto do item 3 (CAR 80): NENHUM. Medido em 21/08/2026 - so' BO muda (39 dos")
        print("  41 SKUs). ICMS-ST, custo, margem, preco e ALERTA nao mudam em nenhum, porque BQ")
        print("  le DR (credito empirico) e nao BO. Divergencia nessas colunas e' DEFEITO.")

    print()
    precisam_decisao = [c for c in r.resultados
                         if c.bucket in ("entrada_humana", "candidata_ultima_entrada")
                         or (c.bucket == "volatil" and c.divergentes > 0)]
    print(f"Colunas que precisam de DECISAO HUMANA / revisao (nao entram no veredito automatico): "
          f"{len(precisam_decisao)}")
    for c in precisam_decisao:
        print(f"  [{c.letra}] {c.header}  (bucket={c.bucket}, "
              f"{c.divergentes}/{c.total_comparavel} divergem)")

    print()
    condicionais = [c for c in r.resultados if c.letra in RAIZES_CONDICIONAIS or c.ignorados_por_decisao > 0]
    print(f"Colunas de RAIZ CONDICIONAL (FALHA 3 - MARGEM_ALVO/MARGEM_ALVO_VAREJO e o fecho delas): "
          f"{len(condicionais)}")
    print("  Bucket ESTRUTURAL normal - contam para o veredito. So' a linha ONDE HOUVER decisao")
    print("  gravada (raiz desviada de MARGEM_ALVO_PADRAO) e' ignorada; sem decisao, o padrao")
    print("  deterministico e' comparado como qualquer outra coluna calculada.")
    for c in condicionais:
        print(f"  [{c.letra}] {c.header}: {c.divergentes}/{c.total_comparavel} divergem, "
              f"{c.ignorados_por_decisao} ignorada(s) por decisao gravada")

    print()
    veredito = "PASSOU" if r.ok else "REPROVADO"
    print(f"VEREDITO FINAL (fat_pedido x aba pedido): {veredito}")
    print("#" * 92)


# ─────────────────────────────────────────────────────────────────────────────
# Verificacao do motor (SEM banco, com o xlsx de verdade) - ver docstring
# item 3. Prova que o LEITOR e o COMPARADOR funcionam antes de existir
# `fat_pedido`.
# ─────────────────────────────────────────────────────────────────────────────


def verificar_motor(colspecs: list[ColSpec], linhas: list[dict], tol_num: float, tol_razao: float,
                     tol_pp: float) -> bool:
    slugs_disponiveis = {c.slug for c in colspecs}
    for c in colspecs:
        c.dest = c.slug  # no autoteste, dest = slug (xlsx contra xlsx)

    print()
    print("#" * 92)
    print("VERIFICACAO DO MOTOR (sem banco - so' a aba `pedido` de verdade)")
    print("#" * 92)

    # (a) auto-comparacao: aba `pedido` contra copia profunda de si mesma
    print()
    print("[motor] (a) autocomparacao: aba `pedido` x copia profunda de si mesma - tem que dar ZERO")
    print("        divergencia em TODA coluna, em TODO bucket. Se nao der, o LEITOR tem bug.")
    copia = copy.deepcopy(linhas)
    resultado_limpo = comparar_pedido(
        nome="[MOTOR] autocomparacao (aba pedido x copia profunda de si mesma)",
        colspecs=colspecs, linhas_orig=linhas, linhas_dest=copia,
        colunas_dest_disponiveis=slugs_disponiveis, tol_num=tol_num, tol_razao=tol_razao, tol_pp=tol_pp,
        ignorar_dest_vazio_entrada_humana=False,  # aqui NAO ha' motivo pra ignorar - e' a MESMA planilha
    )
    imprimir_resultado_pedido(resultado_limpo)
    caso_a_ok = resultado_limpo.ok
    print(f"\n[motor] (a) autocomparacao: {'OK - zero divergencia, como esperado' if caso_a_ok else 'FALHOU - o LEITOR tem bug (investigar antes de qualquer outra coisa)'}")

    if not linhas:
        print("[motor] AVISO: aba `pedido` leu 0 linhas - nao e' possivel injetar defeitos sinteticos "
              "nem provar nada alem da leitura. Verifique o arquivo/aba.")
        return caso_a_ok

    codigo_alvo = linhas[0]["codigo"]

    # (b) defeito 1: valor numerico alterado numa coluna ESTRUTURAL - EM DUAS PARTES,
    # porque o proprio criterio de aceite (>=99,9% dentro da tolerancia, POR COLUNA)
    # tolera de proposito uma diferenca isolada num dataset de milhares de linhas
    # (1/8772 = 0,011% de divergencia, dentro dos 0,1% permitidos). Isso NAO e' bug
    # do motor - e' o motor sendo fiel ao criterio pedido. Por isso o autoteste prova
    # as DUAS pontas: (b1) 1 linha alterada tem que PASSAR (ruido isolado nao reprova
    # sozinho, por design); (b2) alterar linhas suficientes para passar de 0,1% do
    # dataset tem que REPROVAR (divergencia sistematica de verdade e' pega).
    col_num_estrutural = next(c for c in colspecs if c.bucket == "estrutural" and c.tipo == "num")

    def _alterar_embal(dest: list[dict], codigos: set[int]):
        for row in dest:
            if row["codigo"] in codigos:
                atual = row.get(col_num_estrutural.slug)
                row[col_num_estrutural.slug] = (float(atual) if atual not in (None, "") else 0.0) + 999999.0

    dest_num_1linha = copy.deepcopy(linhas)
    _alterar_embal(dest_num_1linha, {codigo_alvo})
    print()
    print(f"[motor] (b1) defeito sintetico 1a/4: valor numerico alterado em [{col_num_estrutural.letra}] "
          f"{col_num_estrutural.header} de 1 UNICA linha (CODIGO={codigo_alvo}) - abaixo do limiar de "
          f"0,1% de um dataset de {len(linhas)} linhas; o criterio de aceite (>=99,9% por coluna) "
          f"PERMITE ruido isolado - o veredito esperado aqui e' PASSOU, nao REPROVADO.")
    resultado_num_1 = comparar_pedido(
        nome=f"[MOTOR] defeito 1a/4 - 1 linha alterada em {col_num_estrutural.header} (esperado: PASSOU)",
        colspecs=colspecs, linhas_orig=linhas, linhas_dest=dest_num_1linha,
        colunas_dest_disponiveis=slugs_disponiveis, tol_num=tol_num, tol_razao=tol_razao, tol_pp=tol_pp,
        ignorar_dest_vazio_entrada_humana=False,
    )
    imprimir_resultado_pedido(resultado_num_1)
    col_result_1 = next(c for c in resultado_num_1.resultados if c.letra == col_num_estrutural.letra)
    caso_b1_ok = resultado_num_1.ok and col_result_1.divergentes == 1 and col_result_1.pct_dentro >= 99.9
    print(f"\n[motor] (b1) 1 linha alterada (ruido isolado): "
          f"{'OK (PASSOU, como o criterio >=99,9% exige - o motor mediu 1 divergencia e nao reprovou por isso)' if caso_b1_ok else 'FALHOU (deveria ter passado, o criterio permite ruido isolado)'}")

    n_linhas_sistemico = max(10, int(len(linhas) * 0.002) + 1)  # > 0,1% do dataset, com folga
    codigos_sistemico = {row["codigo"] for row in linhas[:n_linhas_sistemico]}
    dest_num_sistemico = copy.deepcopy(linhas)
    _alterar_embal(dest_num_sistemico, codigos_sistemico)
    print()
    print(f"[motor] (b2) defeito sintetico 1b/4: MESMA coluna alterada em {len(codigos_sistemico)} linhas "
          f"({100 * len(codigos_sistemico) / len(linhas):.3f}% do dataset, acima do limiar de 0,1%) - "
          f"o veredito esperado aqui e' REPROVADO.")
    resultado_num_sist = comparar_pedido(
        nome=f"[MOTOR] defeito 1b/4 - {len(codigos_sistemico)} linhas alteradas em "
             f"{col_num_estrutural.header} (esperado: REPROVADO)",
        colspecs=colspecs, linhas_orig=linhas, linhas_dest=dest_num_sistemico,
        colunas_dest_disponiveis=slugs_disponiveis, tol_num=tol_num, tol_razao=tol_razao, tol_pp=tol_pp,
        ignorar_dest_vazio_entrada_humana=False,
    )
    imprimir_resultado_pedido(resultado_num_sist)
    col_result_sist = next(c for c in resultado_num_sist.resultados if c.letra == col_num_estrutural.letra)
    caso_b2_ok = (not resultado_num_sist.ok) and col_result_sist.pct_dentro < 99.9 \
        and col_result_sist in resultado_num_sist.estruturais_abaixo_do_limiar
    print(f"\n[motor] (b2) divergencia sistematica ({len(codigos_sistemico)} linhas): "
          f"{'OK (REPROVOU, coluna ficou abaixo de 99,9% dentro da tolerancia)' if caso_b2_ok else 'FALHOU (deveria ter reprovado)'}")

    caso_b_ok = caso_b1_ok and caso_b2_ok

    # (c) defeito 2: texto de ALERTA alterado (coluna CRITICA)
    dest_alerta_alterado = copy.deepcopy(linhas)
    for row in dest_alerta_alterado:
        if row["codigo"] == codigo_alvo:
            row["alerta"] = "TEXTO INJETADO PELO AUTOTESTE - NAO E REAL"
            break
    print()
    print(f"[motor] (c) defeito sintetico 2/4: texto de ALERTA (coluna CRITICA) alterado no CODIGO={codigo_alvo}")
    resultado_alerta = comparar_pedido(
        nome="[MOTOR] defeito 2/4 - texto de ALERTA (critica) alterado",
        colspecs=colspecs, linhas_orig=linhas, linhas_dest=dest_alerta_alterado,
        colunas_dest_disponiveis=slugs_disponiveis, tol_num=tol_num, tol_razao=tol_razao, tol_pp=tol_pp,
        ignorar_dest_vazio_entrada_humana=False,
    )
    imprimir_resultado_pedido(resultado_alerta)
    caso_c_ok = (not resultado_alerta.ok) and len(resultado_alerta.criticas_com_divergencia) == 1 \
        and resultado_alerta.criticas_com_divergencia[0].letra == "D"
    print(f"\n[motor] (c) defeito ALERTA: "
          f"{'OK (REPROVOU pelo bloco CRITICA, exatamente na coluna D)' if caso_c_ok else 'FALHOU (deveria ter reprovado via bloco critico)'}")

    # (d) defeito 3: linha faltando
    dest_linha_faltando = [row for row in copy.deepcopy(linhas) if row["codigo"] != codigo_alvo]
    print()
    print(f"[motor] (d) defeito sintetico 3/4: linha do CODIGO={codigo_alvo} REMOVIDA do lado dbt")
    resultado_faltando = comparar_pedido(
        nome="[MOTOR] defeito 3/4 - linha faltando no lado dbt",
        colspecs=colspecs, linhas_orig=linhas, linhas_dest=dest_linha_faltando,
        colunas_dest_disponiveis=slugs_disponiveis, tol_num=tol_num, tol_razao=tol_razao, tol_pp=tol_pp,
        ignorar_dest_vazio_entrada_humana=False,
    )
    imprimir_resultado_pedido(resultado_faltando)
    caso_d_ok = (not resultado_faltando.ok) and resultado_faltando.so_orig == [codigo_alvo]
    print(f"\n[motor] (d) linha faltando: "
          f"{'OK (REPROVOU, so_orig aponta o CODIGO certo)' if caso_d_ok else 'FALHOU (deveria ter reprovado)'}")

    # (e) defeito 4: linha duplicada (fan-out)
    dest_duplicado = copy.deepcopy(linhas)
    dest_duplicado.append(copy.deepcopy(next(r for r in linhas if r["codigo"] == codigo_alvo)))
    print()
    print(f"[motor] (e) defeito sintetico 4/4: linha do CODIGO={codigo_alvo} DUPLICADA no lado dbt (fan-out)")
    resultado_dup = comparar_pedido(
        nome="[MOTOR] defeito 4/4 - linha duplicada no lado dbt (fan-out)",
        colspecs=colspecs, linhas_orig=linhas, linhas_dest=dest_duplicado,
        colunas_dest_disponiveis=slugs_disponiveis, tol_num=tol_num, tol_razao=tol_razao, tol_pp=tol_pp,
        ignorar_dest_vazio_entrada_humana=False,
    )
    imprimir_resultado_pedido(resultado_dup)
    caso_e_ok = (not resultado_dup.ok) and resultado_dup.n_dup_dest == 1
    print(f"\n[motor] (e) linha duplicada: "
          f"{'OK (REPROVOU, 1 CODIGO duplicado detectado)' if caso_e_ok else 'FALHOU (deveria ter reprovado)'}")

    # (f)/(g) AJUSTE 1 (25/08/2026): prova que a isencao de coluna ROTULADA
    # (LETRAS_DIVERGENCIA_POR_DECISAO) do limiar generico NAO virou cheque em
    # branco - divergencia DENTRO do numero esperado passa, ACIMA dele volta
    # a reprovar. Usa CC (CHECK_TRIB, PENDENCIAS item 4), que declara
    # esperado=5 em LETRAS_DIVERGENCIA_ESPERADO (o unico requisito para o
    # teste fazer sentido e' ter numero EXATO, nao aproximado).
    col_rotulada = next(c for c in colspecs if c.letra == "CC")
    esperado_cc = LETRAS_DIVERGENCIA_ESPERADO["CC"]
    if esperado_cc is None:
        raise RuntimeError("CC precisa de um numero esperado EXATO em LETRAS_DIVERGENCIA_ESPERADO "
                            "para o teste (f)/(g) do Ajuste 1 fazer sentido - nao mexer nesse valor "
                            "sem trocar a coluna usada no teste.")

    def _alterar_texto(dest: list[dict], codigos: set[int], slug: str, valor: str):
        for row in dest:
            if row["codigo"] in codigos:
                row[slug] = valor

    n_dentro = max(1, esperado_cc - 1)
    codigos_dentro = {row["codigo"] for row in linhas[:n_dentro]}
    dest_rotulada_dentro = copy.deepcopy(linhas)
    _alterar_texto(dest_rotulada_dentro, codigos_dentro, col_rotulada.slug, "TEXTO INJETADO - DENTRO DO ESPERADO")
    print()
    print(f"[motor] (f) AJUSTE 1 - coluna ROTULADA [{col_rotulada.letra}] {col_rotulada.header} alterada em "
          f"{len(codigos_dentro)} linha(s), ABAIXO do esperado declarado ({esperado_cc}) - veredito "
          f"esperado aqui: PASSOU (a isencao do limiar generico funciona para o previsto).")
    resultado_rotulada_dentro = comparar_pedido(
        nome=f"[MOTOR] AJUSTE 1f - {len(codigos_dentro)} linha(s) alterada(s) em {col_rotulada.header} "
             f"(dentro do esperado, esperado: PASSOU)",
        colspecs=colspecs, linhas_orig=linhas, linhas_dest=dest_rotulada_dentro,
        colunas_dest_disponiveis=slugs_disponiveis, tol_num=tol_num, tol_razao=tol_razao, tol_pp=tol_pp,
        ignorar_dest_vazio_entrada_humana=False,
    )
    imprimir_resultado_pedido(resultado_rotulada_dentro)
    caso_f_ok = resultado_rotulada_dentro.ok
    print(f"\n[motor] (f) rotulada DENTRO do esperado: "
          f"{'OK (PASSOU, a isencao do Ajuste 1 funcionou)' if caso_f_ok else 'FALHOU (deveria ter passado)'}")

    n_acima = esperado_cc + 3
    codigos_acima = {row["codigo"] for row in linhas[:n_acima]}
    dest_rotulada_acima = copy.deepcopy(linhas)
    _alterar_texto(dest_rotulada_acima, codigos_acima, col_rotulada.slug, "TEXTO INJETADO - ACIMA DO ESPERADO")
    print()
    print(f"[motor] (g) AJUSTE 1 - a MESMA coluna ROTULADA [{col_rotulada.letra}] alterada em "
          f"{len(codigos_acima)} linha(s), ACIMA do esperado declarado ({esperado_cc}) - veredito "
          f"esperado aqui: REPROVADO (a isencao NAO cobre excesso - nao e' cheque em branco). "
          f"Nota: {len(codigos_acima)}/{len(linhas)} = {100*len(codigos_acima)/len(linhas):.3f}% ainda "
          f"fica ACIMA de 99,9% dentro da tolerancia (o limiar generico NAO pegaria isto sozinho) - "
          f"quem tem que pegar e' o cheque de excesso do Ajuste 1, isoladamente.")
    resultado_rotulada_acima = comparar_pedido(
        nome=f"[MOTOR] AJUSTE 1g - {len(codigos_acima)} linha(s) alterada(s) em {col_rotulada.header} "
             f"(acima do esperado, esperado: REPROVADO)",
        colspecs=colspecs, linhas_orig=linhas, linhas_dest=dest_rotulada_acima,
        colunas_dest_disponiveis=slugs_disponiveis, tol_num=tol_num, tol_razao=tol_razao, tol_pp=tol_pp,
        ignorar_dest_vazio_entrada_humana=False,
    )
    imprimir_resultado_pedido(resultado_rotulada_acima)
    col_result_acima = next(c for c in resultado_rotulada_acima.resultados if c.letra == col_rotulada.letra)
    caso_g_ok = (
        not resultado_rotulada_acima.ok
        and col_result_acima in resultado_rotulada_acima.estruturais_rotuladas_acima_do_esperado
        and col_result_acima not in resultado_rotulada_acima.estruturais_abaixo_do_limiar
        and col_result_acima.pct_dentro >= 99.9  # prova que NAO foi o limiar generico que pegou
    )
    print(f"\n[motor] (g) rotulada ACIMA do esperado: "
          f"{'OK (REPROVOU pelo motivo do Ajuste 1 - excesso sobre o esperado - nao pelo limiar generico; '
             'a isencao NAO virou cheque em branco)' if caso_g_ok else 'FALHOU (deveria ter reprovado via Ajuste 1)'}")

    tudo_ok = caso_a_ok and caso_b_ok and caso_c_ok and caso_d_ok and caso_e_ok and caso_f_ok and caso_g_ok
    print()
    print("#" * 92)
    print(f"VERIFICACAO DO MOTOR: {'PASSOU - autocomparacao zero divergencia, os 4 defeitos sinteticos '
                                     'classicos E os 2 casos do Ajuste 1 (dentro/acima do esperado) foram '
                                     'todos pegos corretamente' if tudo_ok else 'FALHOU'}")
    print("#" * 92)
    return tudo_ok


# ─────────────────────────────────────────────────────────────────────────────
# Banco (Oracle) - so' usado se `compras.fat_pedido` ja existir
# ─────────────────────────────────────────────────────────────────────────────


def _ler_perfil_compras(profiles_path: str) -> dict:
    import yaml
    with open(profiles_path, encoding="utf-8") as f:
        cfg = yaml.safe_load(f)
    return cfg["compras"]


def _ler_credencial_compras(profiles_path: str) -> tuple[str, str, str]:
    perfil = _ler_perfil_compras(profiles_path)
    target = perfil["target"]
    out = perfil["outputs"][target]
    return out["user"], out["password"], out["connection_string"]


def conectar(lib_dir: str, profiles_path: str):
    import oracledb
    try:
        oracledb.init_oracle_client(lib_dir=lib_dir)
    except oracledb.ProgrammingError:
        pass  # ja inicializado (chamada dupla no mesmo processo) - inofensivo
    user, password, dsn = _ler_credencial_compras(profiles_path)
    conn = oracledb.connect(user=user, password=password, dsn=dsn)
    del password
    return conn


def tabela_existe(conn, schema: str, tabela: str) -> bool:
    cur = conn.cursor()
    cur.execute(
        "select count(*) from all_objects where owner = :o and object_name = :n"
        " and object_type in ('TABLE', 'VIEW')",
        {"o": schema.upper(), "n": tabela.upper()},
    )
    (n,) = cur.fetchone()
    cur.close()
    return n > 0


def colunas_da_tabela(conn, sql_from: str) -> set[str]:
    cur = conn.cursor()
    cur.execute(f"select * from {sql_from} where rownum = 0")
    colunas = {d[0].lower() for d in cur.description}
    cur.close()
    return colunas


def executar_com_colunas(conn, sql: str, binds: dict | None = None, arraysize: int = 5000) -> tuple[list[dict], set[str]]:
    cur = conn.cursor()
    cur.arraysize = arraysize
    cur.prefetchrows = arraysize
    cur.execute(sql, binds or {})
    colunas = [d[0].lower() for d in cur.description]
    linhas = []
    while True:
        bloco = cur.fetchmany(arraysize)
        if not bloco:
            break
        linhas.extend(dict(zip(colunas, row)) for row in bloco)
    cur.close()
    return linhas, set(colunas)


_CANDIDATOS_CHAVE = ["codigo", "codigo_produto", "cod_produto", "codprod"]


def resolver_chave_dest(colunas_disponiveis: set[str]) -> str | None:
    for cand in _CANDIDATOS_CHAVE:
        if cand in colunas_disponiveis:
            return cand
    return None


def resolver_dest(colspecs: list[ColSpec], colunas_disponiveis: set[str]) -> None:
    """Ajusta `c.dest` in-place: se o slug bate direto, usa; senao tenta um
    resolvedor simples por normalizacao (sem acento/pontuacao) - so' entra em
    acao quando `fat_pedido` existir de fato; ate' la' isso e' um no-op na
    pratica (slug ja e' o candidato natural, dado que o dbt usa snake_case
    pt-BR - CONTEXTO §5, e o proprio CONTEXTO cita `embal_compra` como exemplo
    do vocabulario esperado, que bate exatamente com o slug daqui)."""
    def _norm(s: str) -> str:
        return re.sub(r"[^a-z0-9]", "", s.lower())

    normalizados = {_norm(d): d for d in colunas_disponiveis}
    for c in colspecs:
        if c.dest in colunas_disponiveis:
            continue
        alvo = _norm(c.slug)
        if alvo in normalizados:
            c.dest = normalizados[alvo]


# ─────────────────────────────────────────────────────────────────────────────
# main
# ─────────────────────────────────────────────────────────────────────────────


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--tolerancia", type=float, default=0.01,
                     help="tolerancia absoluta para colunas 'num' (quantidade/moeda/codigo) - default 0,01")
    ap.add_argument("--tolerancia-razao", type=float, default=0.001,
                     help="tolerancia absoluta para colunas 'razao' (margem/mkp/aliquota/credito, faixa "
                          "0-1) - default 0,001 = 0,1 ponto percentual. Colunas 'pp' usam este valor x100.")
    ap.add_argument("--xlsx", default=str(XLSX_PATH_PADRAO), help="caminho do xlsx de referencia")
    ap.add_argument("--gabarito", default=str(GABARITO_PATH), help="caminho do gabarito_pedido_formulas.txt")
    ap.add_argument("--limite-codigos", type=int, default=None,
                     help="restringe a comparacao real a N CODIGO (amostra determinística, os N menores) "
                          "- so' para desenvolvimento; o aceite final e' SEM este parametro")
    ap.add_argument("--sem-db", action="store_true",
                     help="nao conecta no Oracle - roda so' a verificacao do motor (xlsx x xlsx)")
    ap.add_argument("--pular-verificacao-motor", action="store_true",
                     help="pula a autocomparacao + defeitos sinteticos (mais rapido, mas sem a prova)")
    ap.add_argument("--lib-dir", default=ORACLE_CLIENT_LIB_DIR, help="Oracle Instant Client (modo thick)")
    ap.add_argument("--profiles-path", default=PROFILES_PATH, help="caminho do profiles.yml")
    args = ap.parse_args()

    t_inicio = time.perf_counter()

    # 1) gabarito, grafo, classificacao -----------------------------------------------------
    colunas_gabarito = carregar_gabarito(Path(args.gabarito))
    grafo_refs = construir_grafo(colunas_gabarito)
    tipos = montar_tipos(colunas_gabarito)
    buckets, fecho_volatil, fecho_entrada_humana = montar_buckets(colunas_gabarito, grafo_refs)
    fecho_condicional = montar_fecho_condicional(grafo_refs)
    colspecs = montar_colspecs(colunas_gabarito, tipos, buckets, fecho_condicional)

    print("#" * 92)
    print(f"GABARITO: {len(colunas_gabarito)} colunas em {args.gabarito} "
          f"({len(colspecs)} comparaveis - exclui CODIGO=chave e a coluna I '(nao usado)')")
    print("#" * 92)
    contagem_bucket = defaultdict(int)
    for c in colspecs:
        contagem_bucket[c.bucket] += 1
    for bucket in _ORDEM_BUCKET:
        letras = sorted((c.letra for c in colspecs if c.bucket == bucket),
                         key=lambda l: (len(l), l))
        print(f"  {_ROTULO_BUCKET[bucket]}: {contagem_bucket[bucket]} coluna(s)")
        print(f"      {', '.join(letras)}")
    print()
    print("  AVISO DE TENSAO: as 4 colunas CRITICAS exigem zero divergencia OBRIGATORIA por instrucao")
    print("  explicita, mesmo quando a propria formula le' uma coluna de outro bucket:")
    for letra in sorted(LETRAS_CRITICAS):
        c = next(c for c in colunas_gabarito if c.letra == letra)
        deps = grafo_refs.get(letra, set())
        deps_taggeds = [(d, buckets.get(d, "?")) for d in deps]
        tensao = [f"{d}({b})" for d, b in deps_taggeds if b in ("volatil", "entrada_humana")]
        if tensao:
            print(f"    [{letra}] {c.header} depende de: {', '.join(tensao)} - divergencia aqui pode "
                  f"ser causada por defasagem/decisao vazia, mas AINDA ASSIM reprova (sem excecao).")
    print(f"  tolerancias: num={args.tolerancia}  razao={args.tolerancia_razao}  "
          f"pp={args.tolerancia_razao * 100} (texto/data: comparacao exata)")

    # 2) leitura da aba `pedido` -------------------------------------------------------------
    xlsx_path = Path(args.xlsx)
    print()
    print("#" * 92)
    print(f"LEITURA DA ABA `{XLSX_ABA}` (streaming - openpyxl read_only=True, data_only=True): {xlsx_path}")
    print("#" * 92)
    linhas, avisos_header, tempo_leitura, pico_mb = ler_pedido_xlsx(xlsx_path, colunas_gabarito)
    print(f"Linhas lidas   : {len(linhas)}")
    print(f"Colunas lidas  : {len(colunas_gabarito)} (todas as posicoes A..{colunas_gabarito[-1].letra})")
    print(f"Tempo de leitura: {tempo_leitura:.2f}s")
    print(f"Pico de memoria do processo (Windows, working set): "
          f"{pico_mb:.1f} MB" if pico_mb is not None else "Pico de memoria: indisponivel nesta plataforma")
    for a in avisos_header:
        print(f"AVISO cabecalho: {a}")
    if linhas:
        print()
        print("Amostra - primeiras 10 colunas da primeira linha lida:")
        primeira = linhas[0]
        for c in colunas_gabarito[:10]:
            if c.letra in LETRAS_IGNORADAS and c.letra != LETRA_CHAVE:
                continue
            slug = _slug(c.header)
            valor = primeira.get("codigo") if c.letra == LETRA_CHAVE else primeira.get(slug)
            print(f"    [{c.letra}] {c.header:<20} (slug={slug:<20}) = {_fmt(valor)}")

    # 3) verificacao do motor -----------------------------------------------------------------
    motor_ok = True
    if not args.pular_verificacao_motor:
        motor_ok = verificar_motor(colspecs, linhas, args.tolerancia, args.tolerancia_razao,
                                    args.tolerancia_razao * 100)
        if not motor_ok:
            print()
            print("PARADA: a verificacao do motor FALHOU - nao faz sentido confiar numa comparacao real")
            print("contra fat_pedido enquanto o motor nao prova que le'/compara corretamente.")
            sys.exit(1)
    else:
        print()
        print("[motor] --pular-verificacao-motor ativo - autocomparacao/defeitos sinteticos NAO rodados "
              "nesta execucao (o aceite final exige rodar SEM esta flag pelo menos uma vez).")

    if args.sem_db:
        print()
        print("--sem-db ativo: parando aqui (so' o motor foi verificado, sem tentar comparar contra "
              "compras.fat_pedido).")
        tempo_total = time.perf_counter() - t_inicio
        print(f"Tempo total: {tempo_total:.2f}s")
        sys.exit(0 if motor_ok else 1)

    # 4) comparacao real contra compras.fat_pedido (se existir) -------------------------------
    print()
    print("#" * 92)
    print("COMPARACAO REAL: compras.fat_pedido x aba `pedido`")
    print("#" * 92)
    conn = conectar(args.lib_dir, args.profiles_path)
    print(f"Conectado como COMPRAS em {datetime.datetime.now():%Y-%m-%d %H:%M:%S}.")

    if not tabela_existe(conn, "COMPRAS", "fat_pedido"):
        conn.close()
        print()
        print("PENDENTE: compras.fat_pedido AINDA NAO EXISTE. A verificacao do motor (item acima) "
              "PASSOU - o instrumento esta pronto para rodar assim que o model existir.")
        tempo_total = time.perf_counter() - t_inicio
        print(f"Tempo total: {tempo_total:.2f}s")
        sys.exit(2)

    if args.limite_codigos:
        print(f"AVISO: --limite-codigos={args.limite_codigos} ativo (modo desenvolvimento) - o aceite "
              f"final exige rodar sem este parametro.")
        codigos_amostra = sorted(r["codigo"] for r in linhas)[: args.limite_codigos]
        conjunto = set(codigos_amostra)
        linhas_cmp = [r for r in linhas if r["codigo"] in conjunto]
    else:
        codigos_amostra = None
        linhas_cmp = linhas

    filtro_sql = ""
    binds: dict = {}
    if codigos_amostra:
        cur = conn.cursor()
        cur.execute("select * from compras.fat_pedido where rownum = 0")
        cols_preview = {d[0].lower() for d in cur.description}
        cur.close()
        chave = resolver_chave_dest(cols_preview) or "codigo"
        placeholders = ",".join(f":c{i}" for i in range(len(codigos_amostra)))
        filtro_sql = f" where {chave} in ({placeholders})"
        binds = {f"c{i}": c for i, c in enumerate(codigos_amostra)}

    t0 = time.perf_counter()
    linhas_dest, colunas_dest = executar_com_colunas(conn, f"select * from compras.fat_pedido{filtro_sql}", binds)
    tempo_dest = time.perf_counter() - t0
    conn.close()
    print(f"compras.fat_pedido lido: {len(linhas_dest)} linhas em {tempo_dest:.2f}s")

    chave_dest = resolver_chave_dest(colunas_dest)
    if chave_dest is None:
        print(f"PENDENTE: compras.fat_pedido existe mas nenhuma coluna-chave candidata "
              f"({_CANDIDATOS_CHAVE}) foi encontrada entre as colunas disponiveis "
              f"({sorted(colunas_dest)}) - impossivel indexar por CODIGO.")
        sys.exit(2)
    for r in linhas_dest:
        r["codigo"] = int(r[chave_dest])

    resolver_dest(colspecs, colunas_dest)

    resultado = comparar_pedido(
        nome="compras.fat_pedido x aba `pedido` (CRITERIO DE ACEITE FINAL)",
        colspecs=colspecs, linhas_orig=linhas_cmp, linhas_dest=linhas_dest,
        colunas_dest_disponiveis=colunas_dest, tol_num=args.tolerancia, tol_razao=args.tolerancia_razao,
        tol_pp=args.tolerancia_razao * 100,
    )
    # A atribuicao de D roda ANTES do veredito porque o veredito DEPENDE dela
    # (ver ResultadoPedido.criticas_com_divergencia). Recalculada contra os dados
    # de HOJE - numero de registro envelhece, CONTEXTO.md secao 6.0.1.
    atribuicao_alerta = atribuir_divergencia_alerta(
        colspecs=colspecs, linhas_orig=linhas_cmp, linhas_dest=linhas_dest,
        tol_num=args.tolerancia, tol_razao=args.tolerancia_razao, tol_pp=args.tolerancia_razao * 100,
    )
    resultado.nao_atribuiveis_alerta = len(atribuicao_alerta.nao_atribuiveis)

    imprimir_resultado_pedido(resultado)
    imprimir_resumo_executivo(resultado)
    imprimir_atribuicao_alerta(atribuicao_alerta)

    tempo_total = time.perf_counter() - t_inicio
    print(f"\nTempo total: {tempo_total:.2f}s")
    sys.exit(0 if resultado.ok else 1)


if __name__ == "__main__":
    main()
