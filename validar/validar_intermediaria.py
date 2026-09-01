"""Prova que `compras.int_venda_mensal_pivot` reproduz o Power Query `shared
fVendaMes` do gabarito (`docs/gabarito_powerquery.m`) - a Etapa 3 (camada
intermediaria de venda mensal, ja pivotada por SKU).

DESENHO DA VALIDACAO (leia antes de mexer) - ver CONTEXTO.md secao 6.3
------------------------------------------------------------------------------
Duas frentes, e so' a primeira decide o veredito (exit code):

  * FRENTE 1 (decide) - reimplementacao INDEPENDENTE em Python de
    `shared fVendaMes`, lida a partir do `.m` (NUNCA do SQL do model - se eu
    lesse o SQL do outro agente para escrever a referencia, os dois erram
    junto e a validacao nao vale nada). A reimplementacao le
    `compras.int_venda_mensal` (ja validado: zero divergencia em 18/18
    colunas nos meses fechados - `validar_relatorios.py`) e compara contra
    `compras.int_venda_mensal_pivot`. Os DOIS lados leem tabelas MATERIALIZADAS
    do schema `compras`, a poucos segundos de diferenca um do outro - nao ha
    consulta ao vivo ao CEDEP em nenhum dos dois, entao nao ha a defasagem
    "CEDEP vendeu entre os dois SELECTs" que forcou o mensal (Etapa 2) a
    separar mes corrente como informativo (CONTEXTO 6.3). Aqui QATUAL/VATUAL/
    CLI_*_ATUAL/DIAS_SEM_EST_ATUAL/FORA_DE_LINHA sao validados com a MESMA
    exigencia que Q00..Q11 - zero divergencia nas 35 colunas, sem exceber.
    (O unico risco residual de defasagem e' `int_venda_mensal_pivot` ter sido
    CONSTRUIDO a partir de uma foto MAIS ANTIGA de `int_venda_mensal` que a
    lida aqui - o script compara `last_ddl_time` dos dois e avisa se for o
    caso; nao afrouxa o veredito por isso, so' avisa.)

  * FRENTE 2 (informativa, NUNCA reprova) - `compras.int_venda_mensal_pivot`
    (o MODEL de verdade, nao a reimplementacao Python) contra a aba
    `fVendaMes` de `referencia/MODELO_COMPRAS_CEDEP_v10.xlsx` - uma FOTO
    ANTIGA (o WinThor mudou desde que a planilha foi atualizada). O valor
    dela e' outro: e' a UNICA comparacao que nao depende da minha propria
    leitura do `.m` estar certa - se a Frente 1 e o SQL do model errassem do
    MESMO jeito (os dois interpretando o `.m` errado da mesma forma), a
    Frente 1 sozinha nao pegaria isso; comparar o MODEL direto contra a
    planilha pega diferenca SISTEMATICA (coluna deslocada, offset trocado,
    sinal invertido) mesmo nesse cenario. Divergencia aqui e' esperada e
    **nao reprova o aceite** - o que importa e' ordem de grandeza/estrutura
    batendo, e destacar coluna com divergencia em quase 100% das linhas
    (sinal de erro sistematico) vs poucas linhas (deriva normal do dado).

Uso:
    python validar_intermediaria.py [--tolerancia 0.01]
                                     [--somente frente1|frente2|ambos]
                                     [--limite-codigos N] [--sem-xlsx]
                                     [--xlsx CAMINHO] [--autoteste]

Aceite: zero divergencia nas 35 colunas de `compras.int_venda_mensal_pivot`
contra a reimplementacao Python da Frente 1 (int_venda_mensal x fVendaMes/.m).

--autoteste roda um caso SINTETICO, sem banco nenhum, que injeta um defeito
deliberado (offset Q00/Q01 trocado num SKU) e prova que o motor de comparacao
REPROVA - existe justamente para provar que este validador pega defeito, nao
so' que ele passa.

O script NAO afrouxa nada para "passar" - ele mede e imprime; ajustar o model
para o numero fechar e' trabalho de outro agente, e so' depois de alguem
decidir que o gabarito e' que esta certo.
"""
from __future__ import annotations

import argparse
import datetime
import os
import sys
import time
from dataclasses import dataclass, field
from pathlib import Path

import oracledb
import openpyxl
import yaml

# ─────────────────────────────────────────────────────────────────────────────
# Caminhos fixos do ambiente (ver CONTEXTO.md §3)
# ─────────────────────────────────────────────────────────────────────────────
ORACLE_CLIENT_LIB_DIR = r"C:\Oracle\instantclient_21_17"
PROFILES_PATH = os.path.expanduser(r"~\.dbt\profiles.yml")
REPO_DIR = Path(__file__).resolve().parent.parent
XLSX_PATH_PADRAO = REPO_DIR / "referencia" / "MODELO_COMPRAS_CEDEP_v10.xlsx"
XLSX_ABA = "fVendaMes"

MESES_PT = ["jan", "fev", "mar", "abr", "mai", "jun", "jul", "ago", "set", "out", "nov", "dez"]


def rotulo_mes(dt: datetime.date) -> str:
    """Replica `Rotulo` do .m: "jul/26"."""
    return f"{MESES_PT[dt.month - 1]}/{dt.year % 100:02d}"


def _add_months(d: datetime.date, n: int) -> datetime.date:
    m = d.month - 1 + n
    y = d.year + m // 12
    m = m % 12 + 1
    return d.replace(year=y, month=m, day=1)


# ─────────────────────────────────────────────────────────────────────────────
# Dataclasses / motor de comparacao (mesmo desenho de validar_relatorios.py -
# fan-out reprova, contagem de linha diferente reprova, chave duplicada
# reprova; ver comentarios em `ResultadoRelatorio.ok`)
# ─────────────────────────────────────────────────────────────────────────────


@dataclass
class ColSpec:
    orig: str   # nome logico (chave no dict Python / no dict do xlsx), lowercase
    dest: str   # nome resolvido na tabela dbt (lowercase)
    tipo: str   # txt | int | num
    rotulo: str


def _normaliza_vazio(v):
    """'' (celula vazia do Excel / IFERROR(...,"")) equivale a NULL - trate as
    duas como equivalentes (ver instrucao do orquestrador)."""
    if isinstance(v, str) and v.strip() == "":
        return None
    return v


def _norm_texto(v):
    v = _normaliza_vazio(v)
    return v


def _norm_num(v):
    v = _normaliza_vazio(v)
    if v is None:
        return None
    return float(v)


def comparar_valor(tipo: str, orig, dest, tolerancia: float):
    """Devolve (igual: bool, diff_ordenacao: float)."""
    if tipo == "txt":
        a, b = _norm_texto(orig), _norm_texto(dest)
        igual = a == b
        return igual, (0.0 if igual else 1.0)
    a, b = _norm_num(orig), _norm_num(dest)
    if a is None and b is None:
        return True, 0.0
    if a is None or b is None:
        return False, float("inf")
    diff = abs(a - b)
    return diff <= tolerancia, diff


def _fmt(v) -> str:
    if v is None:
        return "NULL"
    if isinstance(v, float):
        return f"{v:.4f}".rstrip("0").rstrip(".")
    s = str(v)
    return s if len(s) <= 40 else s[:37] + "..."


@dataclass
class ResultadoColuna:
    coluna_orig: str
    coluna_dest: str
    rotulo: str
    total_comparavel: int
    divergentes: int
    maior_diferenca: float
    exemplos: list[tuple]  # (chave_legivel, valor_orig, valor_dest, diff)

    @property
    def pct(self) -> float:
        return 100.0 * self.divergentes / self.total_comparavel if self.total_comparavel else 0.0


@dataclass
class ResultadoRelatorio:
    nome: str
    tempo_orig_s: float
    tempo_dest_s: float
    n_orig: int
    n_dest: int
    n_chaves_duplicadas_orig: int
    n_chaves_duplicadas_dest: int
    so_orig: list
    so_dest: list
    colunas_nao_encontradas: list[str]
    resultados_coluna: list[ResultadoColuna] = field(default_factory=list)
    informativo: bool = False
    motivo_informativo: str = ""
    nunca_reprova: bool = False  # Frente 2: mesmo so_dest/duplicata NUNCA reprova o aceite

    @property
    def ok(self) -> bool:
        """Brecha 1 (fan-out): chave duplicada em qualquer lado, ou contagem de
        linha divergente, REPROVA aqui - nao e' so' aviso impresso."""
        return (
            not self.so_orig
            and not self.so_dest
            and not self.colunas_nao_encontradas
            and self.n_chaves_duplicadas_orig == 0
            and self.n_chaves_duplicadas_dest == 0
            and self.n_orig == self.n_dest
            and all(r.divergentes == 0 for r in self.resultados_coluna)
        )

    @property
    def reprova(self) -> bool:
        if self.nunca_reprova:
            return False
        return not self.ok


def _indexar(linhas: list[dict], chave_cols: list[str]) -> tuple[dict, int]:
    idx: dict = {}
    duplicadas = 0
    for row in linhas:
        k = tuple(row[c] for c in chave_cols)
        if k in idx:
            duplicadas += 1
        idx[k] = row
    return idx, duplicadas


def comparar_relatorio(
    nome: str,
    colspecs: list[ColSpec],
    linhas_orig: list[dict],
    linhas_dest: list[dict],
    colunas_dest_disponiveis: set[str],
    tempo_orig_s: float,
    tempo_dest_s: float,
    tolerancia: float,
    fmt_chave=lambda k: f"CODIGO={k[0]}",
    informativo: bool = False,
    motivo_informativo: str = "",
    nunca_reprova: bool = False,
) -> ResultadoRelatorio:
    idx_orig, dup_orig = _indexar(linhas_orig, ["codigo"])
    idx_dest, dup_dest = _indexar(linhas_dest, ["codigo"])

    chaves_orig = set(idx_orig)
    chaves_dest = set(idx_dest)
    so_orig = sorted(chaves_orig - chaves_dest)
    so_dest = sorted(chaves_dest - chaves_orig)
    comuns = chaves_orig & chaves_dest

    colunas_nao_encontradas = [
        f"{c.orig} -> {c.dest}" for c in colspecs if c.dest not in colunas_dest_disponiveis
    ]

    resultados: list[ResultadoColuna] = []
    for c in colspecs:
        if c.dest not in colunas_dest_disponiveis:
            continue
        total = 0
        divergentes = 0
        maior_diff = 0.0
        exemplos: list[tuple] = []
        for k in comuns:
            ov = idx_orig[k].get(c.orig)
            dv = idx_dest[k].get(c.dest)
            total += 1
            igual, diff = comparar_valor(c.tipo, ov, dv, tolerancia)
            if not igual:
                divergentes += 1
                diff_ordenacao = diff if diff != float("inf") else 1e18
                maior_diff = max(maior_diff, diff_ordenacao if diff_ordenacao < 1e18 else maior_diff)
                exemplos.append((fmt_chave(k), ov, dv, diff_ordenacao))
        exemplos.sort(key=lambda t: t[3], reverse=True)
        resultados.append(
            ResultadoColuna(
                coluna_orig=c.orig,
                coluna_dest=c.dest,
                rotulo=c.rotulo,
                total_comparavel=total,
                divergentes=divergentes,
                maior_diferenca=maior_diff,
                exemplos=exemplos[:3],
            )
        )

    resultados.sort(key=lambda r: r.divergentes, reverse=True)

    return ResultadoRelatorio(
        nome=nome,
        tempo_orig_s=tempo_orig_s,
        tempo_dest_s=tempo_dest_s,
        n_orig=len(linhas_orig),
        n_dest=len(linhas_dest),
        n_chaves_duplicadas_orig=dup_orig,
        n_chaves_duplicadas_dest=dup_dest,
        so_orig=so_orig,
        so_dest=so_dest,
        colunas_nao_encontradas=colunas_nao_encontradas,
        resultados_coluna=resultados,
        informativo=informativo,
        motivo_informativo=motivo_informativo,
        nunca_reprova=nunca_reprova,
    )


def _veredito_str(r: ResultadoRelatorio) -> str:
    if r.nunca_reprova:
        return "INFORMATIVO (nunca reprova o aceite - ver docstring, Frente 2)"
    if r.informativo:
        return "INFORMATIVO (nao entra no veredito)" if not r.reprova else "REPROVADO"
    return "PASSOU" if r.ok else "REPROVADO"


def imprimir_resultado(r: ResultadoRelatorio):
    print()
    print("=" * 88)
    print(f"RELATORIO: {r.nome}")
    if r.nunca_reprova:
        print("(INFORMATIVO - Frente 2, foto antiga da planilha. NUNCA reprova o aceite,")
        print(" mesmo com so_dest/duplicata/contagem diferente - ver docstring do script.)")
    elif r.informativo and r.motivo_informativo:
        print(f"(INFORMATIVO por padrao. {r.motivo_informativo})")
    print("=" * 88)
    print(f"Lado ORIGEM (gabarito): {r.n_orig:>8} linhas em {r.tempo_orig_s:8.2f}s")
    print(f"Lado DESTINO (dbt)    : {r.n_dest:>8} linhas em {r.tempo_dest_s:8.2f}s")
    if r.n_chaves_duplicadas_orig:
        print(f"FALHA: {r.n_chaves_duplicadas_orig} CODIGO(s) duplicado(s) no lado origem "
              f"(deveria ser 1 linha por CODIGO)")
    if r.n_chaves_duplicadas_dest:
        print(f"FALHA: {r.n_chaves_duplicadas_dest} CODIGO(s) duplicado(s) no lado dbt "
              f"(fan-out) - reprova mesmo se os valores da ultima linha batem"
              f"{' (mas nao aqui - informativo)' if r.nunca_reprova else ''}")
    if r.n_orig != r.n_dest:
        print(f"FALHA: contagem de linhas diverge (origem={r.n_orig}, dbt={r.n_dest})")

    print()
    print("-- Diferencas de CONJUNTO --")
    print(f"So' no lado origem (gabarito): {len(r.so_orig)}" +
          ("  (esperado - foto antiga: SKU novo desde a planilha)" if r.nunca_reprova and r.so_orig else ""))
    for k in r.so_orig[:3]:
        print(f"    exemplo: CODIGO={k[0]}")
    print(f"So' no lado dbt              : {len(r.so_dest)}" +
          ("  (esperado - foto antiga: SKU saiu de linha/base desde a planilha)"
           if r.nunca_reprova and r.so_dest else ""))
    for k in r.so_dest[:3]:
        print(f"    exemplo: CODIGO={k[0]}")

    if r.colunas_nao_encontradas:
        print()
        print("-- Colunas esperadas SEM correspondente encontrado no lado dbt --")
        for c in r.colunas_nao_encontradas:
            print(f"    {c}")

    divergentes_cols = [c for c in r.resultados_coluna if c.divergentes > 0]
    ok_cols = [c for c in r.resultados_coluna if c.divergentes == 0]
    print()
    print(f"-- Por coluna, ordenado por gravidade ({len(divergentes_cols)} com divergencia, "
          f"{len(ok_cols)} identicas de {len(r.resultados_coluna)} comparadas) --")
    if not divergentes_cols:
        print("    (nenhuma)")
    for c in divergentes_cols:
        print()
        print(f"  [{c.coluna_orig}] -> [{c.coluna_dest}]  ({c.rotulo})")
        print(f"      {c.divergentes}/{c.total_comparavel} linhas divergem ({c.pct:.2f}%), "
              f"maior diferenca absoluta = {c.maior_diferenca:.4f}")
        for chave, ov, dv, diff in c.exemplos:
            print(f"      exemplo {chave}: origem={_fmt(ov)}  dbt={_fmt(dv)}  diff={diff:.4f}"
                  if diff < 1e18 else
                  f"      exemplo {chave}: origem={_fmt(ov)}  dbt={_fmt(dv)}  (nulo de um lado)")

    if r.nunca_reprova and divergentes_cols:
        print()
        print("-- Resumo Frente 2: sistematico (>=90% das linhas) vs deriva normal --")
        for c in divergentes_cols:
            marca = "<<< SUSPEITO: possivel erro SISTEMATICO (coluna deslocada/offset " \
                    "trocado/sinal invertido)" if c.pct >= 90.0 else "deriva normal (foto antiga)"
            print(f"    {c.coluna_orig:<20} {c.pct:6.2f}%  {marca}")

    print()
    print(f"VEREDITO {r.nome}: {_veredito_str(r)}")


# ─────────────────────────────────────────────────────────────────────────────
# Conexao
# ─────────────────────────────────────────────────────────────────────────────


def _ler_perfil_compras(profiles_path: str) -> dict:
    with open(profiles_path, encoding="utf-8") as f:
        cfg = yaml.safe_load(f)
    return cfg["compras"]


def _ler_credencial_compras(profiles_path: str) -> tuple[str, str, str]:
    """Le usuario/senha/connection_string de profiles.yml. A senha NUNCA e'
    impressa nem gravada em lugar algum - so' vive nesta variavel local e no
    dicionario de conexao do oracledb."""
    perfil = _ler_perfil_compras(profiles_path)
    target = perfil["target"]
    out = perfil["outputs"][target]
    return out["user"], out["password"], out["connection_string"]


def conectar(lib_dir: str, profiles_path: str) -> oracledb.Connection:
    try:
        oracledb.init_oracle_client(lib_dir=lib_dir)
    except oracledb.ProgrammingError:
        pass  # ja inicializado (chamada dupla no mesmo processo) - inofensivo
    user, password, dsn = _ler_credencial_compras(profiles_path)
    conn = oracledb.connect(user=user, password=password, dsn=dsn)
    del password
    return conn


def executar_com_colunas(conn: oracledb.Connection, sql: str, binds: dict | None = None,
                          arraysize: int = 5000) -> tuple[list[dict], set[str]]:
    cur = conn.cursor()
    cur.arraysize = arraysize
    cur.prefetchrows = arraysize
    cur.execute(sql, binds or {})
    colunas = [d[0].lower() for d in cur.description]
    linhas = []
    while True:
        bloco = cur.fetchmany(arraysize)
        if not bloco:
            break
        linhas.extend(dict(zip(colunas, row)) for row in bloco)
    cur.close()
    return linhas, set(colunas)


def tabela_existe(conn: oracledb.Connection, schema: str, tabela: str) -> bool:
    cur = conn.cursor()
    cur.execute(
        "select count(*) from all_objects where owner = :o and object_name = :n"
        " and object_type in ('TABLE', 'VIEW')",
        {"o": schema.upper(), "n": tabela.upper()},
    )
    (n,) = cur.fetchone()
    cur.close()
    return n > 0


def data_construcao(conn: oracledb.Connection, schema: str, tabela: str) -> datetime.datetime | None:
    cur = conn.cursor()
    cur.execute(
        "select last_ddl_time from all_objects where owner = :o and object_name = :n"
        " and object_type = 'TABLE'",
        {"o": schema.upper(), "n": tabela.upper()},
    )
    row = cur.fetchone()
    cur.close()
    return row[0] if row else None


# ─────────────────────────────────────────────────────────────────────────────
# Reimplementacao Python de `shared fVendaMes` (docs/gabarito_powerquery.m)
# ─────────────────────────────────────────────────────────────────────────────

# 35 colunas na ordem de `OrdemFinal` do .m (sem CODIGO, que e' a chave)
NOMES_Q = [f"q{i:02d}" for i in range(12)]
NOMES_V = [f"v{i:02d}" for i in range(6)]
COLUNAS_LOGICAS_VALOR = NOMES_Q + NOMES_V + [
    "qatual", "vatual", "cli_atac_atual", "cli_var_atual", "dias_sem_est_atual", "fora_de_linha",
    "tx_devolucao_3m", "cli_atac_00", "cli_var_00", "dias_sem_est_00",
    "med02", "med03", "med04", "med06", "med09", "med12",
]


def _offset(mes: datetime.date, mes_ref: datetime.date) -> int:
    """`OFFSET` do .m: (ano_ref*12+mes_ref) - (ano*12+mes) - 1."""
    return (mes_ref.year * 12 + mes_ref.month) - (mes.year * 12 + mes.month) - 1


def construir_gabarito(rows: list[dict]) -> tuple[dict[int, dict], datetime.date | None, int]:
    """Reimplementacao linha a linha de `shared fVendaMes` (docs/gabarito_powerquery.m),
    a partir de linhas de `compras.int_venda_mensal` (mes, codigo, qtd=quantidade_liquida,
    valor=valor_liquido, qt_fat=quantidade_total (BRUTO faturado - unico uso de bruto,
    so' em TX_DEVOLUCAO_3M), qt_dev=quantidade_devolucao (BRUTO devolvido), cli_atacado,
    cli_varejo, dias_sem_estoque, fora_de_linha).

    Devolve (dict codigo -> registro com as 35 colunas logicas, mes_ref, qtd_meses_fechados
    GLOBAIS - usado so' para decidir se MED09/MED12 sao null).
    """
    if not rows:
        return {}, None, 0

    mes_ref = max(r["mes"] for r in rows)

    # NOTA sobre MED09/MED12: a condicao literal do .m e' `List.Count(NomesQ) >= 9`,
    # e NomesQ = List.Transform({0..11}, ...) tem SEMPRE 12 elementos - lida ao pe' da
    # letra essa condicao e' sempre verdadeira (parece codigo morto/bug do autor do .m,
    # que provavelmente queria checar quantos meses fechados EXISTEM de fato, nao o
    # tamanho da lista de rotulos). Aqui reproduzimos a INTENCAO de negocio descrita
    # pelo orquestrador (MED09/MED12 nulos se nao houver 9/12 meses fechados de
    # historico), medida pelo numero de meses fechados DISTINTOS presentes em TODO o
    # dataset (nao por SKU) - com o historico padrao de 24 meses (compras_meses_historico
    # no dbt_project.yml) isso nunca diverge do resultado da leitura literal do .m,
    # entao nao ha' risco pratico de "corrigir" o gabarito para um numero errado.
    meses_fechados_globais = {r["mes"] for r in rows if _offset(r["mes"], mes_ref) >= 0}
    qtd_meses_fechados = len(meses_fechados_globais)

    todos_codigos = sorted({int(r["codigo"]) for r in rows})
    por_codigo: dict[int, list[dict]] = {c: [] for c in todos_codigos}
    for r in rows:
        por_codigo[int(r["codigo"])].append(r)

    resultado: dict[int, dict] = {}
    for codigo in todos_codigos:
        rec: dict = {"codigo": codigo}
        for nome in NOMES_Q:
            rec[nome] = 0.0
        for nome in NOMES_V:
            rec[nome] = 0.0
        rec["qatual"] = 0.0
        rec["vatual"] = 0.0
        rec["cli_atac_atual"] = 0
        rec["cli_var_atual"] = 0
        rec["dias_sem_est_atual"] = 0
        rec["fora_de_linha"] = "N"  # nulo -> 'N', unica excecao ao nulo->0
        rec["tx_devolucao_3m"] = 0.0
        rec["cli_atac_00"] = 0
        rec["cli_var_00"] = 0
        rec["dias_sem_est_00"] = 0

        fat3 = 0.0
        dev3 = 0.0
        for r in por_codigo[codigo]:
            off = _offset(r["mes"], mes_ref)
            if off == -1:
                # MES = MesAtual -> QATUAL/VATUAL/CLI_*_ATUAL/DIAS_SEM_EST_ATUAL/FORA_DE_LINHA
                rec["qatual"] = float(r["qtd"] or 0)
                rec["vatual"] = float(r["valor"] or 0)
                rec["cli_atac_atual"] = int(r["cli_atacado"] or 0)
                rec["cli_var_atual"] = int(r["cli_varejo"] or 0)
                rec["dias_sem_est_atual"] = int(r["dias_sem_estoque"] or 0)
                fdl = r.get("fora_de_linha")
                rec["fora_de_linha"] = fdl if fdl not in (None, "") else "N"
            elif 0 <= off <= 11:
                rec[f"q{off:02d}"] = float(r["qtd"] or 0)
                if off <= 5:
                    rec[f"v{off:02d}"] = float(r["valor"] or 0)
                if off == 0:
                    rec["cli_atac_00"] = int(r["cli_atacado"] or 0)
                    rec["cli_var_00"] = int(r["cli_varejo"] or 0)
                    rec["dias_sem_est_00"] = int(r["dias_sem_estoque"] or 0)
            # off > 11 (fora da janela Q/V) ou off < -1 (mes futuro, nao deveria existir):
            # ignorado aqui de proposito - nao contribui pra nenhuma coluna do pivot.

            if 0 <= off < 3:
                fat3 += float(r.get("qt_fat") or 0)
                dev3 += float(r.get("qt_dev") or 0)

        rec["tx_devolucao_3m"] = 0.0 if fat3 == 0 else dev3 / fat3

        def soma_q(a: int, b: int) -> float:
            return sum(rec[f"q{i:02d}"] for i in range(a, b + 1))

        rec["med02"] = soma_q(0, 1) / 2
        rec["med03"] = soma_q(0, 2) / 3
        rec["med04"] = soma_q(0, 3) / 4
        rec["med06"] = soma_q(0, 5) / 6
        rec["med09"] = soma_q(0, 8) / 9 if qtd_meses_fechados >= 9 else None
        rec["med12"] = soma_q(0, 11) / 12 if qtd_meses_fechados >= 12 else None

        resultado[codigo] = rec

    return resultado, mes_ref, qtd_meses_fechados


def ler_venda_mensal(conn: oracledb.Connection, codigos: list[int] | None = None) -> list[dict]:
    filtro = ""
    binds: dict = {}
    if codigos:
        placeholders = ",".join(f":c{i}" for i in range(len(codigos)))
        filtro = f" where codigo_produto in ({placeholders})"
        binds = {f"c{i}": c for i, c in enumerate(codigos)}
    sql = f"""
        select mes, codigo_produto as codigo, quantidade_liquida as qtd,
               valor_liquido as valor, quantidade_total as qt_fat,
               quantidade_devolucao as qt_dev, quantidade_clientes_atacado as cli_atacado,
               quantidade_clientes_varejo as cli_varejo, dias_sem_estoque,
               fora_de_linha
          from compras.int_venda_mensal{filtro}
    """
    linhas, _ = executar_com_colunas(conn, sql, binds)
    for r in linhas:
        m = r["mes"]
        r["mes"] = m.date() if isinstance(m, datetime.datetime) else m
    return linhas


def _amostra_codigos(conn: oracledb.Connection, n: int) -> list[int]:
    """Amostra DETERMINISTICA dos N menores codigo_produto (mesmo criterio de
    ordenacao-antes-do-rownum de validar_relatorios.py, para nao repetir a
    Brecha 5 la' documentada)."""
    cur = conn.cursor()
    cur.execute(
        "select codigo from (select distinct codigo_produto as codigo from "
        "compras.int_venda_mensal order by codigo_produto) where rownum <= :n",
        {"n": n},
    )
    codigos = [row[0] for row in cur.fetchall()]
    cur.close()
    return codigos


# ─────────────────────────────────────────────────────────────────────────────
# Resolucao de colunas do lado dbt (nome exato de `int_venda_mensal_pivot`
# ainda nao e' garantido - o model esta' sendo escrito por outro agente
# enquanto este validador e' escrito. Resolve por candidatos, na convencao
# snake_case pt-BR do projeto - ver CONTEXTO.md §5)
# ─────────────────────────────────────────────────────────────────────────────

_CANDIDATOS_CHAVE = ["codigo_produto", "codigo", "cod_produto"]
_CANDIDATOS_EXTRA = {
    "cli_atac_atual": ["cli_atac_atual", "cli_atacado_atual"],
    "cli_var_atual": ["cli_var_atual", "cli_varejo_atual"],
    "dias_sem_est_atual": ["dias_sem_est_atual", "dias_sem_estoque_atual"],
    "cli_atac_00": ["cli_atac_00", "cli_atacado_00"],
    "cli_var_00": ["cli_var_00", "cli_varejo_00"],
    "dias_sem_est_00": ["dias_sem_est_00", "dias_sem_estoque_00"],
}


def resolver_chave_dest(colunas_disponiveis: set[str]) -> str | None:
    for cand in _CANDIDATOS_CHAVE:
        if cand in colunas_disponiveis:
            return cand
    return None


def resolver_colspecs(colunas_disponiveis: set[str]) -> list[ColSpec]:
    specs: list[ColSpec] = []
    for nome in NOMES_Q:
        specs.append(ColSpec(orig=nome, dest=nome, tipo="num",
                              rotulo=f"{nome.upper()} (qtd liquida, offset {int(nome[1:])})"))
    for nome in NOMES_V:
        specs.append(ColSpec(orig=nome, dest=nome, tipo="num",
                              rotulo=f"{nome.upper()} (valor liquido, offset {int(nome[1:])})"))
    rotulos_extra = {
        "qatual": "QATUAL (qtd liquida, mes corrente)",
        "vatual": "VATUAL (valor liquido, mes corrente)",
        "cli_atac_atual": "CLI_ATAC_ATUAL",
        "cli_var_atual": "CLI_VAR_ATUAL",
        "dias_sem_est_atual": "DIAS_SEM_EST_ATUAL",
        "fora_de_linha": "FORA_DE_LINHA",
        "tx_devolucao_3m": "TX_DEVOLUCAO_3M (unico campo com quantidade BRUTA)",
        "cli_atac_00": "CLI_ATAC_00",
        "cli_var_00": "CLI_VAR_00",
        "dias_sem_est_00": "DIAS_SEM_EST_00",
        "med02": "MED02", "med03": "MED03", "med04": "MED04",
        "med06": "MED06", "med09": "MED09", "med12": "MED12",
    }
    for nome, rotulo in rotulos_extra.items():
        tipo = "txt" if nome == "fora_de_linha" else "num"
        dest = nome
        for cand in _CANDIDATOS_EXTRA.get(nome, [nome]):
            if cand in colunas_disponiveis:
                dest = cand
                break
        specs.append(ColSpec(orig=nome, dest=dest, tipo=tipo, rotulo=rotulo))
    return specs


# ─────────────────────────────────────────────────────────────────────────────
# Leitura da aba fVendaMes do xlsx - STREAMING (arquivo tem 25 MB, ver
# CONTEXTO.md/instrucao do orquestrador: read_only=True, nunca carregar tudo)
# ─────────────────────────────────────────────────────────────────────────────

# ordem posicional exata de OrdemFinal no .m (35 colunas) - o cabecalho real
# traz o rotulo do mes colado (ex. "Q00_jul/26"), entao mapeamos por POSICAO,
# nao pelo nome completo (o rotulo do mes muda a cada atualizacao da planilha)
ORDEM_LOGICA_XLSX = ["codigo"] + NOMES_Q + NOMES_V + COLUNAS_LOGICAS_VALOR[len(NOMES_Q) + len(NOMES_V):]


def ler_fvendames_xlsx(caminho: Path, limite_linhas: int | None = None) -> list[dict]:
    wb = openpyxl.load_workbook(str(caminho), read_only=True, data_only=True)
    try:
        if XLSX_ABA not in wb.sheetnames:
            raise RuntimeError(f"aba '{XLSX_ABA}' nao encontrada em {caminho}. "
                                f"Abas disponiveis: {wb.sheetnames}")
        ws = wb[XLSX_ABA]
        it = ws.iter_rows(values_only=True)
        header = next(it)
        n_cols = min(len(header), len(ORDEM_LOGICA_XLSX))

        # sanidade posicional: os prefixos Q00.. / V00.. tem que bater com o
        # rotulo esperado na mesma posicao (nao trava a execucao - so' avisa,
        # a comparacao inteira e' informativa mesmo)
        avisos = []
        for i in range(n_cols):
            esperado = ORDEM_LOGICA_XLSX[i].upper()
            real = str(header[i] or "").upper()
            if not real.startswith(esperado):
                avisos.append(f"posicao {i}: esperado prefixo '{esperado}', cabecalho real '{header[i]}'")

        linhas: list[dict] = []
        for row in it:
            if row is None or all(v is None for v in row):
                continue
            codigo_bruto = row[0]
            try:
                codigo = int(codigo_bruto)
            except (TypeError, ValueError):
                continue
            rec = {"codigo": codigo}
            for i in range(1, n_cols):
                rec[ORDEM_LOGICA_XLSX[i]] = _normaliza_vazio(row[i])
            linhas.append(rec)
            if limite_linhas and len(linhas) >= limite_linhas:
                break
        return linhas, avisos
    finally:
        wb.close()


# ─────────────────────────────────────────────────────────────────────────────
# FRENTE 1 - decide o veredito
# ─────────────────────────────────────────────────────────────────────────────


def validar_frente1(conn: oracledb.Connection, tolerancia: float,
                     limite_codigos: int | None) -> tuple[ResultadoRelatorio | None, str | None]:
    print()
    print("[frente1] METODO: reimplementacao Python de `shared fVendaMes` (docs/gabarito_"
          "powerquery.m), lida a partir de compras.int_venda_mensal (ja validado), "
          "comparada contra compras.int_venda_mensal_pivot. Os dois lados leem tabelas "
          "MATERIALIZADAS do schema compras - sem consulta ao vivo ao CEDEP - entao nao "
          "ha' defasagem por hora do dia. Aceite: zero divergencia nas 35 colunas, SEM "
          "excecao para o mes corrente.")

    if not tabela_existe(conn, "COMPRAS", "int_venda_mensal"):
        return None, "compras.int_venda_mensal nao existe - impossivel construir o gabarito"

    codigos_filtro = None
    if limite_codigos:
        codigos_filtro = _amostra_codigos(conn, limite_codigos)
        print(f"[frente1] AVISO: --limite-codigos={limite_codigos} ativo (modo desenvolvimento) "
              f"- o aceite final exige rodar sem este parametro.")

    t0 = time.perf_counter()
    linhas_venda_mensal = ler_venda_mensal(conn, codigos_filtro)
    gabarito, mes_ref, qtd_meses_fechados = construir_gabarito(linhas_venda_mensal)
    tempo_orig = time.perf_counter() - t0
    print(f"[frente1] int_venda_mensal lido: {len(linhas_venda_mensal)} linhas -> "
          f"gabarito Python construido: {len(gabarito)} SKUs distintos em {tempo_orig:.2f}s")
    if mes_ref:
        print(f"[frente1] mes_ref (MesAtual do .m) = {rotulo_mes(mes_ref)} ({mes_ref:%Y-%m-%d}); "
              f"meses fechados globais no dataset = {qtd_meses_fechados} "
              f"(MED09 {'calculado' if qtd_meses_fechados >= 9 else 'NULL'}, "
              f"MED12 {'calculado' if qtd_meses_fechados >= 12 else 'NULL'})")

    if not tabela_existe(conn, "COMPRAS", "int_venda_mensal_pivot"):
        print("[frente1] compras.int_venda_mensal_pivot AINDA NAO EXISTE. Prova de execucao "
              "da reimplementacao Python: OK.")
        if gabarito:
            exemplo_codigo = next(iter(sorted(gabarito)))
            print(f"[frente1] amostra do gabarito (CODIGO={exemplo_codigo}): "
                  f"{ {k: _fmt(v) for k, v in gabarito[exemplo_codigo].items()} }")
        return None, "int_venda_mensal_pivot (model ainda nao existe) - comparacao pendente"

    build_venda_mensal = data_construcao(conn, "COMPRAS", "int_venda_mensal")
    build_pivot = data_construcao(conn, "COMPRAS", "int_venda_mensal_pivot")
    if build_venda_mensal and build_pivot and build_venda_mensal > build_pivot:
        bvm = f"{build_venda_mensal:%Y-%m-%d %H:%M:%S}"
        bp = f"{build_pivot:%Y-%m-%d %H:%M:%S}"
        print(f"[frente1] AVISO: int_venda_mensal foi reconstruido ({bvm}) DEPOIS de "
              f"int_venda_mensal_pivot ({bp}) - o pivot pode estar construido sobre uma foto "
              f"mais antiga. Rode `dbt run --select int_venda_mensal_pivot` antes de confiar "
              f"no veredito.")

    filtro_sql = ""
    binds = {}
    if codigos_filtro:
        chave = "codigo_produto"  # tenta o nome mais provavel primeiro; ajustado abaixo se preciso
        cur = conn.cursor()
        cur.execute("select * from compras.int_venda_mensal_pivot where rownum = 0")
        cols_disp_preview = {d[0].lower() for d in cur.description}
        cur.close()
        chave = resolver_chave_dest(cols_disp_preview) or "codigo_produto"
        placeholders = ",".join(f":c{i}" for i in range(len(codigos_filtro)))
        filtro_sql = f" where {chave} in ({placeholders})"
        binds = {f"c{i}": c for i, c in enumerate(codigos_filtro)}

    t0 = time.perf_counter()
    linhas_pivot, colunas_pivot = executar_com_colunas(
        conn, f"select * from compras.int_venda_mensal_pivot{filtro_sql}", binds
    )
    tempo_dest = time.perf_counter() - t0
    print(f"[frente1] int_venda_mensal_pivot lido: {len(linhas_pivot)} linhas em {tempo_dest:.2f}s")

    chave_dest = resolver_chave_dest(colunas_pivot)
    if chave_dest is None:
        return None, (
            f"int_venda_mensal_pivot existe mas nenhuma coluna-chave candidata "
            f"({_CANDIDATOS_CHAVE}) foi encontrada entre as colunas disponiveis "
            f"({sorted(colunas_pivot)}) - nao e' possivel indexar por CODIGO"
        )
    for r in linhas_pivot:
        r["codigo"] = int(r[chave_dest])

    colspecs = resolver_colspecs(colunas_pivot)

    resultado = comparar_relatorio(
        nome="int_venda_mensal_pivot x fVendaMes/.m (reimplementacao Python - DECIDE O VEREDITO)",
        colspecs=colspecs,
        linhas_orig=list(gabarito.values()),
        linhas_dest=linhas_pivot,
        colunas_dest_disponiveis=colunas_pivot,
        tempo_orig_s=tempo_orig,
        tempo_dest_s=tempo_dest,
        tolerancia=tolerancia,
    )
    return resultado, None


# ─────────────────────────────────────────────────────────────────────────────
# FRENTE 2 - informativa, nunca reprova
# ─────────────────────────────────────────────────────────────────────────────


def validar_frente2(conn: oracledb.Connection, xlsx_path: Path, tolerancia: float,
                     limite_codigos: int | None) -> tuple[ResultadoRelatorio | None, str | None]:
    print()
    print(f"[frente2] METODO (informativo, NUNCA reprova): compras.int_venda_mensal_pivot "
          f"(o MODEL de verdade) contra a aba '{XLSX_ABA}' de {xlsx_path.name} - FOTO ANTIGA "
          f"da planilha (WinThor mudou desde entao). Le em streaming (openpyxl read_only=True) "
          f"por causa dos 25 MB do arquivo.")

    if not tabela_existe(conn, "COMPRAS", "int_venda_mensal_pivot"):
        return None, "int_venda_mensal_pivot (model ainda nao existe) - Frente 2 pendente"

    if not xlsx_path.exists():
        return None, f"{xlsx_path} nao encontrado - Frente 2 pendente"

    t0 = time.perf_counter()
    linhas_xlsx, avisos = ler_fvendames_xlsx(xlsx_path)
    tempo_orig = time.perf_counter() - t0
    print(f"[frente2] aba '{XLSX_ABA}' lida: {len(linhas_xlsx)} linhas em {tempo_orig:.2f}s")
    for a in avisos:
        print(f"[frente2] AVISO cabecalho: {a}")

    if limite_codigos:
        codigos_amostra = sorted(r["codigo"] for r in linhas_xlsx)[:limite_codigos]
        conjunto = set(codigos_amostra)
        linhas_xlsx = [r for r in linhas_xlsx if r["codigo"] in conjunto]
        print(f"[frente2] AVISO: --limite-codigos={limite_codigos} ativo - amostra reduzida a "
              f"{len(linhas_xlsx)} linhas do xlsx.")

    codigos_filtro = sorted({r["codigo"] for r in linhas_xlsx}) if limite_codigos else None

    filtro_sql = ""
    binds = {}
    if codigos_filtro:
        cur = conn.cursor()
        cur.execute("select * from compras.int_venda_mensal_pivot where rownum = 0")
        cols_disp_preview = {d[0].lower() for d in cur.description}
        cur.close()
        chave = resolver_chave_dest(cols_disp_preview) or "codigo_produto"
        placeholders = ",".join(f":c{i}" for i in range(len(codigos_filtro)))
        filtro_sql = f" where {chave} in ({placeholders})"
        binds = {f"c{i}": c for i, c in enumerate(codigos_filtro)}

    t0 = time.perf_counter()
    linhas_pivot, colunas_pivot = executar_com_colunas(
        conn, f"select * from compras.int_venda_mensal_pivot{filtro_sql}", binds
    )
    tempo_dest = time.perf_counter() - t0
    print(f"[frente2] int_venda_mensal_pivot lido: {len(linhas_pivot)} linhas em {tempo_dest:.2f}s")

    chave_dest = resolver_chave_dest(colunas_pivot)
    if chave_dest is None:
        return None, f"int_venda_mensal_pivot sem coluna-chave reconhecida - Frente 2 pendente"
    for r in linhas_pivot:
        r["codigo"] = int(r[chave_dest])

    colspecs = resolver_colspecs(colunas_pivot)

    resultado = comparar_relatorio(
        nome="int_venda_mensal_pivot x fVendaMes (xlsx, FOTO ANTIGA - informativo)",
        colspecs=colspecs,
        linhas_orig=linhas_xlsx,
        linhas_dest=linhas_pivot,
        colunas_dest_disponiveis=colunas_pivot,
        tempo_orig_s=tempo_orig,
        tempo_dest_s=tempo_dest,
        tolerancia=tolerancia,
        informativo=True,
        motivo_informativo=(
            "Foto antiga da planilha - o WinThor mudou desde a ultima atualizacao do xlsx. "
            "Divergencia esperada e NAO reprova. O valor aqui e' pegar diferenca SISTEMATICA "
            "(coluna deslocada, offset trocado, sinal invertido) que a Frente 1 nao pegaria "
            "se as duas implementacoes (Python e SQL) errassem do mesmo jeito."
        ),
        nunca_reprova=True,
    )
    return resultado, None


# ─────────────────────────────────────────────────────────────────────────────
# Autoteste sintetico - SEM banco nenhum. Prova que o motor de comparacao
# PEGA DEFEITO (nao so' que ele passa quando tudo esta certo).
# ─────────────────────────────────────────────────────────────────────────────


def _linhas_sinteticas() -> list[dict]:
    """4 meses de historico + mes corrente para o SKU 1001, 2 meses para o
    SKU 2002 - o suficiente pra exercitar offset 0/1/2 (Q00/Q01/Q02, V00/V01,
    TX_DEVOLUCAO_3M) e o caso "poucos meses de historico" (2002)."""
    def d(y, m):
        return datetime.date(y, m, 1)

    rows = [
        # codigo 1001: corrente=ago/26 (offset -1), depois offset 0,1,2
        {"mes": d(2026, 8), "codigo": 1001, "qtd": 10, "valor": 100, "qt_fat": 12, "qt_dev": 2,
         "cli_atacado": 3, "cli_varejo": 5, "dias_sem_estoque": 1, "fora_de_linha": "N"},
        {"mes": d(2026, 7), "codigo": 1001, "qtd": 20, "valor": 200, "qt_fat": 22, "qt_dev": 2,
         "cli_atacado": 4, "cli_varejo": 6, "dias_sem_estoque": 0, "fora_de_linha": "N"},  # offset 0
        {"mes": d(2026, 6), "codigo": 1001, "qtd": 30, "valor": 300, "qt_fat": 32, "qt_dev": 2,
         "cli_atacado": 2, "cli_varejo": 7, "dias_sem_estoque": 2, "fora_de_linha": "N"},  # offset 1
        {"mes": d(2026, 5), "codigo": 1001, "qtd": 40, "valor": 400, "qt_fat": 42, "qt_dev": 2,
         "cli_atacado": 1, "cli_varejo": 8, "dias_sem_estoque": 0, "fora_de_linha": "N"},  # offset 2
        # codigo 2002: so' 2 meses de historico (primeira venda recente)
        {"mes": d(2026, 8), "codigo": 2002, "qtd": 5, "valor": 50, "qt_fat": 5, "qt_dev": 0,
         "cli_atacado": 1, "cli_varejo": 1, "dias_sem_estoque": 0, "fora_de_linha": "S"},
        {"mes": d(2026, 7), "codigo": 2002, "qtd": 6, "valor": 60, "qt_fat": 6, "qt_dev": 0,
         "cli_atacado": 1, "cli_varejo": 1, "dias_sem_estoque": 0, "fora_de_linha": "S"},
    ]
    return rows


def autoteste_sintetico() -> bool:
    print("#" * 88)
    print("AUTOTESTE SINTETICO (sem banco) - prova que o motor de comparacao PEGA DEFEITO")
    print("#" * 88)

    linhas = _linhas_sinteticas()
    gabarito, mes_ref, qtd_fechados = construir_gabarito(linhas)
    print(f"\ngabarito construido: {len(gabarito)} SKUs, mes_ref={rotulo_mes(mes_ref)}, "
          f"meses fechados globais={qtd_fechados}")
    print(f"gabarito[1001] = { {k: _fmt(v) for k, v in gabarito[1001].items()} }")
    print(f"gabarito[2002] = { {k: _fmt(v) for k, v in gabarito[2002].items()} }")

    # Q00(1001) deveria ser 20 (mes 07/26, offset 0) e Q01(1001) deveria ser 30 (mes 06/26,
    # offset 1) - conferido pela conta manual acima.
    assert gabarito[1001]["q00"] == 20.0, gabarito[1001]["q00"]
    assert gabarito[1001]["q01"] == 30.0, gabarito[1001]["q01"]
    # TX_DEVOLUCAO_3M(1001) = (dev0+dev1+dev2)/(fat0+fat1+fat2) = (2+2+2)/(22+32+42) = 6/96
    assert abs(gabarito[1001]["tx_devolucao_3m"] - 6 / 96) < 1e-9
    # MED09/MED12 tem que ser None: dataset sintetico so' tem 3 meses fechados no total
    assert gabarito[1001]["med09"] is None
    assert gabarito[1001]["med12"] is None
    # MED02(1001) = (Q00+Q01)/2 = (20+30)/2 = 25
    assert gabarito[1001]["med02"] == 25.0

    colspecs = resolver_colspecs(set(COLUNAS_LOGICAS_VALOR))  # todas as colunas "existem" no teste

    # ── Caso 1: dest = copia EXATA do gabarito -> tem que PASSAR, zero divergencia ──
    dest_limpo = [dict(rec) for rec in gabarito.values()]
    resultado_limpo = comparar_relatorio(
        nome="[AUTOTESTE] caso limpo (dest = copia exata do gabarito)",
        colspecs=colspecs, linhas_orig=list(gabarito.values()), linhas_dest=dest_limpo,
        colunas_dest_disponiveis=set(COLUNAS_LOGICAS_VALOR), tempo_orig_s=0.0, tempo_dest_s=0.0,
        tolerancia=0.01,
    )
    imprimir_resultado(resultado_limpo)
    caso1_ok = resultado_limpo.ok
    print(f"\n[AUTOTESTE] caso limpo: {'OK (passou, como esperado)' if caso1_ok else 'FALHOU (deveria ter passado!)'}")

    # ── Caso 2: dest com Q00/Q01 TROCADOS no SKU 1001 (simula bug de "offset trocado" ──
    # na traducao pra SQL) - MED02 NAO muda (soma de dois termos trocados e' a mesma
    # soma), o que prova por que o CHECK POR COLUNA individual e' indispensavel: uma
    # comparacao so' de agregado (MED02) deixaria esse bug passar batido.
    dest_quebrado = [dict(rec) for rec in gabarito.values()]
    for rec in dest_quebrado:
        if rec["codigo"] == 1001:
            rec["q00"], rec["q01"] = rec["q01"], rec["q00"]
    resultado_quebrado = comparar_relatorio(
        nome="[AUTOTESTE] caso QUEBRADO (Q00/Q01 trocados no SKU 1001 - offset trocado)",
        colspecs=colspecs, linhas_orig=list(gabarito.values()), linhas_dest=dest_quebrado,
        colunas_dest_disponiveis=set(COLUNAS_LOGICAS_VALOR), tempo_orig_s=0.0, tempo_dest_s=0.0,
        tolerancia=0.01,
    )
    imprimir_resultado(resultado_quebrado)
    q00_result = next(c for c in resultado_quebrado.resultados_coluna if c.coluna_orig == "q00")
    q01_result = next(c for c in resultado_quebrado.resultados_coluna if c.coluna_orig == "q01")
    med02_result = next(c for c in resultado_quebrado.resultados_coluna if c.coluna_orig == "med02")
    caso2_ok = (
        not resultado_quebrado.ok
        and q00_result.divergentes == 1
        and q01_result.divergentes == 1
        and med02_result.divergentes == 0  # prova o ponto: agregado nao pega, coluna pega
    )
    print(f"\n[AUTOTESTE] caso quebrado: "
          f"{'OK (REPROVOU como esperado, e MED02 continuou batendo - prova que so agregado nao pegaria)' if caso2_ok else 'FALHOU (deveria ter reprovado com Q00/Q01 divergentes)'}")

    # ── Caso 3: dest com CODIGO duplicado (fan-out) -> tem que REPROVAR ──
    dest_duplicado = [dict(rec) for rec in gabarito.values()] + [dict(gabarito[1001])]
    resultado_duplicado = comparar_relatorio(
        nome="[AUTOTESTE] caso FAN-OUT (CODIGO 1001 duplicado no lado dbt)",
        colspecs=colspecs, linhas_orig=list(gabarito.values()), linhas_dest=dest_duplicado,
        colunas_dest_disponiveis=set(COLUNAS_LOGICAS_VALOR), tempo_orig_s=0.0, tempo_dest_s=0.0,
        tolerancia=0.01,
    )
    imprimir_resultado(resultado_duplicado)
    caso3_ok = (not resultado_duplicado.ok) and resultado_duplicado.n_chaves_duplicadas_dest == 1
    print(f"\n[AUTOTESTE] caso fan-out: "
          f"{'OK (REPROVOU como esperado)' if caso3_ok else 'FALHOU (deveria ter reprovado por chave duplicada)'}")

    tudo_ok = caso1_ok and caso2_ok and caso3_ok
    print()
    print("#" * 88)
    print(f"AUTOTESTE SINTETICO: {'PASSOU - o motor de comparacao PEGA os 3 defeitos injetados '
                                   'e NAO acusa falso positivo no caso limpo' if tudo_ok else 'FALHOU'}")
    print("#" * 88)
    return tudo_ok


# ─────────────────────────────────────────────────────────────────────────────
# main
# ─────────────────────────────────────────────────────────────────────────────


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--tolerancia", type=float, default=0.01,
                     help="tolerancia absoluta para colunas numericas (default 0.01)")
    ap.add_argument("--somente", choices=["frente1", "frente2", "ambos"], default="ambos",
                     help="qual frente validar (default ambos)")
    ap.add_argument("--limite-codigos", type=int, default=None,
                     help="restringe a comparacao a N CODIGO (amostra determinística) - so' "
                          "para desenvolvimento; o aceite final e' SEM este parametro")
    ap.add_argument("--sem-xlsx", action="store_true", help="pula a Frente 2 (nao le o xlsx)")
    ap.add_argument("--xlsx", default=str(XLSX_PATH_PADRAO), help="caminho do xlsx de referencia")
    ap.add_argument("--lib-dir", default=ORACLE_CLIENT_LIB_DIR, help="Oracle Instant Client (modo thick)")
    ap.add_argument("--profiles-path", default=PROFILES_PATH, help="caminho do profiles.yml")
    ap.add_argument("--autoteste", action="store_true",
                     help="roda so' o autoteste sintetico (sem banco) e sai - prova que o motor "
                          "de comparacao pega defeito injetado de proposito")
    args = ap.parse_args()

    if args.autoteste:
        ok = autoteste_sintetico()
        sys.exit(0 if ok else 1)

    if args.limite_codigos:
        print(f"AVISO: --limite-codigos={args.limite_codigos} ativo - isto e' um modo de "
              f"desenvolvimento. O aceite final exige rodar SEM este parametro.")

    t_inicio = time.perf_counter()
    conn = conectar(args.lib_dir, args.profiles_path)
    print(f"Conectado como COMPRAS em {datetime.datetime.now():%Y-%m-%d %H:%M:%S}.")
    print()
    print("#" * 88)
    print("METODO (ver CONTEXTO.md 6.3 e docstring deste script): Frente 1 decide o veredito -")
    print("reimplementacao Python de shared fVendaMes (.m) x int_venda_mensal_pivot, os dois lados")
    print("materializados no schema compras (sem consulta ao vivo ao CEDEP, sem defasagem por hora).")
    print("Frente 2 e' informativa e NUNCA reprova - int_venda_mensal_pivot x aba fVendaMes do xlsx,")
    print("uma foto antiga da planilha.")
    print("#" * 88)

    resultados: list[ResultadoRelatorio] = []
    pendencias: list[str] = []

    if args.somente in ("frente1", "ambos"):
        r, pendencia = validar_frente1(conn, args.tolerancia, args.limite_codigos)
        if r is not None:
            resultados.append(r)
        if pendencia:
            pendencias.append(pendencia)

    if args.somente in ("frente2", "ambos") and not args.sem_xlsx:
        r, pendencia = validar_frente2(conn, Path(args.xlsx), args.tolerancia, args.limite_codigos)
        if r is not None:
            resultados.append(r)
        if pendencia:
            pendencias.append(pendencia)

    conn.close()

    for r in resultados:
        imprimir_resultado(r)

    tempo_total = time.perf_counter() - t_inicio

    resultados_veredito = [r for r in resultados if not r.nunca_reprova]
    resultados_informativos = [r for r in resultados if r.nunca_reprova]

    print()
    print("#" * 88)
    if pendencias:
        print("PENDENTE:")
        for p in pendencias:
            print(f"  - {p}")
    for r in resultados_veredito:
        print(f"{r.nome}: {_veredito_str(r)}")
    for r in resultados_informativos:
        print(f"{r.nome}: {_veredito_str(r)}")
    print(f"Tempo total: {tempo_total:.2f}s")
    print("#" * 88)

    if pendencias and not resultados_veredito:
        sys.exit(2)
    if any(r.reprova for r in resultados):
        sys.exit(1)
    sys.exit(0)


if __name__ == "__main__":
    main()
